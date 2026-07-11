from datetime import date as date_cls, datetime, time, timedelta
from itertools import combinations

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Count, Prefetch, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from accounts.decorators import role_required
from core.models import Notification, notify
from professors.models import ProfessorAvailability, ProfessorProfile
from students.models import StudentProfile, StudentReference, normalize_matricule

from .forms import (
    DeadlineForm,
    HistoricalDefenseForm,
    JuryForm,
    JuryGenerationForm,
    JuryAddMemberForm,
    JuryMembersForSlotForm,
    JurySmartMembersForm,
    JuryStudentAssignForm,
    PFERequestDecisionForm,
    PFERequestForm,
    PlanningGenerationForm,
    TargetedJuryGenerationForm,
)

from .models import (
    Deadline,
    DefenseSchedule,
    Evaluation,
    Jury,
    JuryMember,
    JuryStudent,
    PFERequest,
    Result,
    FiliereExpert,
    GenerationReport,
    mention_for_average,
    corrected_breakdown,
)

from .pdf import simple_pdf_response


DEFENSE_DURATION_MINUTES = 20

# Salles de soutenance (7). Chaque jury occupe une salle ; deux jurys simultanés
# ne peuvent pas partager la même salle → au plus 7 jurys en parallèle.
DEFENSE_SALLES = ["Amphi", "Salle 1", "Salle 2", "Salle 3", "Salle 4", "Salle 8", "Salle 9"]
MAX_SIMULTANEOUS_JURIES = len(DEFENSE_SALLES)

# Fenêtre des soutenances : du 03/07/2026 au 10/07/2026 inclus (week-end compris).
DEFENSE_START = date_cls(2026, 7, 3)
DEFENSE_DEADLINE = date_cls(2026, 7, 10)

# Créneaux de soutenance (matin / après-midi, exception vendredi) : voir
# professors.slots. Un jury doit tenir entièrement dans un seul créneau.
from professors import slots as defense_slots
from professors.slots import slots_for as defense_slots_for


def _has_real_grades(js):
    """Vrai si l'étudiant a de VRAIES notes : au moins une évaluation ENVOYÉE,
    ou un résultat publié / calculé (moyenne renseignée). Un objet Result vide
    (créé par erreur lors d'une tentative de publication prématurée, sans note)
    ne compte PAS — l'étudiant reste retirable et reprogrammable."""
    if js.evaluations.filter(is_submitted=True).exists():
        return True
    res = getattr(js, "result", None)
    return bool(res and (res.is_published or res.average is not None))


def _slot_label_at(date, start_time):
    """'morning'/'afternoon'/None selon le créneau contenant start_time."""
    touched = defense_slots.slots_touched(
        date, start_time,
        (datetime.combine(date, start_time) + timedelta(minutes=1)).time(),
    )
    if defense_slots.MORNING in touched:
        return defense_slots.MORNING
    if defense_slots.AFTERNOON in touched:
        return defense_slots.AFTERNOON
    return None


def _salle_occupee(defense_date, start_time, end_time, salle):
    """Vrai si la salle est déjà occupée par un jury sur [start_time, end_time]."""
    if not salle:
        return False
    return DefenseSchedule.objects.filter(
        jury_student__jury__defense_date=defense_date,
        jury_student__jury__salle=salle,
        start_time__lt=end_time,
        end_time__gt=start_time,
    ).exists()


def _choisir_salle_libre(defense_date, start_time, end_time, pool=None):
    """Renvoie une salle libre sur le créneau [start_time, end_time], ou None si
    toutes les salles du pool sont occupées à ce moment-là."""
    for salle in (pool or DEFENSE_SALLES):
        if salle and not _salle_occupee(defense_date, start_time, end_time, salle):
            return salle
    return None


@login_required
@role_required(["student"])
def submit_pfe_request(request):
    student = get_object_or_404(
        StudentProfile.objects.select_related("encadrant", "user"),
        user=request.user,
    )

    deadline = Deadline.objects.filter(
        is_active=True
    ).order_by("-deadline_date").first()

    if deadline and deadline.is_closed():
        messages.error(request, "La date limite de dépôt des demandes est dépassée.")
        return redirect("student_dashboard")

    existing_request = PFERequest.objects.filter(student=student).first()

    if existing_request and existing_request.status in [
        PFERequest.STATUS_PENDING_PROFESSOR,
        PFERequest.STATUS_PENDING_ADMIN,
        PFERequest.STATUS_ACCEPTED,
    ]:
        messages.warning(request, "Vous avez déjà une demande en cours ou acceptée.")
        return redirect("student_dashboard")

    if request.method == "POST":
        form = PFERequestForm(
            request.POST,
            request.FILES,
            instance=existing_request,
        )

        if form.is_valid():
            pfe_request = form.save(commit=False)
            pfe_request.student = student
            pfe_request.status = PFERequest.STATUS_PENDING_PROFESSOR

            pfe_request.professor_comment = None
            pfe_request.admin_comment = None
            pfe_request.professor_reviewed_at = None
            pfe_request.admin_reviewed_at = None
            pfe_request.reviewed_by_professor = None
            pfe_request.reviewed_by_admin = None
            pfe_request.reviewed_by = None
            pfe_request.reviewed_at = None

            pfe_request.save()

            messages.success(
                request,
                "Votre demande a été envoyée à votre encadrant."
            )
            return redirect("student_dashboard")
    else:
        form = PFERequestForm(instance=existing_request)

    return render(request, "soutenances/submit_pfe_request.html", {
        "form": form,
        "student": student,
        "deadline": deadline,
        "existing_request": existing_request,
    })


@login_required
@role_required(["admin"])
def admin_pfe_requests(request):
    requests = PFERequest.objects.select_related(
        "student",
        "student__user",
        "student__encadrant",
    ).order_by("-submitted_at")

    # Étudiants ayant soutenu = résultat publié par le chef de département.
    defended_student_ids = set(
        Result.objects.filter(is_published=True).values_list(
            "jury_student__student_id", flat=True
        )
    )

    for demande in requests:
        demande.has_defended = demande.student_id in defended_student_ids

    # Demandes bloquées chez l'encadrant (jamais traitées), étudiant sans jury.
    blocked_count = PFERequest.objects.filter(
        status=PFERequest.STATUS_PENDING_PROFESSOR,
        student__jury_assignment__isnull=True,
    ).count()

    return render(request, "soutenances/admin_pfe_requests.html", {
        "requests": requests,
        "blocked_count": blocked_count,
    })


def _notify_student_decision(pfe_request, accepted):
    if accepted:
        notify(
            getattr(pfe_request.student, "user", None),
            "Demande acceptée",
            "Votre demande de soutenance a été acceptée par le département de l'IUP.",
            "/student-dashboard/",
            category=Notification.CATEGORY_REQUEST,
        )
    else:
        notify(
            getattr(pfe_request.student, "user", None),
            "Demande refusée par le département",
            "Votre demande de soutenance a été refusée par le département de l'IUP.",
            "/student-dashboard/",
            category=Notification.CATEGORY_REQUEST,
        )


@login_required
@role_required(["admin"])
def admin_pfe_quick_accept(request, pk):
    """Accepter une demande (déjà validée par l'encadrant) sans ouvrir le dossier."""
    pfe_request = get_object_or_404(
        PFERequest.objects.select_related("student", "student__user"), pk=pk
    )
    if request.method == "POST":
        if pfe_request.status != PFERequest.STATUS_PENDING_ADMIN:
            messages.error(request, "Cette demande n'est pas en attente du département.")
        else:
            pfe_request.admin_accept(request.user)
            _notify_student_decision(pfe_request, True)
            messages.success(request, f"Demande de {pfe_request.student.full_name} acceptée.")
    return redirect("admin_pfe_requests")


@login_required
@role_required(["admin"])
def admin_pfe_quick_refuse(request, pk):
    """Refuser une demande (en attente du département) sans ouvrir le dossier."""
    pfe_request = get_object_or_404(
        PFERequest.objects.select_related("student", "student__user"), pk=pk
    )
    if request.method == "POST":
        if pfe_request.status != PFERequest.STATUS_PENDING_ADMIN:
            messages.error(request, "Cette demande n'est pas en attente du département.")
        else:
            comment = (request.POST.get("comment") or "").strip()
            pfe_request.admin_refuse(request.user, comment or None)
            _notify_student_decision(pfe_request, False)
            messages.success(request, f"Demande de {pfe_request.student.full_name} refusée.")
    return redirect("admin_pfe_requests")


@login_required
@role_required(["admin"])
@transaction.atomic
def admin_add_historical_defense(request):
    """Enregistre manuellement une soutenance déjà réalisée avant la plateforme
    (étudiant de la liste officielle non inscrit) : crée son profil, son jury,
    sa date et son résultat publié. Il bascule alors en « inscrit » + « soutenu »."""
    from decimal import Decimal
    from django.utils import timezone as _tz
    from accounts.models import CustomUser

    if request.method == "POST":
        form = HistoricalDefenseForm(request.POST)
        if form.is_valid():
            reference = form.cleaned_data.get("reference")
            student = form.cleaned_data.get("student_profile")
            encadrant = form.cleaned_data["encadrant"]
            president = form.cleaned_data["president"]
            member = form.cleaned_data.get("member")
            defense_date = form.cleaned_data["defense_date"]
            salle = form.cleaned_data.get("salle") or ""
            final_note = form.cleaned_data["final_note"]
            matricule = form.cleaned_data["matricule"]

            if student is None:
                # Étudiant non inscrit : créer un compte technique + profil.
                base_username = f"hist_{matricule.lower()}"
                username = base_username
                suffix = 1
                while CustomUser.objects.filter(username=username).exists():
                    suffix += 1
                    username = f"{base_username}_{suffix}"
                user = CustomUser.objects.create(
                    username=username,
                    role=CustomUser.ROLE_STUDENT,
                    is_active=False,
                )
                user.set_unusable_password()
                user.save()

                student = StudentProfile.objects.create(
                    user=user,
                    matricule=matricule,
                    full_name=reference.full_name,
                    filiere=(reference.filiere or ""),
                    encadrant=encadrant,
                )

            # Demande de soutenance (acceptée) : la crée si absente, sinon la
            # marque acceptée (l'étudiant sort de « sans demande »).
            pfe, _created = PFERequest.objects.get_or_create(student=student)
            if pfe.status != PFERequest.STATUS_ACCEPTED:
                pfe.status = PFERequest.STATUS_ACCEPTED
                pfe.save(update_fields=["status"])

            jury = Jury.objects.create(
                name=f"Soutenance {student.full_name} ({defense_date.strftime('%d/%m/%Y')})",
                defense_date=defense_date,
                salle=salle,
                is_validated=True,
            )
            members = [encadrant, president]
            if member and member.id not in (encadrant.id, president.id):
                members.append(member)
            for professor in members:
                JuryMember.objects.create(jury=jury, professor=professor)

            js = JuryStudent.objects.create(
                student=student, jury=jury, president=president
            )

            # Créditer CHAQUE membre : une évaluation « envoyée » (verrouillée)
            # avec la note finale reportée sur les 3 critères (détail non
            # disponible pour l'historique). Ainsi chaque membre est compté
            # comme ayant noté l'étudiant (recap prof, espace prof, stats) et la
            # soutenance apparaît « terminée » dans son espace.
            for professor in members:
                Evaluation.objects.create(
                    jury_student=js,
                    professor=professor,
                    rapport_note=final_note,
                    presentation_note=final_note,
                    questions_note=final_note,
                    is_submitted=True,
                    is_locked=True,
                    submitted_at=_tz.now(),
                )

            Result.objects.create(
                jury_student=js,
                average=final_note,
                note_gap_value=Decimal("0"),
                has_note_gap_alert=False,
                is_published=True,
                published_at=_tz.now(),
            )

            messages.success(
                request,
                f"Soutenance historique enregistrée pour {student.full_name} "
                f"(note {final_note}/20). Il compte désormais parmi les étudiants soutenus."
            )
            return redirect("admin_students_overview")
    else:
        form = HistoricalDefenseForm()

    return render(request, "soutenances/admin_historical_defense.html", {"form": form})


@login_required
@role_required(["admin"])
def admin_regularize_historical(request):
    """Régularise les soutenances historiques enregistrées AVANT le crédit
    automatique des membres : pour chaque résultat PUBLIÉ dont l'étudiant n'a
    AUCUNE évaluation envoyée, crée une évaluation « envoyée » (verrouillée) par
    membre du jury — encadrant compris — avec la note finale reportée sur les 3
    critères. Les notes publiées ne sont PAS modifiées."""
    if request.method != "POST":
        return redirect("admin_students_overview")

    from django.utils import timezone as _tz

    juries_done = 0
    evals_created = 0
    results = (
        Result.objects.filter(is_published=True, average__isnull=False)
        .select_related("jury_student__jury")
        .prefetch_related(
            "jury_student__jury__members__professor",
            "jury_student__evaluations",
        )
    )
    for res in results:
        js = res.jury_student
        # Défense normale déjà notée -> on ne touche pas.
        if js.evaluations.filter(is_submitted=True).exists():
            continue
        members = [m.professor for m in js.jury.members.all()]
        if not members:
            continue
        created_here = 0
        for prof in members:
            if Evaluation.objects.filter(jury_student=js, professor=prof).exists():
                continue
            Evaluation.objects.create(
                jury_student=js,
                professor=prof,
                rapport_note=res.average,
                presentation_note=res.average,
                questions_note=res.average,
                is_submitted=True,
                is_locked=True,
                submitted_at=_tz.now(),
            )
            created_here += 1
        if created_here:
            juries_done += 1
            evals_created += created_here

    if juries_done:
        messages.success(
            request,
            f"{juries_done} soutenance(s) historique(s) régularisée(s) — "
            f"{evals_created} évaluation(s) créée(s). Les membres du jury et les "
            f"encadrants sont désormais crédités."
        )
    else:
        messages.info(request, "Aucune soutenance historique à régulariser.")
    return redirect("admin_students_overview")


@login_required
@role_required(["admin"])
def admin_pfe_accept_all(request):
    """Accepte toutes les demandes déjà validées par les encadrants (en attente
    du département)."""
    if request.method == "POST":
        pending = PFERequest.objects.filter(
            status=PFERequest.STATUS_PENDING_ADMIN
        ).select_related("student", "student__user")
        count = 0
        for pfe_request in pending:
            pfe_request.admin_accept(request.user)
            _notify_student_decision(pfe_request, True)
            count += 1
        messages.success(request, f"{count} demande(s) acceptée(s).")
    return redirect("admin_pfe_requests")


@login_required
@role_required(["admin"])
def admin_pfe_accept_blocked(request):
    """Considère PRÊTS les étudiants dont la demande est restée bloquée chez
    l'encadrant (jamais traitée) : le département court-circuite la validation
    encadrant et accepte leurs demandes en masse. Ils deviennent programmables
    par la génération automatique (ex. fenêtre 08/07 → 10/07)."""
    if request.method == "POST":
        blocked = PFERequest.objects.filter(
            status=PFERequest.STATUS_PENDING_PROFESSOR,
            student__jury_assignment__isnull=True,
        ).select_related("student", "student__user")
        count = 0
        for pfe_request in blocked:
            pfe_request.admin_accept(request.user)
            _notify_student_decision(pfe_request, True)
            count += 1
        messages.success(
            request,
            f"{count} demande(s) bloquée(s) chez l'encadrant acceptée(s) — "
            f"ces étudiants sont maintenant prêts. Lancez « Générer "
            f"automatiquement » avec la fenêtre de dates souhaitée "
            f"(ex. 08/07 → 10/07)."
        )
    return redirect("admin_pfe_requests")


@login_required
@role_required(["admin"])
def admin_pfe_request_detail(request, pk):
    pfe_request = get_object_or_404(
        PFERequest.objects.select_related(
            "student",
            "student__user",
            "student__encadrant",
            "reviewed_by",
            "reviewed_by_admin",
            "reviewed_by_professor",
        ),
        pk=pk,
    )

    decision_form = PFERequestDecisionForm()

    REUPLOAD_FIELDS = {
        "authorization": "authorization_document",
        "attestation": "attestation_stage",
        "rapport": "rapport_stage",
    }

    if request.method == "POST" and request.POST.get("action") == "request_reupload":
        docs = [d for d in request.POST.getlist("reupload_document") if d in REUPLOAD_FIELDS]
        comment = (request.POST.get("reupload_comment") or "").strip()
        if not docs:
            messages.error(request, "Choisissez au moins un document à faire redéposer.")
        else:
            pfe_request.reupload_document = ",".join(docs)
            pfe_request.reupload_comment = comment
            update_fields = ["reupload_document", "reupload_comment"]
            # On vide chaque pièce concernée pour obliger l'étudiant à la redéposer.
            for doc in docs:
                field = REUPLOAD_FIELDS[doc]
                setattr(pfe_request, field, None)
                update_fields.append(field)
            pfe_request.save(update_fields=update_fields)
            notify(
                getattr(pfe_request.student, "user", None),
                "Document(s) à redéposer",
                f"Le département demande le redépôt : {pfe_request.reupload_documents_display}. {comment}".strip(),
                "/student-dashboard/",
                category=Notification.CATEGORY_DOCUMENT,
            )
            messages.success(
                request,
                "Demande de redépôt envoyée à l'étudiant (visible aussi par l'encadrant)."
            )
        return redirect("admin_pfe_request_detail", pk=pfe_request.pk)

    if request.method == "POST":
        decision_form = PFERequestDecisionForm(request.POST)

        if decision_form.is_valid():
            action = request.POST.get("action")
            comment = decision_form.cleaned_data.get("comment")

            if action == "accept":
                if pfe_request.status != PFERequest.STATUS_PENDING_ADMIN:
                    messages.error(
                        request,
                        "Cette demande doit d'abord être validée par l'encadrant."
                    )
                    return redirect("admin_pfe_request_detail", pk=pfe_request.pk)

                pfe_request.admin_accept(request.user)
                notify(
                    getattr(pfe_request.student, "user", None),
                    "Demande acceptée",
                    "Votre demande de soutenance a été acceptée par le département de l'IUP.",
                    "/student-dashboard/",
                    category=Notification.CATEGORY_REQUEST,
                )
                messages.success(request, "La demande a été acceptée avec succès.")
                return redirect("admin_pfe_requests")

            if action == "refuse":
                if pfe_request.status != PFERequest.STATUS_PENDING_ADMIN:
                    messages.error(
                        request,
                        "Cette demande doit d'abord être validée par l'encadrant."
                    )
                    return redirect("admin_pfe_request_detail", pk=pfe_request.pk)

                pfe_request.admin_refuse(request.user, comment)
                notify(
                    getattr(pfe_request.student, "user", None),
                    "Demande refusée par le département",
                    "Votre demande de soutenance a été refusée par le département de l'IUP. Consultez le commentaire.",
                    "/student-dashboard/",
                    category=Notification.CATEGORY_REQUEST,
                )
                messages.success(request, "La demande a été refusée.")
                return redirect("admin_pfe_requests")

            messages.error(request, "Action invalide.")

    return render(request, "soutenances/admin_pfe_request_detail.html", {
        "pfe_request": pfe_request,
        "decision_form": decision_form,
    })


@login_required
@role_required(["admin"])
def admin_deadline(request):
    deadline = Deadline.objects.filter(
        is_active=True
    ).order_by("-deadline_date").first()

    if request.method == "POST":
        form = DeadlineForm(request.POST, instance=deadline)

        if form.is_valid():
            saved = form.save()

            if saved.is_active:
                Deadline.objects.exclude(pk=saved.pk).update(is_active=False)

            messages.success(request, "Date limite mise à jour.")
            return redirect("admin_deadline")
    else:
        form = DeadlineForm(instance=deadline)

    deadlines = Deadline.objects.order_by("-deadline_date")

    return render(request, "soutenances/admin_deadline.html", {
        "form": form,
        "deadlines": deadlines,
    })


@login_required
@role_required(["admin"])
def admin_expert_groups(request):
    """Gestion des groupes d'experts par filière (un expert = professeur de
    référence de la filière, distinct de l'encadrant lors des jurys)."""
    filieres = [c for c in StudentProfile.FILIERE_CHOICES if c[0]]
    professors = list(ProfessorProfile.objects.order_by("full_name"))

    if request.method == "POST":
        for filiere_value, _label in filieres:
            posted_ids = set(request.POST.getlist(f"expert_{filiere_value}"))
            posted_ids = {int(pid) for pid in posted_ids if pid.isdigit()}

            FiliereExpert.objects.filter(filiere=filiere_value).exclude(
                professor_id__in=posted_ids
            ).delete()

            existing = set(
                FiliereExpert.objects.filter(filiere=filiere_value).values_list(
                    "professor_id", flat=True
                )
            )
            for pid in posted_ids - existing:
                FiliereExpert.objects.create(filiere=filiere_value, professor_id=pid)

        messages.success(request, "Groupes d'experts mis à jour.")
        return redirect("admin_expert_groups")

    experts_by_filiere = {value: set() for value, _ in filieres}
    for entry in FiliereExpert.objects.all():
        if entry.filiere in experts_by_filiere:
            experts_by_filiere[entry.filiere].add(entry.professor_id)

    groups = []
    for value, label in filieres:
        selected_ids = experts_by_filiere.get(value, set())
        groups.append({
            "value": value,
            "label": label,
            "count": len(selected_ids),
            "professors": [
                {"id": p.id, "full_name": p.full_name, "checked": p.id in selected_ids}
                for p in professors
            ],
        })

    return render(request, "soutenances/admin_expert_groups.html", {
        "groups": groups,
    })


@login_required
@role_required(["admin"])
def admin_jury_list(request):
    # Les étudiants de chaque jury sont préchargés DANS L'ORDRE des passages
    # (horaire croissant) pour que le template les affiche 9:00, 9:20, …
    students_ordered = (
        JuryStudent.objects
        .select_related("student", "student__encadrant", "president", "schedule")
        .order_by("schedule__start_time", "id")
    )
    juries = Jury.objects.prefetch_related(
        "members__professor",
        Prefetch("students", queryset=students_ordered),
    ).order_by("defense_date", "name")

    # Filtre par jour (optionnel).
    day_filter = (request.GET.get("day") or "").strip()
    selected_day = None
    if day_filter:
        try:
            selected_day = date_cls.fromisoformat(day_filter)
            juries = juries.filter(defense_date=selected_day)
        except ValueError:
            selected_day = None

    pending_students_count = StudentProfile.objects.filter(
        pfe_request__status=PFERequest.STATUS_ACCEPTED,
        jury_assignment__isnull=True,
    ).count()

    future_availabilities_count = ProfessorAvailability.objects.filter(
        date__gte=timezone.localdate()
    ).count()

    draft_juries_count = juries.filter(is_validated=False).count()
    assigned_students_count = JuryStudent.objects.count()

    # Répartition des jurys validés :
    # - Publiés    : date pas encore passée (soutenance à venir)
    # - Terminés   : date passée ET les 3 membres ont noté tous les étudiants
    # - En attente : date passée mais notes pas toutes saisies
    today = timezone.localdate()
    submitted_counts = dict(
        Evaluation.objects.filter(is_submitted=True)
        .values("jury_student_id")
        .annotate(c=Count("id"))
        .values_list("jury_student_id", "c")
    )

    upcoming_juries_count = 0
    completed_juries_count = 0
    awaiting_notes_count = 0
    past_count = 0
    future_count = 0

    juries = list(juries)  # on évalue une fois et on attribue une catégorie
    for jury in juries:
        if not jury.is_validated:
            jury.category = "draft"
            continue
        # Un jury dont TOUS les étudiants sont notés par les 3 membres bascule
        # dans « Passés » (Terminé), même si sa date est encore future.
        jury_students = list(jury.students.all())
        all_graded = bool(jury_students) and all(
            submitted_counts.get(js.id, 0) >= 3 for js in jury_students
        )
        if all_graded:
            jury.category = "past"
            past_count += 1
            completed_juries_count += 1
            continue
        if jury.defense_date and jury.defense_date >= today:
            jury.category = "future"
            upcoming_juries_count += 1
            future_count += 1
            continue
        # Jury validé, date passée, notes incomplètes.
        jury.category = "past"
        past_count += 1
        awaiting_notes_count += 1

    # ── Rapport jurys par jour (matin / après-midi), recalculé à chaque
    #    affichage (donc à jour après ajout/suppression). Basé sur TOUS les
    #    jurys (indépendamment du filtre jour courant).
    from collections import defaultdict
    all_juries = Jury.objects.all()
    first_start = {}
    for row in (
        DefenseSchedule.objects
        .filter(jury_student__jury__in=all_juries)
        .order_by("start_time")
        .values("jury_student__jury_id", "start_time")
    ):
        jid = row["jury_student__jury_id"]
        if jid not in first_start:
            first_start[jid] = row["start_time"]
    day_map = defaultdict(lambda: {"morning": 0, "afternoon": 0})
    available_days = set()
    for j in all_juries:
        if not j.defense_date:
            continue
        available_days.add(j.defense_date)
        st = first_start.get(j.pk)
        # Matin / après-midi robuste : basé sur l'heure de début (>= 14:00 =
        # après-midi). Reste correct même hors des bornes strictes du créneau
        # (ex. vendredi après-midi commençant à 15:20, dans le « trou » 12h–16h).
        if st and st >= time(14, 0):
            day_map[j.defense_date]["afternoon"] += 1
        else:
            day_map[j.defense_date]["morning"] += 1
    day_report = [
        {
            "date": d, "morning": day_map[d]["morning"],
            "afternoon": day_map[d]["afternoon"],
            "total": day_map[d]["morning"] + day_map[d]["afternoon"],
        }
        for d in sorted(day_map.keys())
    ]

    # Ordonner les jurys par date puis par HORAIRE de début (ordre croissant) —
    # et non par nom (qui trierait « Jury 10 » avant « Jury 2 »).
    juries.sort(key=lambda j: (
        j.defense_date or date_cls.max,
        first_start.get(j.pk) or time(23, 59),
    ))

    return render(request, "soutenances/admin_jury_list.html", {
        "juries": juries,
        "generation_form": JuryGenerationForm(),
        "pending_students_count": pending_students_count,
        "future_availabilities_count": future_availabilities_count,
        "duration_minutes": DEFENSE_DURATION_MINUTES,
        "total_juries_count": len(juries),
        "published_juries_count": upcoming_juries_count,
        "completed_juries_count": completed_juries_count,
        "awaiting_notes_count": awaiting_notes_count,
        "draft_juries_count": draft_juries_count,
        "past_count": past_count,
        "future_count": future_count,
        "assigned_students_count": assigned_students_count,
        "day_report": day_report,
        "available_days": sorted(available_days),
        "selected_day": selected_day,
    })


def _next_week_monday_friday():
    today = timezone.localdate()
    days_ahead = (0 - today.weekday()) % 7 or 7  # lundi prochain (semaine suivante)
    monday = today + timedelta(days=days_ahead)
    return monday, monday + timedelta(days=4)


def _parse_by_filiere_params(request):
    monday, friday = _next_week_monday_friday()
    try:
        sd = date_cls.fromisoformat((request.POST.get("start_date") or "").strip())
    except ValueError:
        sd = monday
    try:
        ed = date_cls.fromisoformat((request.POST.get("end_date") or "").strip())
    except ValueError:
        ed = friday
    try:
        cap = int(request.POST.get("max_simultaneous") or len(DEFENSE_SALLES))
    except ValueError:
        cap = len(DEFENSE_SALLES)
    cap = max(1, min(cap, len(DEFENSE_SALLES)))
    return sd, ed, cap


@login_required
@role_required(["admin"])
def admin_generate_by_filiere(request):
    """Teste l'algorithme ALTERNATIF par filière. GET : formulaire (lundi→
    vendredi prochain pré-rempli). POST preview : aperçu NON destructif.
    L'application se fait via admin_apply_by_filiere."""
    monday, friday = _next_week_monday_friday()
    if request.method == "POST":
        sd, ed, cap = _parse_by_filiere_params(request)
        if ed < sd:
            messages.error(request, "La date de fin doit être postérieure ou égale au début.")
            return redirect("admin_generate_by_filiere")
        payload = run_by_filiere(sd, ed, cap, commit=False)
        # On SAUVEGARDE l'aperçu pour pouvoir le reconsulter (le bouton
        # « Dernier test par filière »), même après avoir changé d'onglet.
        payload["kind"] = "by_filiere"
        payload["start_date"] = sd.isoformat()
        payload["end_date"] = ed.isoformat()
        payload["max_simultaneous"] = cap
        GenerationReport.objects.create(data=payload)
        old_ids = list(
            GenerationReport.objects.filter(data__kind="by_filiere")
            .order_by("-created_at").values_list("id", flat=True)[5:]
        )
        if old_ids:
            GenerationReport.objects.filter(id__in=old_ids).delete()
        return render(request, "soutenances/admin_by_filiere_preview.html", {
            "report": payload,
            "start_date": sd.isoformat(),
            "end_date": ed.isoformat(),
            "max_simultaneous": cap,
        })
    return render(request, "soutenances/admin_by_filiere_form.html", {
        "start_date": monday.isoformat(),
        "end_date": friday.isoformat(),
        "max_simultaneous": len(DEFENSE_SALLES),
        "salles": DEFENSE_SALLES,
    })


@login_required
@role_required(["admin"])
def admin_by_filiere_last_preview(request):
    """Rouvre le dernier aperçu « par filière » sauvegardé (pour le consulter
    à nouveau et éventuellement l'appliquer)."""
    entry = (
        GenerationReport.objects.filter(data__kind="by_filiere")
        .order_by("-created_at").first()
    )
    if not entry:
        messages.info(
            request,
            "Aucun test par filière enregistré. Lancez d'abord un aperçu."
        )
        return redirect("admin_generate_by_filiere")
    data = entry.data
    return render(request, "soutenances/admin_by_filiere_preview.html", {
        "report": data,
        "start_date": data.get("start_date", ""),
        "end_date": data.get("end_date", ""),
        "max_simultaneous": data.get("max_simultaneous", len(DEFENSE_SALLES)),
        "saved_preview": True,
        "saved_at": data.get("generated_at", ""),
    })


@login_required
@role_required(["admin"])
def admin_apply_by_filiere(request):
    """Applique l'algorithme par filière : REMPLACE les jurys publiés de la
    fenêtre (hors jurys contenant des étudiants déjà notés) par la nouvelle
    génération, qui devient publiée. Action irréversible (confirmée)."""
    if request.method != "POST":
        return redirect("admin_generate_by_filiere")
    sd, ed, cap = _parse_by_filiere_params(request)
    if ed < sd:
        messages.error(request, "Dates invalides.")
        return redirect("admin_generate_by_filiere")
    payload = run_by_filiere(sd, ed, cap, commit=True)
    GenerationReport.objects.create(data=payload)
    messages.success(
        request,
        f"Génération par filière appliquée : {payload.get('replaced_juries', 0)} "
        f"jury(s) publié(s) remplacé(s), {payload.get('created', 0)} nouveau(x) "
        f"jury(s), {payload.get('assigned', 0)} étudiant(s) programmé(s)."
    )
    if payload.get("skipped_graded"):
        messages.info(
            request,
            f"{payload['skipped_graded']} jury(s) conservé(s) car ils contiennent "
            f"des étudiants déjà notés."
        )
    return redirect("admin_jury_list")


@login_required
@role_required(["admin"])
def admin_generate_juries(request):
    if request.method == "POST":
        form = JuryGenerationForm(request.POST)
        if not form.is_valid():
            return render(request, "soutenances/admin_generate_auto.html", {
                "form": form,
                "salles": DEFENSE_SALLES,
            })

        result = generate_smart_juries(
            start_date=form.cleaned_data["start_date"],
            end_date=form.cleaned_data["end_date"],
            max_simultaneous=form.cleaned_data["max_simultaneous"],
        )
        # Rapport sérialisé : affiché maintenant ET sauvegardé pour pouvoir le
        # reconsulter depuis la page Jurys (bouton « Dernier rapport »).
        payload = build_generation_report_payload(result)
        GenerationReport.objects.create(data=payload)
        old_ids = list(
            GenerationReport.objects.order_by("-created_at")
            .values_list("id", flat=True)[10:]
        )
        if old_ids:
            GenerationReport.objects.filter(id__in=old_ids).delete()
        return render(request, "soutenances/admin_generation_report.html", {
            "report": payload,
        })

    # GET : afficher le formulaire de paramètres (dates + nombre max de jurys).
    form = JuryGenerationForm(initial={
        "start_date": DEFENSE_START,
        "end_date": DEFENSE_DEADLINE,
        "max_simultaneous": len(DEFENSE_SALLES),
    })
    return render(request, "soutenances/admin_generate_auto.html", {
        "form": form,
        "salles": DEFENSE_SALLES,
    })


@login_required
@role_required(["admin"])
def admin_last_generation_report(request):
    """Reconsulter le dernier rapport de génération automatique (hors tests
    par filière, qui ont leur propre bouton)."""
    entry = None
    for r in GenerationReport.objects.order_by("-created_at")[:30]:
        if (r.data or {}).get("kind") != "by_filiere":
            entry = r
            break
    if not entry:
        messages.info(
            request,
            "Aucun rapport de génération enregistré pour le moment. "
            "Lancez une génération automatique pour en créer un."
        )
        return redirect("admin_jury_list")
    return render(request, "soutenances/admin_generation_report.html", {
        "report": entry.data,
    })


def build_generation_report_payload(result):
    """Convertit le résultat de génération en dictionnaire sérialisable
    (JSON) consommé par le template du rapport et stocké en base."""

    def fmt_d(d):
        return d.strftime("%d/%m/%Y") if hasattr(d, "strftime") else str(d)

    def fmt_t(t):
        return t.strftime("%H:%M") if hasattr(t, "strftime") else str(t)

    report = result.get("report", {})

    juries = []
    for e in report.get("juries", []):
        juries.append({
            "name": e.get("jury_name", ""),
            "salle": e.get("salle", ""),
            "members": e.get("members", []),
            "filiere": e.get("filiere", ""),
            "date": fmt_d(e.get("defense_date")),
            "start": fmt_t(e.get("slot_start")),
            "capacity": e.get("capacity", 0),
            "students": [
                {
                    "name": s.get("name", ""),
                    "matricule": s.get("matricule", ""),
                    "encadrant": s.get("encadrant", ""),
                    "time": f"{fmt_t(s.get('start_time'))} → {fmt_t(s.get('end_time'))}",
                }
                for s in e.get("students_scheduled", [])
            ],
        })

    # Bilan par jour : nombre de jurys et d'étudiants par date.
    day_map = {}
    for e in juries:
        entry = day_map.setdefault(
            e["date"], {"date": e["date"], "juries": 0, "students": 0}
        )
        entry["juries"] += 1
        entry["students"] += len(e["students"])
    by_day = list(day_map.values())

    errors = []
    for err in result.get("errors", []):
        student = err.get("student")
        errors.append({
            "student": (getattr(student, "full_name", "") or "(nom absent)"),
            "matricule": getattr(student, "matricule", ""),
            "encadrant": (
                student.encadrant.full_name
                if getattr(student, "encadrant", None) else "—"
            ),
            "message": err.get("message", ""),
        })

    return {
        "generated_at": timezone.localtime().strftime("%d/%m/%Y %H:%M"),
        "total_ready": report.get("total_ready", 0),
        "created": result.get("created", 0),
        "assigned": result.get("assigned", 0),
        "filled_existing": report.get("filled_existing", 0),
        "replaced_juries": report.get("replaced_juries", 0),
        "skipped_graded": report.get("skipped_graded", 0),
        "filled_details": [
            {
                "name": f.get("name", ""),
                "matricule": f.get("matricule", ""),
                "encadrant": f.get("encadrant", ""),
                "jury_name": f.get("jury_name", ""),
                "date": fmt_d(f.get("defense_date")),
                "time": f"{fmt_t(f.get('start_time'))} → {fmt_t(f.get('end_time'))}",
                "salle": f.get("salle", ""),
                "is_validated": f.get("is_validated", False),
            }
            for f in report.get("filled_details", [])
        ],
        "errors_count": len(errors),
        "feasibility": report.get("feasibility", {}),
        "priority_usage": report.get("priority_usage", []),
        "by_day": by_day,
        "by_encadrant": report.get("by_encadrant", []),
        "juries": juries,
        "errors": errors,
    }


def _targeted_students_grouped():
    """Étudiants acceptés sans jury, groupés par encadrant (pour la sélection
    par encadrant dans la génération ciblée)."""
    from collections import OrderedDict

    students = StudentProfile.objects.filter(
        pfe_request__status=PFERequest.STATUS_ACCEPTED,
        jury_assignment__isnull=True,
        encadrant__isnull=False,
    ).select_related("encadrant").order_by("encadrant__full_name", "full_name")

    groups = OrderedDict()
    for student in students:
        key = student.encadrant.full_name
        groups.setdefault(key, []).append(student)
    return [{"encadrant": name, "students": rows} for name, rows in groups.items()]


@login_required
@role_required(["admin"])
def admin_generate_juries_targeted(request):
    """Génération ciblée : dates + nombre de jurys + étudiants + professeurs
    + salles choisis par l'admin."""
    if request.method == "POST":
        form = TargetedJuryGenerationForm(request.POST)
        # Dates multiples et salles lues directement (widgets dynamiques).
        defense_dates = []
        for raw in request.POST.getlist("defense_dates"):
            raw = (raw or "").strip()
            if not raw:
                continue
            try:
                defense_dates.append(date_cls.fromisoformat(raw))
            except ValueError:
                pass
        valid_salles = {choice for choice, _ in Jury.SALLE_CHOICES if choice}
        salles = [s for s in request.POST.getlist("salles") if s in valid_salles]

        if form.is_valid():
            if not defense_dates:
                messages.error(request, "Choisissez au moins une date de soutenance.")
                return render(request, "soutenances/admin_generate_targeted.html", {
                    "form": form, "salle_choices": Jury.SALLE_CHOICES,
                    "students_grouped": _targeted_students_grouped(),
                })

            result = generate_targeted_juries(
                defense_dates,
                list(form.cleaned_data["students"]),
                list(form.cleaned_data["professors"]),
                form.cleaned_data["num_juries"],
                salles,
            )
            if result.get("error"):
                messages.error(request, result["error"])
                return render(request, "soutenances/admin_generate_targeted.html", {
                    "form": form, "salle_choices": Jury.SALLE_CHOICES,
                    "students_grouped": _targeted_students_grouped(),
                })

            dates_label = ", ".join(d.strftime("%d/%m/%Y") for d in sorted(set(defense_dates)))
            messages.success(
                request,
                f"{result['created']} jury(s) créé(s), "
                f"{result['assigned']} étudiant(s) programmé(s) sur : {dates_label}."
            )
            if result["skipped"]:
                detail = "; ".join(
                    f"{s.full_name} ({why})" for s, why in result["skipped"][:8]
                )
                messages.warning(
                    request,
                    f"{len(result['skipped'])} étudiant(s) non programmé(s) : {detail}"
                )
            return redirect("admin_jury_list")
    else:
        form = TargetedJuryGenerationForm()

    return render(request, "soutenances/admin_generate_targeted.html", {
        "form": form, "salle_choices": Jury.SALLE_CHOICES,
        "students_grouped": _targeted_students_grouped(),
    })


@transaction.atomic
def generate_targeted_juries(defense_dates, students, professors, num_juries, salles=None):
    """Génère les jurys ciblés sur une ou plusieurs dates. Pour chaque date, on
    place jusqu'à num_juries jurys parmi les étudiants encore non programmés ;
    les étudiants restants passent à la date suivante. Respecte la date limite,
    les disponibilités, la priorité mono-filière + expert (≠ encadrant), et
    répartit les salles fournies en rotation."""
    result = {
        "created": 0, "assigned": 0, "scheduled": 0,
        "skipped": [], "juries": [], "error": None,
    }

    dates = sorted({d for d in (defense_dates or []) if d})
    if not dates:
        result["error"] = "Choisissez au moins une date de soutenance."
        return result
    if any(d < DEFENSE_START or d > DEFENSE_DEADLINE for d in dates):
        result["error"] = (
            f"Les soutenances doivent se tenir entre le "
            f"{DEFENSE_START.strftime('%d/%m/%Y')} et le "
            f"{DEFENSE_DEADLINE.strftime('%d/%m/%Y')} (inclus)."
        )
        return result

    prof_ids_all = {p.id for p in professors}
    not_selected = [s for s in students if s.encadrant_id not in prof_ids_all]
    for student in not_selected:
        result["skipped"].append((student, "encadrant non sélectionné"))

    remaining = [s for s in students if s.encadrant_id in prof_ids_all]
    placed_ids = set()
    salle_index = 0

    for defense_date in dates:
        sub = [s for s in remaining if s.id not in placed_ids]
        if not sub:
            break
        day = _generate_juries_one_date(
            defense_date, sub, professors, num_juries, salles, salle_index
        )
        result["created"] += day["created"]
        result["assigned"] += day["assigned"]
        result["scheduled"] += day["scheduled"]
        result["juries"].extend(day["juries"])
        placed_ids.update(day["placed_ids"])
        salle_index = day["salle_index"]

    for student in remaining:
        if student.id not in placed_ids:
            result["skipped"].append(
                (student, "aucune date/disponibilité ne permet de le programmer")
            )

    return result


def _generate_juries_one_date(defense_date, students, professors, num_juries, salles, salle_index):
    """Place jusqu'à num_juries jurys sur UNE date. Retourne un dict avec
    created/assigned/scheduled/juries, placed_ids (étudiants programmés) et
    salle_index (position de rotation des salles atteinte)."""
    from collections import defaultdict

    result = {
        "created": 0, "assigned": 0, "scheduled": 0,
        "juries": [], "skipped": [], "placed_ids": [], "salle_index": salle_index,
    }
    salles = salles or []

    prof_by_id = {p.id: p for p in professors}
    prof_ids = set(prof_by_id)

    # Experts par filière, restreints aux professeurs sélectionnés.
    experts_by_filiere = defaultdict(set)
    for entry in FiliereExpert.objects.filter(professor_id__in=prof_ids):
        experts_by_filiere[entry.filiere].add(entry.professor_id)
    all_expert_ids = set().union(*experts_by_filiere.values()) if experts_by_filiere else set()

    # 1. Étudiants exploitables (acceptés + encadrant sélectionné).
    ready = []
    for student in students:
        if student.encadrant_id not in prof_ids:
            result["skipped"].append((student, "encadrant non sélectionné"))
        else:
            ready.append(student)

    if not ready:
        return result

    # 2. Regroupement par encadrant (en conservant la filière).
    by_enc = defaultdict(list)
    for student in ready:
        by_enc[student.encadrant_id].append(student)

    enc_filiere = {}
    for eid, group in by_enc.items():
        enc_filiere[eid] = group[0].filiere or ""

    # 3. Répartition des encadrants dans num_juries bandes, priorité mono-filière
    #    et au plus 2 encadrants par jury (1 place réservée à l'expert).
    n = max(1, num_juries)
    buckets = [
        {"filiere": None, "enc_ids": [], "students": []} for _ in range(n)
    ]
    enc_sorted = sorted(by_enc.keys(), key=lambda e: (enc_filiere[e], -len(by_enc[e])))

    for eid in enc_sorted:
        fil = enc_filiere[eid]
        target = None
        # a) privilégier une bande vide (maximise le parallélisme et garde chaque
        #    jury mono-filière tant qu'il reste des bandes libres)
        empty = [b for b in buckets if not b["enc_ids"]]
        if empty:
            target = empty[0]
            target["filiere"] = fil
        else:
            # b) sinon, bande de même filière avec une place d'encadrant libre
            same = [b for b in buckets if b["filiere"] == fil and len(b["enc_ids"]) < 2]
            if same:
                target = min(same, key=lambda b: len(b["students"]))
            else:
                # c) repli : bande avec place d'encadrant libre (filières mêlées)
                room = [b for b in buckets if len(b["enc_ids"]) < 2]
                if room:
                    target = min(room, key=lambda b: len(b["students"]))
        if target is None:
            for student in by_enc[eid]:
                result["skipped"].append((student, "pas assez de jurys disponibles"))
            continue
        target["enc_ids"].append(eid)
        target["students"].extend(by_enc[eid])

    # Les encadrants ne servent que dans leur propre jury : on les exclut des
    # pools d'experts / remplissage pour ne pas les rendre indisponibles ailleurs.
    all_enc_ids = set(by_enc.keys())
    used_prof_ids = set()  # professeurs déjà engagés (jurys parallèles disjoints)

    # Heures de début candidates : grille de 20 min à l'intérieur de chaque
    # créneau (matin 9h–14h, après-midi 15h–19h). Un jury ne peut pas démarrer
    # dans la coupure 14h–15h.
    day_slots = defense_slots_for(defense_date)
    candidate_starts = []
    for slot_start, slot_end in day_slots:
        cursor_t = datetime.combine(defense_date, slot_start)
        slot_end_dt = datetime.combine(defense_date, slot_end)
        while cursor_t < slot_end_dt:
            candidate_starts.append(cursor_t.time())
            cursor_t += timedelta(minutes=DEFENSE_DURATION_MINUTES)

    def block_fits_slot(start, block_minutes):
        """Le bloc [start, start+durée] doit tenir entièrement dans un créneau."""
        start_dt = datetime.combine(defense_date, start)
        end_dt = start_dt + timedelta(minutes=block_minutes)
        for slot_start, slot_end in day_slots:
            if start >= slot_start and end_dt <= datetime.combine(defense_date, slot_end):
                return True
        return False

    def earliest_start_for(encadrant_objs, n_students):
        """Renvoie (heure_début, encadrants_disponibles) : d'abord un créneau où
        TOUS les encadrants sont disponibles pour tout le bloc (qui doit tenir
        dans un seul créneau matin/après-midi), sinon le créneau qui en rend le
        plus disponibles."""
        block = DEFENSE_DURATION_MINUTES * n_students
        for start in candidate_starts:
            if not block_fits_slot(start, block):
                continue
            if all(
                is_professor_available(p, defense_date, start, block)
                for p in encadrant_objs
            ):
                return start, list(encadrant_objs)
        best = None
        for start in candidate_starts:
            if not block_fits_slot(start, block):
                continue
            avail = [
                p for p in encadrant_objs
                if is_professor_available(p, defense_date, start, block)
            ]
            if avail and (best is None or len(avail) > len(best[1])):
                best = (start, avail)
        return best if best else (None, [])

    for bucket in buckets:
        if not bucket["students"]:
            continue

        enc_ids = list(dict.fromkeys(bucket["enc_ids"]))
        enc_objs = [prof_by_id[e] for e in enc_ids]

        # 3a. Choix du créneau de début selon la disponibilité des encadrants.
        start_time, avail_enc = earliest_start_for(enc_objs, len(bucket["students"]))
        if start_time is None:
            for student in bucket["students"]:
                result["skipped"].append(
                    (student, "aucune disponibilité commune des encadrants ce jour")
                )
            continue

        kept_enc = [p.id for p in avail_enc]
        for eid in enc_ids:
            if eid not in kept_enc:
                p = prof_by_id[eid]
                for student in [s for s in bucket["students"] if s.encadrant_id == eid]:
                    result["skipped"].append(
                        (student, f"encadrant indisponible ({p.full_name})")
                    )

        bucket_students = [s for s in bucket["students"] if s.encadrant_id in kept_enc]
        if not bucket_students:
            continue

        n_students = len(bucket_students)
        block_minutes = DEFENSE_DURATION_MINUTES * n_students
        bucket_filieres = {s.filiere or "" for s in bucket_students}

        members = [prof_by_id[e] for e in kept_enc]
        members_ids = {p.id for p in members}

        def can_use(pid):
            if pid in members_ids or pid in used_prof_ids or pid in all_enc_ids:
                return False
            return is_professor_available(
                prof_by_id[pid], defense_date, start_time, block_minutes
            )

        # 3b. Ajouter un expert de la filière (≠ encadrants), disponible.
        has_expert = False
        for fil in bucket_filieres:
            if len(members) >= 3:
                break
            candidates = [
                pid for pid in experts_by_filiere.get(fil, set())
                if pid not in kept_enc and can_use(pid)
            ]
            if candidates:
                pid = candidates[0]
                members.append(prof_by_id[pid])
                members_ids.add(pid)
                has_expert = True

        # 3c. Compléter à 3 avec des professeurs disponibles. On privilégie les
        #     non-experts pour réserver les experts à leur propre filière.
        non_experts = [pid for pid in prof_ids if pid not in all_expert_ids]
        expert_fillers = [pid for pid in prof_ids if pid in all_expert_ids]
        for pid in non_experts + expert_fillers:
            if len(members) >= 3:
                break
            if can_use(pid):
                members.append(prof_by_id[pid])
                members_ids.add(pid)

        if len(members) < 3:
            for student in bucket_students:
                result["skipped"].append(
                    (student, "pas assez de professeurs disponibles pour compléter le jury")
                )
            continue

        # 4. Salle : une salle libre pour tout le bloc du jury (deux jurys
        #    simultanés ne partagent pas de salle). Sans salle libre → on n'ouvre
        #    pas ce jury (les étudiants passeront à une autre date).
        block_end = (
            datetime.combine(defense_date, start_time)
            + timedelta(minutes=block_minutes)
        ).time()
        salle = _choisir_salle_libre(
            defense_date, start_time, block_end, pool=(salles or DEFENSE_SALLES)
        )
        if salle is None:
            for student in bucket_students:
                result["skipped"].append(
                    (student, "aucune salle disponible à ce créneau")
                )
            continue

        idx = result["created"] + 1
        jury = Jury.objects.create(
            name=f"Jury {defense_date.strftime('%d/%m')} #{idx}",
            defense_date=defense_date,
            salle=salle,
            is_validated=False,
        )
        for professor in members:
            JuryMember.objects.create(jury=jury, professor=professor)
        used_prof_ids.update(members_ids)
        result["created"] += 1

        cursor = datetime.combine(defense_date, start_time)
        scheduled_here = 0
        for student in sorted(bucket_students, key=lambda s: s.full_name.lower()):
            # Président : un expert de la filière de préférence, sinon tout
            # membre ≠ encadrant de l'étudiant.
            student_experts = experts_by_filiere.get(student.filiere or "", set())
            president = next(
                (m for m in members
                 if m.id != student.encadrant_id and m.id in student_experts),
                None,
            ) or next(
                (m for m in members if m.id != student.encadrant_id), None
            )
            js = JuryStudent.objects.create(
                student=student, jury=jury, president=president
            )
            try:
                DefenseSchedule.objects.create(
                    jury_student=js,
                    start_time=cursor.time(),
                    end_time=(cursor + timedelta(minutes=DEFENSE_DURATION_MINUTES)).time(),
                    duration_minutes=DEFENSE_DURATION_MINUTES,
                )
            except ValidationError as exc:
                js.delete()
                result["skipped"].append((student, "; ".join(exc.messages)))
                continue
            result["assigned"] += 1
            result["placed_ids"].append(student.id)
            scheduled_here += 1
            cursor += timedelta(minutes=DEFENSE_DURATION_MINUTES)

        result["scheduled"] += scheduled_here
        result["juries"].append({
            "name": jury.name,
            "salle": jury.get_salle_display() if jury.salle else "",
            "date": defense_date.strftime("%d/%m/%Y"),
            "members": [p.full_name for p in members],
            "count": scheduled_here,
            "has_expert": has_expert,
            "mono_filiere": len(bucket_filieres) == 1,
        })

    return result


@transaction.atomic
def generate_smart_juries(start_date=None, end_date=None, max_simultaneous=None):
    """
    Planificateur des jurys de soutenance.

    Hiérarchie d'objectifs :
      1. (Dur)   Placer TOUS les étudiants prêts dans la fenêtre de soutenances.
      2. (Dur)   Un membre siège soit le matin, soit l'après-midi d'une même
                 journée — jamais les deux.
      3. (Dur)   Encadrant présent (sauf aucune dispo → jury avec expert de la
                 filière), 3 membres, président ≠ encadrant, une salle par jury,
                 20 minutes par étudiant, capacité max de jurys simultanés.
      4. (Max)   Un prof prioritaire par jury (un seul), qui préside ; rotation
                 entre prioritaires pour consommer leurs disponibilités.
      5. (Bonus) Expert de la filière présent ; jury mono-filière ; président
                 stable le plus longtemps possible.

    Stratégie :
      - PRÉ-VOL de faisabilité : offre (demi-journées déclarées, règle 2
        comprise) vs besoin (étudiants x 20 min) par encadrant + plafond global.
      - Placement « LE PLUS CONTRAINT D'ABORD » : les encadrants avec le moins
        de marge sont placés en premier, sur leurs demi-journées rares.
      - CO-PLACEMENT : le 3e siège va de préférence à un autre encadrant ayant
        des étudiants restants (ils passent dans le même jury).
      - Boucle jusqu'à point fixe, puis 2e essai en relâchant les préférences
        (jamais les règles dures), puis passe « encadrant absent ».
    """
    from collections import defaultdict

    if start_date is None:
        start_date = DEFENSE_START
    if end_date is None:
        end_date = DEFENSE_DEADLINE
    cap = max_simultaneous or MAX_SIMULTANEOUS_JURIES
    cap = max(1, min(cap, len(DEFENSE_SALLES)))

    today = timezone.localdate()
    window_start = max(today, start_date)

    # 1. Étudiants prêts (PFE accepté, sans jury, encadrant connu)
    all_ready = list(
        StudentProfile.objects.filter(
            pfe_request__status=PFERequest.STATUS_ACCEPTED,
            jury_assignment__isnull=True,
            encadrant__isnull=False,
        ).select_related("encadrant", "user")
    )

    professors = list(ProfessorProfile.objects.order_by("full_name"))
    prof_by_id = {p.id: p for p in professors}

    experts_by_filiere = defaultdict(set)
    for entry in FiliereExpert.objects.all():
        experts_by_filiere[entry.filiere].add(entry.professor_id)

    result = {
        "created": 0,
        "assigned": 0,
        "scheduled": 0,
        "errors": [],
        "report": {
            "total_ready": len(all_ready),
            "by_encadrant_before": {},
            "juries": [],
            "filled_details": [],
            "feasibility": {"encadrants": [], "global": {}},
            "coverage": {"with_priority": 0, "with_expert": 0},
            "priority_usage": [],
        },
    }

    if not all_ready:
        return result

    if len(professors) < 3:
        for student in all_ready:
            result["errors"].append({
                "student": student,
                "reason": "not_enough_professors",
                "message": "Nombre insuffisant de professeurs pour former un jury.",
            })
        return result

    # 2. Étudiants groupés par encadrant
    students_by_encadrant = defaultdict(list)
    for student in all_ready:
        students_by_encadrant[student.encadrant_id].append(student)
    for key in students_by_encadrant:
        students_by_encadrant[key].sort(key=lambda s: s.full_name.lower())

    for enc_id, students in students_by_encadrant.items():
        if students:
            enc_name = students[0].encadrant.full_name
            result["report"]["by_encadrant_before"][enc_name] = len(students)

    # Comptes initiaux par encadrant (pour le bilan prêts/affectés/restants).
    initial_by_enc = {
        eid: len(sts) for eid, sts in students_by_encadrant.items()
    }

    # 2bis. PHASE DE REMPLISSAGE : avant de créer de NOUVEAUX jurys, on place
    #       d'abord les étudiants sans jury dans les jurys EXISTANTS (brouillons
    #       ET publiés) où leur encadrant est membre et où il reste un créneau
    #       libre. On ne retire/déplace JAMAIS un étudiant déjà présent.
    today = timezone.localdate()
    filled = 0
    for enc_id in list(students_by_encadrant.keys()):
        students = students_by_encadrant[enc_id]
        while students:
            placed = False
            candidate_juries = (
                Jury.objects.filter(
                    members__professor_id=enc_id,
                    defense_date__gte=today,
                ).order_by("defense_date", "id").distinct()
            )
            for jury in candidate_juries:
                jmembers = [m.professor for m in jury.members.select_related("professor")]
                slot = find_free_slot_in_jury(jury, jmembers)
                if slot is None:
                    continue
                student = students[0]
                president = choose_president_for_student(
                    student=student, members=jmembers, defense_date=jury.defense_date,
                )
                try:
                    with transaction.atomic():
                        js = JuryStudent.objects.create(
                            jury=jury, student=student, president=president,
                            encadrant_absent=False,
                        )
                        DefenseSchedule.objects.create(
                            jury_student=js, start_time=slot,
                            duration_minutes=DEFENSE_DURATION_MINUTES,
                        )
                        refresh_jury_name_count(jury)
                except ValidationError:
                    continue
                if jury.is_validated:
                    notify(
                        getattr(student, "user", None),
                        "Soutenance planifiée",
                        f"Votre soutenance est prévue le "
                        f"{jury.defense_date.strftime('%d/%m/%Y')} à "
                        f"{slot.strftime('%H:%M')} (jury « {jury.name} », salle "
                        f"{jury.get_salle_display() or '—'}).",
                        "/student-dashboard/",
                        category=Notification.CATEGORY_JURY,
                    )
                result["report"]["filled_details"].append({
                    "name": student.full_name or "(nom absent)",
                    "matricule": student.matricule,
                    "encadrant": student.encadrant.full_name if student.encadrant else "—",
                    "jury_name": jury.name,
                    "defense_date": jury.defense_date,
                    "start_time": slot,
                    "end_time": slot_end_time(jury.defense_date, slot),
                    "salle": jury.get_salle_display() if jury.salle else "",
                    "is_validated": jury.is_validated,
                })
                students.pop(0)
                filled += 1
                result["assigned"] += 1
                result["scheduled"] += 1
                placed = True
                break
            if not placed:
                break
    result["report"]["filled_existing"] = filled

    # 3. Unités de planification : (date, demi-journée) de la fenêtre
    all_units = []
    d = window_start
    while d <= end_date:
        all_units.append((d, defense_slots.MORNING))
        all_units.append((d, defense_slots.AFTERNOON))
        d += timedelta(days=1)

    def unit_len_slots(date, slot):
        s, e = defense_slots.slot_bounds(date, slot)
        minutes = int(
            (datetime.combine(date, e) - datetime.combine(date, s)).total_seconds() // 60
        )
        return minutes // DEFENSE_DURATION_MINUTES

    # Disponibilités par (prof, unité), en minutes utiles.
    prof_unit_minutes = defaultdict(int)
    for row in ProfessorAvailability.objects.filter(
        date__gte=window_start, date__lte=end_date
    ):
        for slot in defense_slots.slots_touched(row.date, row.start_time, row.end_time):
            s, e = defense_slots.slot_bounds(row.date, slot)
            start = max(row.start_time, s)
            end = min(row.end_time, e)
            minutes = int(
                (datetime.combine(row.date, end) - datetime.combine(row.date, start)).total_seconds() // 60
            )
            if minutes >= DEFENSE_DURATION_MINUTES:
                prof_unit_minutes[(row.professor_id, row.date, slot)] += minutes

    def prof_units(pid):
        return [
            (dte, slot) for (dte, slot) in all_units
            if prof_unit_minutes.get((pid, dte, slot), 0) >= DEFENSE_DURATION_MINUTES
        ]

    def supply_slots_for(pid):
        """Offre sous la règle stricte : au plus UNE demi-journée par jour."""
        per_day = {}
        for (dte, slot) in prof_units(pid):
            mins = prof_unit_minutes[(pid, dte, slot)]
            per_day[dte] = max(per_day.get(dte, 0), mins)
        return sum(m // DEFENSE_DURATION_MINUTES for m in per_day.values())

    # 4. PRÉ-VOL : faisabilité par encadrant + plafond global (rapport)
    for enc_id, studs in students_by_encadrant.items():
        supply = supply_slots_for(enc_id)
        need = len(studs)
        enc = prof_by_id.get(enc_id)
        result["report"]["feasibility"]["encadrants"].append({
            "encadrant": enc.full_name if enc else "?",
            "students": need,
            "supply_slots": supply,
            "missing_slots": max(0, need - supply),
            "ok": supply >= need,
        })
    result["report"]["feasibility"]["encadrants"].sort(
        key=lambda r: (r["ok"], -r["missing_slots"])
    )
    global_capacity = sum(unit_len_slots(dte, slot) for (dte, slot) in all_units) * cap
    result["report"]["feasibility"]["global"] = {
        "students": len(all_ready),
        "capacity_slots": global_capacity,
        "ok": len(all_ready) <= global_capacity,
    }

    jury_index = 1

    def unit_start_times(date, slot):
        s, e = defense_slots.slot_bounds(date, slot)
        cur = datetime.combine(date, s)
        limit = datetime.combine(date, e)
        out = []
        while cur + timedelta(minutes=DEFENSE_DURATION_MINUTES) <= limit:
            out.append(cur.time())
            cur += timedelta(minutes=DEFENSE_DURATION_MINUTES)
        return out

    def free_at(p, date, t, slot):
        return (
            is_professor_available(p, date, t, DEFENSE_DURATION_MINUTES)
            and not professor_has_conflict(p, date, t, DEFENSE_DURATION_MINUTES)
            and not professor_busy_other_slot(p, date, slot)
        )

    # PK des jurys créés durant CE run (seuls fusionnables entre eux ; les
    # jurys existants ne sont jamais modifiés par la génération).
    created_ids = set()

    def register_jury(jury, members, sel, sel_slots, date, t, experts_here):
        nonlocal jury_index
        created_new = getattr(jury, "_created_new", True)
        # Jurys créés durant CE run : seuls eux sont fusionnables entre eux.
        created_ids.add(jury.pk)
        if created_new:
            result["created"] += 1
            jury_index += 1
            if any(getattr(m, "is_priority", False) for m in members):
                result["report"]["coverage"]["with_priority"] += 1
            if experts_here:
                result["report"]["coverage"]["with_expert"] += 1
        result["assigned"] += len(sel)
        result["scheduled"] += len(sel)

        # Entrée de rapport : fusionnée si le jury a été PROLONGÉ (mêmes
        # membres, même jour, même demi-journée → un seul jury).
        report_entry = None
        for e in result["report"]["juries"]:
            if e.get("jury_pk") == jury.pk:
                report_entry = e
                break
        if report_entry is None:
            report_entry = {
                "jury_pk": jury.pk,
                "jury_name": jury.name,
                "salle": jury.get_salle_display() if jury.salle else "",
                "members": [p.full_name for p in members],
                "defense_date": date,
                "slot_start": t,
                "capacity": 0,
                "students_scheduled": [],
            }
            result["report"]["juries"].append(report_entry)
        else:
            report_entry["slot_start"] = min(report_entry["slot_start"], t)
        for student, slot_start in zip(sel, sel_slots):
            slot_end = (
                datetime.combine(date, slot_start)
                + timedelta(minutes=DEFENSE_DURATION_MINUTES)
            ).time()
            report_entry["students_scheduled"].append({
                "name": student.full_name or "(nom absent)",
                "matricule": student.matricule,
                "encadrant": student.encadrant.full_name if student.encadrant else "—",
                "start_time": slot_start,
                "end_time": slot_end,
            })
        report_entry["capacity"] = len(report_entry["students_scheduled"])
        report_entry["students_scheduled"].sort(key=lambda s: s["start_time"])
        for student in sel:
            pool = students_by_encadrant.get(student.encadrant_id, [])
            if student in pool:
                pool.remove(student)

    def try_place(enc, relaxed=False):
        """Tente de créer UN jury pour l'encadrant `enc` (et co-encadrants).

        Chaque créneau candidat est noté : prioritaire libre = +2, expert de
        la filière libre = +1. On choisit le meilleur créneau (à score égal,
        le plus tôt) au lieu du premier faisable — c'est ce qui maximise la
        présence des prioritaires et des experts sans sacrifier le placement.
        """
        remaining = students_by_encadrant.get(enc.id, [])
        if not remaining:
            return False
        target_filiere = remaining[0].filiere or ""
        experts_target = experts_by_filiere.get(target_filiere, set())

        candidates = {}  # score -> premier candidat validé à ce score
        best_score = -1

        for (date, slot) in prof_units(enc.id):
            if best_score >= 3:
                break
            # Règle stricte : l'encadrant ne siège pas dans les deux
            # demi-journées d'une même date.
            if professor_busy_other_slot(enc, date, slot):
                continue
            starts_here = unit_start_times(date, slot)
            for t in starts_here:
                if best_score >= 3:
                    break
                if jury_slot_capacity_reached(date, t, max_simultaneous=cap):
                    continue
                if not (
                    is_professor_available(enc, date, t, DEFENSE_DURATION_MINUTES)
                    and not professor_has_conflict(enc, date, t, DEFENSE_DURATION_MINUTES)
                ):
                    continue
                others = [
                    p for p in professors
                    if p.id != enc.id and free_at(p, date, t, slot)
                ]
                if len(others) < 2:
                    continue
                # Prioritaire : un seul par jury, le moins chargé (rotation
                # → tous les prioritaires sont consommés).
                prios = sorted(
                    [p for p in others if getattr(p, "is_priority", False)],
                    key=lambda p: (professor_total_scheduled_load(p), p.full_name.lower()),
                )
                chosen_p = prios[0] if prios else None
                expert_free = any(
                    p.id in experts_target
                    and (chosen_p is None or p.id != chosen_p.id)
                    for p in others
                )
                score = (2 if chosen_p is not None else 0) + (1 if expert_free else 0)
                if score <= best_score or score in candidates:
                    continue  # pas mieux qu'un candidat déjà validé

                # ── Validation complète du candidat ──────────────────────
                remaining_enc_ids = {
                    eid for eid, sts in students_by_encadrant.items()
                    if sts and eid != enc.id
                }
                seats = [enc] + ([chosen_p] if chosen_p is not None else [])
                pool_others = [p for p in others if all(p.id != s.id for s in seats)]
                if not relaxed:
                    # Strict : pas de 2e prioritaire dans le même jury.
                    pool_others = [
                        p for p in pool_others if not getattr(p, "is_priority", False)
                    ]
                expert_cands = [p for p in pool_others if p.id in experts_target]
                co_cands = sorted(
                    [p for p in pool_others if p.id in remaining_enc_ids],
                    key=lambda p: (
                        0 if (students_by_encadrant[p.id][0].filiere or "") == target_filiere else 1,
                        -len(students_by_encadrant[p.id]),
                        p.full_name.lower(),
                    ),
                )
                filler_cands = sorted(
                    [p for p in pool_others if p.id not in remaining_enc_ids],
                    key=lambda p: (
                        0 if p.id in experts_target else 1,
                        p.full_name.lower(),
                    ),
                )
                # 3e siège : expert d'abord (objectif « expert si possible »),
                # SAUF si un co-encadrant « serré » (peu de marge) a besoin du
                # co-placement pour que ses étudiants passent.
                tight_co = [
                    p for p in co_cands
                    if supply_slots_for(p.id) - len(students_by_encadrant[p.id]) <= 2
                ]
                if tight_co:
                    third_pref = tight_co + expert_cands + co_cands + filler_cands
                elif expert_cands:
                    third_pref = expert_cands + co_cands + filler_cands
                else:
                    third_pref = co_cands + filler_cands
                for cand in third_pref:
                    if len(seats) >= 3:
                        break
                    if all(cand.id != s.id for s in seats):
                        seats.append(cand)
                if len(seats) < 3:
                    # Dernier recours : tout prof libre (règles dures respectées).
                    for cand in others:
                        if len(seats) >= 3:
                            break
                        if all(cand.id != s.id for s in seats):
                            seats.append(cand)
                if len(seats) < 3:
                    continue
                members = seats[:3]
                block = build_consecutive_available_slots(
                    members=members, defense_date=date, block_start=t,
                    max_slots=40, max_simultaneous=cap,
                )
                if not block:
                    continue
                pool = []
                for m in members:
                    pool.extend(students_by_encadrant.get(m.id, []))
                # Priorité aux étudiants dont l'encadrant a le MOINS de marge
                # (dispo faisable - étudiants restants) : on ne consomme pas la
                # disponibilité d'un prof « serré » pour faire passer les
                # étudiants d'un prof confortable. Puis mono-filière.
                pool.sort(key=lambda s: (
                    supply_slots_for(s.encadrant_id)
                    - len(students_by_encadrant.get(s.encadrant_id, [])),
                    0 if s.encadrant_id == enc.id else 1,
                    0 if (s.filiere or "") == target_filiere else 1,
                    s.full_name.lower(),
                ))
                sel = pool[:len(block)]
                if not sel:
                    continue
                sel_slots = block[:len(sel)]
                block_end = (
                    datetime.combine(date, sel_slots[-1])
                    + timedelta(minutes=DEFENSE_DURATION_MINUTES)
                ).time()
                salle = _choisir_salle_libre(date, t, block_end)
                # Aucune salle libre sur TOUT le bloc : réduire le bloc plutôt
                # que d'abandonner le créneau (une salle peut être libre sur un
                # bloc plus court ; les étudiants restants passeront plus tard).
                while salle is None and len(sel) > 1:
                    sel = sel[:-1]
                    sel_slots = sel_slots[:len(sel)]
                    block_end = (
                        datetime.combine(date, sel_slots[-1])
                        + timedelta(minutes=DEFENSE_DURATION_MINUTES)
                    ).time()
                    salle = _choisir_salle_libre(date, t, block_end)
                if salle is None:
                    continue
                experts_here = {
                    m.id for m in members
                    if m.id in experts_by_filiere.get((sel[0].filiere or ""), set())
                }
                candidates[score] = {
                    "members": members, "sel": sel, "sel_slots": sel_slots,
                    "date": date, "t": t, "salle": salle,
                    "experts_here": experts_here,
                }
                best_score = score

        # Créer le meilleur candidat (score décroissant) ; en cas d'échec de
        # validation, essayer le suivant.
        for score in sorted(candidates.keys(), reverse=True):
            c = candidates[score]
            plan = {
                "members": c["members"],
                "students": c["sel"],
                "defense_date": c["date"],
                "start_times": c["sel_slots"],
                "salle": c["salle"],
                "experts": c["experts_here"],
            }
            try:
                jury = create_grouped_jury_from_plan(
                    plan, jury_index, mergeable_ids=created_ids
                )
            except ValidationError:
                continue
            register_jury(
                jury, c["members"], c["sel"], c["sel_slots"],
                c["date"], c["t"], c["experts_here"],
            )
            return True
        return False

    # 5. PLACEMENT : le plus contraint d'abord, jusqu'à point fixe.
    #    1er tour strict (préférences respectées), 2e tour relâché
    #    (expert / un-seul-prioritaire sacrifiés — jamais les règles dures).
    for relaxed in (False, True):
        progress = True
        while progress and any(students_by_encadrant.values()):
            progress = False
            enc_ids = [eid for eid, sts in students_by_encadrant.items() if sts]
            enc_ids.sort(key=lambda eid: (
                supply_slots_for(eid) - len(students_by_encadrant[eid]),
                -len(students_by_encadrant[eid]),
            ))
            for eid in enc_ids:
                enc = prof_by_id.get(eid)
                if enc is not None and try_place(enc, relaxed=relaxed):
                    progress = True
                    break

    # 6. 2e passe : étudiants dont l'encadrant n'a AUCUNE disponibilité future.
    #    Jury SANS l'encadrant mais avec un expert de la filière (« encadrant
    #    absent »).
    encadrant_has_avail = {}

    def _enc_absent(enc_id):
        if enc_id not in encadrant_has_avail:
            encadrant_has_avail[enc_id] = ProfessorAvailability.objects.filter(
                professor_id=enc_id, date__gte=today
            ).exists()
        return not encadrant_has_avail[enc_id]

    absent_by_fil = defaultdict(list)
    for enc_id, students in students_by_encadrant.items():
        if students and _enc_absent(enc_id):
            for student in students:
                absent_by_fil[student.filiere or ""].append(student)
            students_by_encadrant[enc_id] = []

    if any(absent_by_fil.values()):
        candidate_slots = build_all_future_slot_starts(start_date, end_date)
        for defense_date, block_start in candidate_slots:
            if not any(absent_by_fil.values()):
                break
            current_slot = _slot_label_at(defense_date, block_start)

            progress = True
            while progress and any(absent_by_fil.values()):
                progress = False
                if jury_slot_capacity_reached(defense_date, block_start, max_simultaneous=cap):
                    break
                available_profs = [
                    p for p in professors
                    if free_at(p, defense_date, block_start, current_slot)
                ]
                if len(available_profs) < 3:
                    break

                for fil, studs in list(absent_by_fil.items()):
                    if not studs:
                        continue
                    experts_here = [
                        p for p in available_profs
                        if p.id in experts_by_filiere.get(fil, set())
                    ]
                    if not experts_here:
                        continue  # pas d'expert dispo → remplacement impossible
                    expert = experts_here[0]
                    others_pool = [p for p in available_profs if p.id != expert.id]
                    prio_others = sorted(
                        [p for p in others_pool if getattr(p, "is_priority", False)],
                        key=lambda p: (professor_total_scheduled_load(p), p.full_name.lower()),
                    )
                    non_prio_others = [
                        p for p in others_pool if not getattr(p, "is_priority", False)
                    ]
                    others = (prio_others[:1] + non_prio_others)[:2]
                    if len(others) < 2:
                        others = others_pool[:2]
                    if len(others) < 2:
                        continue
                    members = [expert] + others
                    slots_avail = build_consecutive_available_slots(
                        members=members, defense_date=defense_date,
                        block_start=block_start, max_slots=40,
                        max_simultaneous=cap,
                    )
                    if not slots_avail:
                        continue
                    sel = studs[:len(slots_avail)]
                    sel_slots = slots_avail[:len(sel)]
                    block_end = (
                        datetime.combine(defense_date, sel_slots[-1])
                        + timedelta(minutes=DEFENSE_DURATION_MINUTES)
                    ).time()
                    salle = _choisir_salle_libre(defense_date, block_start, block_end)
                    if salle is None:
                        continue

                    plan = {
                        "members": members,
                        "students": sel,
                        "defense_date": defense_date,
                        "start_times": sel_slots,
                        "salle": salle,
                        "experts": {expert.id},
                        "encadrant_absent": True,
                    }
                    try:
                        jury = create_grouped_jury_from_plan(
                            plan, jury_index, mergeable_ids=created_ids
                        )
                    except ValidationError as exc:
                        for student in sel:
                            result["errors"].append({
                                "student": student, "reason": "validation_error",
                                "message": "; ".join(exc.messages),
                            })
                        continue

                    register_jury(
                        jury, members, sel, sel_slots, defense_date, block_start,
                        {expert.id},
                    )
                    for s in sel:
                        if s in absent_by_fil[fil]:
                            absent_by_fil[fil].remove(s)
                    progress = True
                    break

    # 7. Non placés : raisons précises et chiffrées (issues du pré-vol).
    feas_by_name = {
        r["encadrant"]: r for r in result["report"]["feasibility"]["encadrants"]
    }
    for enc_id, students in students_by_encadrant.items():
        if not students:
            continue
        enc = prof_by_id.get(enc_id)
        row = feas_by_name.get(enc.full_name if enc else "", {})
        if row and not row.get("ok", True):
            reason = "insufficient_availability"
            message = (
                f"Disponibilités insuffisantes de l'encadrant : {row['students']} étudiant(s) "
                f"pour {row['supply_slots']} créneau(x) de 20 min (règle matin OU après-midi "
                f"comprise). Il manque environ {row['missing_slots']} créneau(x) — demander "
                f"plus de disponibilités."
            )
        else:
            reason = "no_slot_found"
            message = "Aucun créneau compatible trouvé (conflits, salles ou capacité)."
        for student in students:
            result["errors"].append({
                "student": student, "reason": reason, "message": message,
            })
    for fil, studs in absent_by_fil.items():
        for student in studs:
            result["errors"].append({
                "student": student,
                "reason": "no_expert",
                "message": "Encadrant sans disponibilité et aucun expert disponible pour le remplacer.",
            })

    # 8. Bilan d'utilisation des profs prioritaires — taux calculé sur la
    #    disponibilité FAISABLE (au plus une demi-journée par jour, règle
    #    stricte : déclarer matin + après-midi ne compte qu'une fois).
    priority_usage = []
    for prof in professors:
        if not getattr(prof, "is_priority", False):
            continue
        used, total, _free = feasible_priority_usage(prof, today)
        priority_usage.append({
            "name": prof.full_name,
            "used": used,
            "total": total,
            "free": total - used,
            "pct": round(100 * used / total) if total else 0,
        })
    result["report"]["priority_usage"] = priority_usage

    # 9. Bilan par encadrant : prêts / affectés / restants / %.
    remaining_by_enc = {
        eid: len(sts) for eid, sts in students_by_encadrant.items()
    }
    for fil, studs in absent_by_fil.items():
        for student in studs:
            remaining_by_enc[student.encadrant_id] = (
                remaining_by_enc.get(student.encadrant_id, 0) + 1
            )
    by_encadrant = []
    for enc_id, ready in initial_by_enc.items():
        left = remaining_by_enc.get(enc_id, 0)
        placed = ready - left
        enc = prof_by_id.get(enc_id)
        by_encadrant.append({
            "encadrant": enc.full_name if enc else "?",
            "ready": ready,
            "placed": placed,
            "remaining": left,
            "pct": round(100 * placed / ready) if ready else 0,
        })
    by_encadrant.sort(key=lambda r: (-r["remaining"], r["encadrant"].lower()))
    result["report"]["by_encadrant"] = by_encadrant

    # 10. Jurys du rapport en ordre chronologique (date puis heure).
    result["report"]["juries"].sort(
        key=lambda e: (e["defense_date"], e["slot_start"])
    )

    # 11. Renumérotation PAR JOUR (matin puis après-midi) de tous les jurys,
    #     puis rafraîchissement des noms dans le rapport.
    renumber_all_juries()
    pk_list = [
        e.get("jury_pk") for e in result["report"]["juries"] if e.get("jury_pk")
    ]
    fresh_names = dict(
        Jury.objects.filter(pk__in=pk_list).values_list("pk", "name")
    )
    for e in result["report"]["juries"]:
        if e.get("jury_pk") in fresh_names:
            e["jury_name"] = fresh_names[e["jury_pk"]]

    return result


@transaction.atomic
def generate_juries_for_date(defense_date=None):
    return generate_smart_juries()


# ══════════════════════════════════════════════════════════════════════════
#  ALGORITHME ALTERNATIF « PAR FILIÈRE » (à tester)
#  - jurys mono-filière ; un expert de la filière maximisé dans chaque jury ;
#    encadrant gardé si disponible, sinon expert ; président prioritaire favorisé.
#  - APERÇU non destructif (transaction annulée) ; APPLICATION = remplace les
#    jurys PUBLIÉS de la fenêtre (hors jurys contenant des étudiants déjà notés).
# ══════════════════════════════════════════════════════════════════════════

class _PreviewRollback(Exception):
    """Sert à annuler la transaction après avoir capturé l'aperçu."""


def _window_published_target(start_date, end_date):
    """Étudiants actuellement dans des jurys PUBLIÉS de la fenêtre, à
    re-générer. On EXCLUT tout jury contenant un étudiant déjà noté (pour ne
    pas toucher aux soutenances terminées)."""
    juries = Jury.objects.filter(
        is_validated=True,
        defense_date__gte=start_date,
        defense_date__lte=end_date,
    ).prefetch_related("students__student__encadrant")
    students, deletable_ids, skipped = [], [], 0
    for j in juries:
        jss = list(j.students.all())
        has_graded = any(
            _has_real_grades(js) for js in jss
        )
        if has_graded:
            skipped += 1
            continue
        deletable_ids.append(j.id)
        students.extend(js.student for js in jss)
    return students, deletable_ids, skipped


def _by_filiere_place(students, start_date, end_date, cap, result):
    """Place les `students` en jurys mono-filière (expert de filière requis,
    encadrant si dispo, président prioritaire favorisé)."""
    from collections import defaultdict

    experts_by_filiere = defaultdict(set)
    for e in FiliereExpert.objects.all():
        experts_by_filiere[e.filiere].add(e.professor_id)
    professors = list(ProfessorProfile.objects.all())

    by_fil = defaultdict(list)
    for s in students:
        by_fil[s.filiere or "?"].append(s)
    for k in by_fil:
        by_fil[k].sort(key=lambda s: (s.full_name or "").lower())

    candidate_slots = build_all_future_slot_starts(start_date, end_date)
    slot_used = defaultdict(set)
    jury_index = 1

    for defense_date, block_start in candidate_slots:
        if not any(by_fil.values()):
            break
        current_slot = _slot_label_at(defense_date, block_start)
        other_slot = (
            defense_slots.AFTERNOON if current_slot == defense_slots.MORNING
            else defense_slots.MORNING
        )
        while any(by_fil.values()):
            if jury_slot_capacity_reached(defense_date, block_start, max_simultaneous=cap):
                break
            blocked = slot_used.get((defense_date, other_slot), set())
            avail = [
                p for p in professors
                if p.id not in blocked
                and is_professor_available(p, defense_date, block_start, DEFENSE_DURATION_MINUTES)
                and not professor_has_conflict(p, defense_date, block_start, DEFENSE_DURATION_MINUTES)
                and not professor_busy_other_slot(p, defense_date, current_slot)
            ]
            if len(avail) < 3:
                break

            # Choisir une filière ayant des étudiants ET un expert disponible.
            chosen = None
            for fil, studs in sorted(by_fil.items(), key=lambda kv: -len(kv[1])):
                if not studs:
                    continue
                experts_here = [p for p in avail if p.id in experts_by_filiere.get(fil, set())]
                if experts_here:
                    chosen = (fil, studs, experts_here)
                    break
            if chosen is None:
                break
            fil, studs, experts_here = chosen

            # Expert : prioritaire d'abord, puis le moins chargé.
            experts_here.sort(key=lambda p: (
                0 if getattr(p, "is_priority", False) else 1,
                professor_total_scheduled_load(p), (p.full_name or "").lower(),
            ))
            expert = experts_here[0]

            # Membres : expert + encadrants dispo de ces étudiants + remplisseurs
            # (prioritaires d'abord pour la présidence).
            enc_ids = {s.encadrant_id for s in studs}
            members = [expert]
            for p in avail:
                if len(members) >= 3:
                    break
                if p.id != expert.id and p.id in enc_ids:
                    members.append(p)
            if len(members) < 3:
                fillers = [p for p in avail if p not in members]
                fillers.sort(key=lambda p: (
                    0 if getattr(p, "is_priority", False) else 1,
                    (p.full_name or "").lower(),
                ))
                for p in fillers:
                    if len(members) >= 3:
                        break
                    members.append(p)
            if len(members) < 3:
                break

            block = build_consecutive_available_slots(
                members=members, defense_date=defense_date,
                block_start=block_start, max_slots=40, max_simultaneous=cap,
            )
            if not block:
                break
            sel = studs[:len(block)]
            sel_slots = block[:len(sel)]
            block_end = slot_end_time(defense_date, sel_slots[-1])
            salle = _choisir_salle_libre(defense_date, block_start, block_end)
            while salle is None and len(sel) > 1:
                sel = sel[:-1]
                sel_slots = sel_slots[:len(sel)]
                block_end = slot_end_time(defense_date, sel_slots[-1])
                salle = _choisir_salle_libre(defense_date, block_start, block_end)
            if salle is None or not sel:
                break

            member_ids = {m.id for m in members}
            president_order = sorted(members, key=lambda p: (
                0 if getattr(p, "is_priority", False) else 1,
                0 if p.id == expert.id else 1,
                professor_load_on_date(p, defense_date),
                (p.full_name or "").lower(),
            ))
            try:
                with transaction.atomic():
                    jury = Jury.objects.create(
                        name=build_grouped_jury_name(sel, defense_date),
                        defense_date=defense_date, salle=salle, is_validated=False,
                    )
                    for m in members:
                        JuryMember.objects.create(jury=jury, professor=m)
                    for s, t in zip(sel, sel_slots):
                        president = next(
                            (p for p in president_order if p.id != s.encadrant_id), None
                        )
                        js = JuryStudent.objects.create(
                            jury=jury, student=s, president=president,
                            encadrant_absent=(s.encadrant_id not in member_ids),
                        )
                        DefenseSchedule.objects.create(
                            jury_student=js, start_time=t,
                            duration_minutes=DEFENSE_DURATION_MINUTES,
                        )
            except ValidationError as exc:
                for s in sel:
                    result["errors"].append({
                        "student": s, "reason": "validation_error",
                        "message": "; ".join(exc.messages),
                    })
                break

            for s in sel:
                by_fil[fil].remove(s)
            slot_used[(defense_date, current_slot)].update(m.id for m in members)
            result["created"] += 1
            result["assigned"] += len(sel)
            jury_index += 1
            entry = {
                "jury_pk": jury.pk, "jury_name": jury.name,
                "salle": jury.get_salle_display() if jury.salle else "",
                "members": [p.full_name for p in members],
                "filiere": fil,
                "defense_date": defense_date, "slot_start": sel_slots[0],
                "capacity": len(sel), "students_scheduled": [],
            }
            for s, t in zip(sel, sel_slots):
                entry["students_scheduled"].append({
                    "name": s.full_name or "(nom absent)", "matricule": s.matricule,
                    "encadrant": s.encadrant.full_name if s.encadrant else "—",
                    "start_time": t, "end_time": slot_end_time(defense_date, t),
                })
            result["report"]["juries"].append(entry)

    # Étudiants non placés.
    for fil, studs in by_fil.items():
        for s in studs:
            result["errors"].append({
                "student": s, "reason": "no_slot_found",
                "message": f"Filière {fil} : aucun créneau/expert disponible.",
            })


def run_by_filiere(start_date, end_date, cap, commit):
    """Supprime les jurys publiés (hors notés) de la fenêtre et régénère par
    filière. commit=False → aperçu (transaction annulée, rien n'est modifié)."""
    def _work():
        students, deletable_ids, skipped = _window_published_target(start_date, end_date)
        result = {
            "created": 0, "assigned": 0, "errors": [],
            "report": {"total_ready": len(students), "juries": [], "replaced_juries": len(deletable_ids), "skipped_graded": skipped},
        }
        # Libère les étudiants en supprimant les jurys publiés (sans notes).
        Jury.objects.filter(id__in=deletable_ids).delete()
        _by_filiere_place(students, start_date, end_date, cap, result)
        if commit:
            # Les nouveaux jurys remplacent les anciens : publiés + notifiés.
            new_ids = [e["jury_pk"] for e in result["report"]["juries"]]
            Jury.objects.filter(id__in=new_ids).update(is_validated=True)
            for jid in new_ids:
                jury = Jury.objects.filter(pk=jid).first()
                if jury:
                    _notify_jury_published(jury)
        return result

    if commit:
        with transaction.atomic():
            return build_generation_report_payload(_work())
    # Aperçu : on exécute puis on annule tout.
    payload = {}
    try:
        with transaction.atomic():
            payload = build_generation_report_payload(_work())
            raise _PreviewRollback()
    except _PreviewRollback:
        pass
    return payload


def build_all_future_slot_starts(start_date=None, end_date=None):
    if start_date is None:
        start_date = DEFENSE_START
    if end_date is None:
        end_date = DEFENSE_DEADLINE

    today = timezone.localdate()
    now_time = timezone.localtime().time()

    starts = set()

    # Fenêtre des soutenances : du start_date au end_date inclus (week-end compris).
    lower = max(today, start_date)
    availabilities = ProfessorAvailability.objects.filter(
        date__gte=lower,
        date__lte=end_date,
    ).order_by(
        "date",
        "start_time",
        "end_time",
    )

    for availability in availabilities:
        for defense_date, start_time in build_slots_from_availability(availability):
            if defense_date > today or start_time > now_time:
                starts.add((defense_date, start_time))

    return sorted(
        starts,
        key=lambda item: (item[0], item[1])
    )


def build_consecutive_available_slots(members, defense_date, block_start, max_slots, max_simultaneous=None):
    if jury_slot_capacity_reached(defense_date, block_start, max_simultaneous=max_simultaneous):
        return []

    # Le jury reste dans le créneau (matin/après-midi) où il démarre : on borne
    # la fin des créneaux consécutifs à la fin de ce créneau.
    slot_label = _slot_label_at(defense_date, block_start)
    slot_end = None
    if slot_label:
        _, slot_end = defense_slots.slot_bounds(defense_date, slot_label)

    slots = []
    cursor = datetime.combine(defense_date, block_start)

    for index in range(max_slots):
        current_time = (
            cursor + timedelta(minutes=index * DEFENSE_DURATION_MINUTES)
        ).time()

        # Ne pas déborder du créneau (matin/après-midi).
        if slot_end is not None:
            end_dt = (
                datetime.combine(defense_date, current_time)
                + timedelta(minutes=DEFENSE_DURATION_MINUTES)
            )
            if end_dt > datetime.combine(defense_date, slot_end):
                break

        all_members_available = True

        for professor in members:
            if not is_professor_available(
                professor,
                defense_date,
                current_time,
                DEFENSE_DURATION_MINUTES,
            ):
                all_members_available = False
                break

            if professor_has_conflict(
                professor,
                defense_date,
                current_time,
                DEFENSE_DURATION_MINUTES,
            ):
                all_members_available = False
                break

        if not all_members_available:
            break

        slots.append(current_time)

    return slots


def create_grouped_jury_from_plan(plan, jury_index, mergeable_ids=None):
    defense_date = plan["defense_date"]
    start_times = plan["start_times"]
    member_ids = sorted(p.id for p in plan["members"])
    slot_label = _slot_label_at(defense_date, start_times[0]) if start_times else None
    block_end = (
        datetime.combine(defense_date, start_times[-1])
        + timedelta(minutes=DEFENSE_DURATION_MINUTES)
    ).time() if start_times else None

    # ── FUSION : mêmes 3 membres + même date + même demi-journée = UN SEUL
    #    jury. On ne prolonge QUE les jurys créés pendant CE run de génération
    #    (mergeable_ids) : les jurys existants (brouillons compris) ne sont
    #    JAMAIS modifiés par la génération automatique.
    existing = None
    merge_pool = Jury.objects.filter(
        defense_date=defense_date, is_validated=False
    )
    if mergeable_ids is None:
        merge_pool = merge_pool.none()
    else:
        merge_pool = merge_pool.filter(pk__in=mergeable_ids)
    for candidate in merge_pool.prefetch_related("members"):
        ids = sorted(m.professor_id for m in candidate.members.all())
        if ids != member_ids:
            continue
        first_sched = DefenseSchedule.objects.filter(
            jury_student__jury=candidate
        ).order_by("start_time").first()
        if first_sched and _slot_label_at(
            defense_date, first_sched.start_time
        ) != slot_label:
            continue
        # Un jury = une salle : sa salle doit être libre sur le nouveau bloc.
        if candidate.salle and block_end and _salle_occupee(
            defense_date, start_times[0], block_end, candidate.salle
        ):
            continue
        existing = candidate
        break

    created_new = existing is None

    encadrant_absent = plan.get("encadrant_absent", False)

    # Ordre des présidents pour ce jury (président fixe autant que possible).
    # On classe les membres une seule fois : prof prioritaire d'abord, puis
    # expert de la filière, puis le moins chargé. Chaque étudiant reçoit le
    # 1er membre de cette liste qui n'est pas son encadrant. Le président
    # préside donc le maximum d'étudiants ; on ne bascule vers le membre
    # suivant que pour ses propres encadrés.
    experts = plan.get("experts", set())
    president_order = sorted(
        plan["members"],
        key=lambda p: (
            0 if getattr(p, "is_priority", False) else 1,
            0 if p.id in experts else 1,
            # Continuité : celui qui préside déjà des étudiants continue en
            # tant que président le plus longtemps possible avant qu'un autre
            # membre prenne le relais.
            -JuryStudent.objects.filter(president=p).count(),
            professor_load_on_date(p, defense_date),
            professor_total_scheduled_load(p),
            p.full_name.lower(),
        ),
    )

    with transaction.atomic():
        if created_new:
            jury = Jury.objects.create(
                name=build_grouped_jury_name(
                    students=plan["students"],
                    defense_date=defense_date,
                ),
                defense_date=defense_date,
                salle=plan.get("salle", ""),
                is_validated=False,
            )
            for professor in plan["members"]:
                JuryMember.objects.create(
                    jury=jury,
                    professor=professor,
                )
        else:
            jury = existing

        for student, start_time in zip(plan["students"], start_times):
            president = next(
                (p for p in president_order if p.id != student.encadrant_id),
                None,
            )

            assignment = JuryStudent.objects.create(
                jury=jury,
                student=student,
                president=president,
                encadrant_absent=encadrant_absent,
            )

            DefenseSchedule.objects.create(
                jury_student=assignment,
                start_time=start_time,
                duration_minutes=DEFENSE_DURATION_MINUTES,
            )

    jury._created_new = created_new
    return jury


def build_grouped_jury_name(students, defense_date):
    """Numérotation PAR JOUR : Jury 1, 2, ... selon les jurys déjà créés à
    cette date. La renumérotation finale (renumber_draft_juries) remet les
    numéros dans l'ordre chronologique des passages."""
    index = Jury.objects.filter(defense_date=defense_date).count() + 1
    n = len(students)
    return f"Jury {index} - {n} étudiant{'s' if n > 1 else ''}"


def find_common_block(members, new_date, n, preferred_salle=""):
    """Cherche un bloc de `n` passages de 20 min consécutifs, le même jour,
    dans une seule demi-journée, où les 3 membres sont TOUS disponibles (dispo
    déclarée + sans conflit), avec une salle libre. Renvoie (start_times,
    salle) ou (None, None) si aucune disponibilité commune n'existe."""
    common = sorted(
        t for (d, t) in get_common_available_slots(members) if d == new_date
    )
    if not common:
        return None, None
    common_set = set(common)
    for start in common:
        label = _slot_label_at(new_date, start)
        if not label:
            continue
        _, hd_end = defense_slots.slot_bounds(new_date, label)
        times = []
        cursor = datetime.combine(new_date, start)
        ok = True
        for _ in range(n):
            t = cursor.time()
            nxt = cursor + timedelta(minutes=DEFENSE_DURATION_MINUTES)
            if t not in common_set or nxt > datetime.combine(new_date, hd_end):
                ok = False
                break
            times.append(t)
            cursor = nxt
        if not ok or len(times) != n:
            continue
        block_end = slot_end_time(new_date, times[-1])
        if preferred_salle and not _salle_occupee(new_date, times[0], block_end, preferred_salle):
            return times, preferred_salle
        salle = _choisir_salle_libre(new_date, times[0], block_end)
        if salle:
            return times, salle
    return None, None


def find_free_slot_in_jury(jury, members=None):
    """Trouve le PREMIER créneau de 20 min libre dans la demi-journée du jury
    (trou dans le programme OU place à la fin), SANS déplacer les étudiants
    déjà présents. Conditions : les 3 membres disponibles + salle libre à ce
    créneau. Renvoie le start_time ou None."""
    if members is None:
        members = [m.professor for m in jury.members.select_related("professor")]
    if len(members) < 3:
        return None
    scheds = list(
        DefenseSchedule.objects.filter(jury_student__jury=jury).order_by("start_time")
    )
    if not scheds:
        return None
    label = _slot_label_at(jury.defense_date, scheds[0].start_time)
    if not label:
        return None
    hd_start, hd_end = defense_slots.slot_bounds(jury.defense_date, label)
    occupied = {s.start_time for s in scheds}
    cursor = datetime.combine(jury.defense_date, hd_start)
    limit = datetime.combine(jury.defense_date, hd_end)
    while cursor + timedelta(minutes=DEFENSE_DURATION_MINUTES) <= limit:
        t = cursor.time()
        if t not in occupied:
            end = slot_end_time(jury.defense_date, t)
            members_ok = all(
                is_professor_available(m, jury.defense_date, t)
                and not professor_has_conflict(m, jury.defense_date, t)
                for m in members
            )
            room_ok = not (
                jury.salle and _salle_occupee(jury.defense_date, t, end, jury.salle)
            )
            if members_ok and room_ok:
                return t
        cursor = cursor + timedelta(minutes=DEFENSE_DURATION_MINUTES)
    return None


def free_slot_at_end_of_jury(jury, members=None):
    """Prochain créneau de 20 min libre à la SUITE des passages d'un jury,
    s'il reste du temps dans sa demi-journée (matin fini avant 14h, après-midi
    avant 19h) et si les 3 membres y sont disponibles + la salle libre.
    Renvoie le start_time ou None."""
    if members is None:
        members = [m.professor for m in jury.members.select_related("professor")]
    if len(members) < 3:
        return None
    scheds = list(
        DefenseSchedule.objects.filter(jury_student__jury=jury).order_by("start_time")
    )
    if not scheds:
        return None
    label = _slot_label_at(jury.defense_date, scheds[0].start_time)
    if not label:
        return None
    _, hd_end = defense_slots.slot_bounds(jury.defense_date, label)
    last = scheds[-1]
    start = last.end_time or slot_end_time(
        jury.defense_date, last.start_time,
        last.duration_minutes or DEFENSE_DURATION_MINUTES,
    )
    # Reste-t-il un créneau de 20 min dans la demi-journée ?
    if (datetime.combine(jury.defense_date, start)
            + timedelta(minutes=DEFENSE_DURATION_MINUTES)
            > datetime.combine(jury.defense_date, hd_end)):
        return None
    # Les 3 membres disponibles et sans conflit à ce créneau.
    for m in members:
        if not is_professor_available(m, jury.defense_date, start):
            return None
        if professor_has_conflict(m, jury.defense_date, start):
            return None
    # Salle libre à ce créneau (aucun autre jury n'y chevauche).
    if jury.salle and _salle_occupee(
        jury.defense_date, start, slot_end_time(jury.defense_date, start), jury.salle
    ):
        return None
    return start


def build_block_from_start(members, on_date, start, n, preferred_salle=""):
    """Vérifie qu'à partir de `start` le `on_date`, il y a `n` passages
    consécutifs où les 3 membres sont disponibles + une salle libre. Renvoie
    (start_times, salle) ou (None, None)."""
    if n <= 0:
        return [], (preferred_salle or "")
    common = {t for (d, t) in get_common_available_slots(members) if d == on_date}
    label = _slot_label_at(on_date, start)
    if not label:
        return None, None
    _, hd_end = defense_slots.slot_bounds(on_date, label)
    times = []
    cursor = datetime.combine(on_date, start)
    for _ in range(n):
        t = cursor.time()
        nxt = cursor + timedelta(minutes=DEFENSE_DURATION_MINUTES)
        if t not in common or nxt > datetime.combine(on_date, hd_end):
            return None, None
        times.append(t)
        cursor = nxt
    block_end = slot_end_time(on_date, times[-1])
    if preferred_salle and not _salle_occupee(on_date, times[0], block_end, preferred_salle):
        return times, preferred_salle
    salle = _choisir_salle_libre(on_date, times[0], block_end)
    if salle:
        return times, salle
    return None, None


def reschedule_options_for_jury(jury, members, n, max_options=60):
    """Liste des créneaux (date + demi-journée) où les 3 membres sont
    disponibles simultanément pour `n` passages consécutifs, avec une salle
    libre. Un créneau par (date, demi-journée), le plus tôt. C'est ce que
    l'admin choisit pour reprogrammer (au lieu d'une date au hasard)."""
    from collections import defaultdict

    need = max(n, 1)
    by_key = defaultdict(set)
    for d, t in get_common_available_slots(members):
        label = _slot_label_at(d, t)
        if label:
            by_key[(d, label)].add(t)

    options = []
    for (d, label) in sorted(by_key):
        _, hd_end = defense_slots.slot_bounds(d, label)
        for start in sorted(by_key[(d, label)]):
            times = []
            cursor = datetime.combine(d, start)
            ok = True
            for _ in range(need):
                t = cursor.time()
                nxt = cursor + timedelta(minutes=DEFENSE_DURATION_MINUTES)
                if t not in by_key[(d, label)] or nxt > datetime.combine(d, hd_end):
                    ok = False
                    break
                times.append(t)
                cursor = nxt
            if not ok:
                continue
            block_end = slot_end_time(d, times[-1])
            salle = (
                jury.salle if (jury.salle and not _salle_occupee(d, times[0], block_end, jury.salle))
                else _choisir_salle_libre(d, times[0], block_end)
            )
            if not salle:
                continue
            options.append({
                "value": f"{d.isoformat()}|{times[0].strftime('%H:%M')}",
                "date": d,
                "start": times[0],
                "end": block_end,
                "half": "Matin" if label == defense_slots.MORNING else "Après-midi",
                "salle": salle,
                "current": (d == jury.defense_date),
            })
            break  # un seul créneau par (date, demi-journée)
        if len(options) >= max_options:
            break
    return options


def recompact_jury_schedule(jury):
    """Réorganise les horaires d'un jury pour supprimer les trous : les
    passages restants sont replacés de façon contiguë (20 min chacun) à
    partir de l'heure du premier passage. S'applique aussi aux jurys
    PUBLIÉS ; dans ce cas, les étudiants dont l'horaire change sont prévenus.
    Utilise update() pour ne pas rejouer la validation de disponibilité
    (le bloc était déjà valide)."""
    scheds = list(
        DefenseSchedule.objects.filter(jury_student__jury=jury)
        .select_related("jury_student__student__user")
        .order_by("start_time", "id")
    )
    if not scheds:
        return
    cursor = datetime.combine(jury.defense_date, scheds[0].start_time)
    changed = []
    for sched in scheds:
        minutes = sched.duration_minutes or DEFENSE_DURATION_MINUTES
        new_start = cursor.time()
        new_end = (cursor + timedelta(minutes=minutes)).time()
        if sched.start_time != new_start or sched.end_time != new_end:
            DefenseSchedule.objects.filter(pk=sched.pk).update(
                start_time=new_start, end_time=new_end,
            )
            changed.append((sched.jury_student, new_start, new_end))
        cursor = cursor + timedelta(minutes=minutes)

    # Jury déjà publié : prévenir les étudiants dont l'horaire a bougé.
    if jury.is_validated and changed:
        for js, new_start, new_end in changed:
            notify(
                getattr(js.student, "user", None),
                "Horaire de soutenance ajusté",
                f"Votre passage est désormais prévu à "
                f"{new_start.strftime('%H:%M')}–{new_end.strftime('%H:%M')} "
                f"le {jury.defense_date.strftime('%d/%m/%Y')} "
                f"(jury « {jury.name} »).",
                "/student-dashboard/",
                category=Notification.CATEGORY_JURY,
            )


def refresh_jury_name_count(jury):
    """Met à jour le nombre d'étudiants dans le nom du jury (« Jury 2 - N
    étudiants »), en conservant son numéro. À appeler après ajout/retrait
    d'un étudiant."""
    import re
    n = jury.students.count()
    m = re.match(r"(Jury\s+\d+)", jury.name or "")
    prefix = m.group(1) if m else f"Jury {jury.pk}"
    if n:
        new_name = f"{prefix} - {n} étudiant{'s' if n > 1 else ''}"
    else:
        new_name = prefix
    if new_name != jury.name:
        Jury.objects.filter(pk=jury.pk).update(name=new_name)
        jury.name = new_name


def renumber_all_juries():
    """Renumérote TOUS les jurys, PAR JOUR : d'abord les jurys du MATIN, puis
    ceux de l'APRÈS-MIDI, chacun dans l'ordre croissant des horaires. Le numéro
    se met donc à jour automatiquement à chaque ajout / suppression de jury."""
    from collections import defaultdict
    from datetime import time as _time

    first_start = {}
    for row in (
        DefenseSchedule.objects.order_by("start_time")
        .values("jury_student__jury_id", "start_time")
    ):
        jid = row["jury_student__jury_id"]
        if jid not in first_start:
            first_start[jid] = row["start_time"]

    by_date = defaultdict(list)
    for jury in Jury.objects.prefetch_related("students"):
        if not jury.defense_date:
            continue
        st = first_start.get(jury.pk)
        label = _slot_label_at(jury.defense_date, st) if st else None
        # Matin = 0, Après-midi = 1, non planifié = 2 (à la fin).
        half = 0 if label == defense_slots.MORNING else (1 if label == defense_slots.AFTERNOON else 2)
        by_date[jury.defense_date].append(
            (half, st or _time(23, 59), jury.pk, jury)
        )

    for defense_date, rows in by_date.items():
        rows.sort(key=lambda r: (r[0], r[1], r[2]))
        for i, (_, _, _, jury) in enumerate(rows, start=1):
            n = jury.students.count()
            new_name = (
                f"Jury {i} - {n} étudiant{'s' if n > 1 else ''}" if n else f"Jury {i}"
            )
            if jury.name != new_name:
                Jury.objects.filter(pk=jury.pk).update(name=new_name)


# Rétrocompat : anciens appels.
def renumber_draft_juries():
    renumber_all_juries()


def recompact_all_juries():
    """Recompacte les horaires de TOUS les jurys (passages contigus, ordre
    croissant, sans trou). Renvoie le nombre de jurys dont l'horaire a changé."""
    changed = 0
    jury_ids = set(
        DefenseSchedule.objects.values_list("jury_student__jury_id", flat=True)
    )
    for jury in Jury.objects.filter(id__in=jury_ids):
        before = list(
            DefenseSchedule.objects.filter(jury_student__jury=jury)
            .order_by("start_time").values_list("pk", "start_time")
        )
        recompact_jury_schedule(jury)
        after = list(
            DefenseSchedule.objects.filter(jury_student__jury=jury)
            .order_by("start_time").values_list("pk", "start_time")
        )
        if before != after:
            changed += 1
    return changed


@login_required
@role_required(["admin"])
def admin_reorder_all_schedules(request):
    """Bouton : réordonne et resserre les horaires de TOUS les jurys."""
    if request.method == "POST":
        changed = recompact_all_juries()
        messages.success(
            request,
            f"Horaires réordonnés et resserrés pour {changed} jury(s) "
            f"(passages contigus, sans trou)."
        )
    return redirect("admin_jury_list")


def choose_president_for_student(student, members, defense_date):
    candidates = [
        professor for professor in members
        if professor.id != student.encadrant_id
    ]

    candidates.sort(
        key=lambda professor: (
            0 if getattr(professor, "is_priority", False) else 1,
            professor_load_on_date(professor, defense_date),
            professor_total_scheduled_load(professor),
            supervised_students_count(professor),
            professor.full_name.lower(),
        )
    )

    return candidates[0] if candidates else None


def calculate_next_defense_slot_for_jury(jury, members):
    """
    Retourne le prochain start_time disponible de 20 min pour ce jury.
    Utilise end_time du dernier créneau existant, ou cherche le premier créneau libre.
    Retourne None si aucun créneau valide n'est trouvé.
    """
    last_schedule = DefenseSchedule.objects.filter(
        jury_student__jury=jury,
    ).order_by("-start_time").first()

    if last_schedule:
        # end_time peut être vide (créneau créé sans save()) : recalculer.
        next_start = last_schedule.end_time or slot_end_time(
            jury.defense_date,
            last_schedule.start_time,
            last_schedule.duration_minutes or DEFENSE_DURATION_MINUTES,
        )
    else:
        # No students yet: find earliest slot where all members are available
        avails = ProfessorAvailability.objects.filter(
            professor__in=members,
            date=jury.defense_date,
        ).order_by("start_time")

        next_start = None
        for avail in avails:
            candidate = avail.start_time
            if all(
                is_professor_available(m, jury.defense_date, candidate, DEFENSE_DURATION_MINUTES)
                for m in members
            ):
                next_start = candidate
                break

        if next_start is None:
            return None

    # Verify availability and no conflict for all members at next_start
    for member in members:
        if not is_professor_available(member, jury.defense_date, next_start, DEFENSE_DURATION_MINUTES):
            return None
        if professor_has_conflict(member, jury.defense_date, next_start, DEFENSE_DURATION_MINUTES):
            return None

    return next_start


def _professors_free_for_halfday(defense_date, slot_label, exclude_ids=None):
    """Professeurs proposables pour un NOUVEAU jury sur (date, demi-journée) :
    disponibilité déclarée couvrant ce créneau ET pas déjà membre d'un jury
    qui chevauche cette demi-journée (ni l'autre demi-journée du jour, règle
    matin OU après-midi)."""
    exclude_ids = exclude_ids or set()
    start, end = defense_slots.slot_bounds(defense_date, slot_label)

    busy_ids = set(
        DefenseSchedule.objects.filter(
            jury_student__jury__defense_date=defense_date,
            start_time__lt=end,
            end_time__gt=start,
        ).values_list(
            "jury_student__jury__members__professor_id", flat=True
        )
    )
    result = []
    for p in ProfessorProfile.objects.order_by("full_name"):
        if p.id in busy_ids or p.id in exclude_ids:
            continue
        if not is_professor_available(p, defense_date, start):
            continue
        if professor_busy_other_slot(p, defense_date, slot_label):
            continue
        result.append(p)
    return result


def _free_rooms_for_halfday(defense_date, slot_label):
    """Salles libres sur toute la demi-journée (aucun jury n'y chevauche)."""
    start, end = defense_slots.slot_bounds(defense_date, slot_label)
    return [
        s for s in DEFENSE_SALLES
        if not _salle_occupee(defense_date, start, end, s)
    ]


@login_required
@role_required(["admin"])
def admin_jury_add_manual(request):
    """Après la génération automatique : ajouter un (ou deux) jury(s) à la main.
    L'admin choisit une date + une demi-journée, la plateforme propose les
    étudiants sans jury, les professeurs disponibles et non déjà affectés sur
    ce créneau, et les salles libres."""
    today = timezone.localdate()

    # Étape 1 : date + demi-journée.
    date_raw = (request.POST.get("defense_date") or request.GET.get("defense_date") or "").strip()
    slot_label = (request.POST.get("slot") or request.GET.get("slot") or "").strip()

    defense_date = None
    if date_raw:
        try:
            defense_date = date_cls.fromisoformat(date_raw)
        except ValueError:
            defense_date = None
    if slot_label not in (defense_slots.MORNING, defense_slots.AFTERNOON):
        slot_label = ""

    start_time_raw = (request.POST.get("start_time") or request.GET.get("start_time") or "").strip()

    context = {
        "today": today.isoformat(),
        "defense_date": date_raw,
        "slot": slot_label,
        "start_time": start_time_raw,
        "slot_choices": [
            (defense_slots.MORNING, "Matin"),
            (defense_slots.AFTERNOON, "Après-midi"),
        ],
        "step2": False,
    }

    if defense_date and slot_label:
        start, end = defense_slots.slot_bounds(defense_date, slot_label)
        context.update({
            "step2": True,
            "slot_start": start,
            "slot_end": end,
            "students": StudentProfile.objects.filter(
                pfe_request__status=PFERequest.STATUS_ACCEPTED,
                jury_assignment__isnull=True,
            ).select_related("encadrant").order_by("encadrant__full_name", "full_name"),
            "professors": _professors_free_for_halfday(defense_date, slot_label),
            "free_rooms": _free_rooms_for_halfday(defense_date, slot_label),
            "slot_display": "Matin" if slot_label == defense_slots.MORNING else "Après-midi",
        })

    if request.method == "POST" and request.POST.get("action") == "create":
        if not (defense_date and slot_label):
            messages.error(request, "Choisissez une date et une demi-journée.")
            return render(request, "soutenances/admin_jury_add_manual.html", context)

        start, end = defense_slots.slot_bounds(defense_date, slot_label)

        # Horaire de début choisi par l'admin (optionnel). Par défaut : début de
        # la demi-journée. Doit rester dans les bornes du créneau.
        begin = start
        begin_invalid = False
        if start_time_raw:
            try:
                hh, mm = start_time_raw.split(":")
                begin = time(int(hh), int(mm))
            except (ValueError, TypeError):
                begin_invalid = True
        if not begin_invalid and not (start <= begin < end):
            begin_invalid = True

        student_ids = [int(x) for x in request.POST.getlist("students") if x.isdigit()]
        prof_ids = [int(x) for x in request.POST.getlist("professors") if x.isdigit()]
        salle = request.POST.get("salle") or ""

        students = list(
            StudentProfile.objects.filter(
                id__in=student_ids,
                pfe_request__status=PFERequest.STATUS_ACCEPTED,
                jury_assignment__isnull=True,
            ).select_related("encadrant")
        )
        members = list(ProfessorProfile.objects.filter(id__in=prof_ids))

        # Revérification serveur (les listes ont pu changer entre-temps).
        free_now_ids = {p.id for p in _professors_free_for_halfday(defense_date, slot_label)}
        invalid = [p for p in members if p.id not in free_now_ids]

        if len(members) != 3:
            messages.error(request, "Sélectionnez exactement 3 professeurs.")
        elif invalid:
            messages.error(
                request,
                "Ces professeurs ne sont plus disponibles/libres sur ce créneau : "
                + ", ".join(p.full_name for p in invalid)
            )
        elif not students:
            messages.error(request, "Sélectionnez au moins un étudiant.")
        elif begin_invalid:
            messages.error(
                request,
                f"L'horaire de début doit être compris entre "
                f"{start.strftime('%H:%M')} et {end.strftime('%H:%M')}."
            )
        elif not salle or _salle_occupee(defense_date, begin, end, salle):
            messages.error(request, "Choisissez une salle libre sur ce créneau.")
        else:
            # Un créneau de 20 min par étudiant, à partir de l'horaire de début.
            # TOUS les étudiants sélectionnés passent, même au-delà de la fin de
            # la demi-journée (décision de l'admin).
            placed = students
            start_times = []
            cur = datetime.combine(defense_date, begin)
            for _ in placed:
                start_times.append(cur.time())
                cur += timedelta(minutes=DEFENSE_DURATION_MINUTES)
            last_end = cur.time()  # fin du dernier passage

            president_pool = members
            warnings = []
            with transaction.atomic():
                jury = Jury.objects.create(
                    name=build_grouped_jury_name(placed, defense_date),
                    defense_date=defense_date,
                    salle=salle,
                    is_validated=False,
                )
                for p in members:
                    JuryMember.objects.create(jury=jury, professor=p)
                schedules = []
                for student, t in zip(placed, start_times):
                    enc_in = any(m.id == student.encadrant_id for m in members)
                    president = choose_president_for_student(
                        student=student, members=president_pool,
                        defense_date=defense_date,
                    )
                    js = JuryStudent.objects.create(
                        jury=jury, student=student, president=president,
                        encadrant_absent=not enc_in,
                    )
                    schedules.append(DefenseSchedule(
                        jury_student=js, start_time=t,
                        end_time=slot_end_time(defense_date, t, DEFENSE_DURATION_MINUTES),
                        duration_minutes=DEFENSE_DURATION_MINUTES,
                    ))
                    if not enc_in:
                        enc = student.encadrant.full_name if student.encadrant else "?"
                        warnings.append(f"{student.full_name} (encadrant {enc})")
                # bulk_create : ne rejoue pas la validation de disponibilité des
                # membres — autorise les passages au-delà de la fin de créneau.
                DefenseSchedule.objects.bulk_create(schedules)
                recompact_jury_schedule(jury)
                refresh_jury_name_count(jury)
                renumber_all_juries()
                jury.refresh_from_db()

            messages.success(
                request,
                f"Jury créé : {jury.name} — {defense_date.strftime('%d/%m/%Y')} "
                f"({context.get('slot_display', slot_label)}) à partir de "
                f"{begin.strftime('%H:%M')} en {salle}, {len(placed)} étudiant(s)."
            )
            if warnings:
                messages.warning(
                    request,
                    "Encadrant hors jury (marqué « encadrant absent ») pour : "
                    + "; ".join(warnings)
                )
            if last_end > end:
                messages.warning(
                    request,
                    f"Certains passages dépassent la fin de la demi-journée "
                    f"({end.strftime('%H:%M')}) — tous les étudiants ont été placés."
                )
            return redirect("admin_jury_detail", pk=jury.pk)

    return render(request, "soutenances/admin_jury_add_manual.html", context)


@login_required
@role_required(["admin"])
def admin_jury_create(request):
    if request.method == "POST":
        form = JuryForm(request.POST)

        if form.is_valid():
            jury = save_jury_with_members(form)
            renumber_all_juries()
            messages.success(request, "Le jury a été créé avec succès.")
            return redirect("admin_jury_detail", pk=jury.pk)
    else:
        form = JuryForm()

    eligible_students = StudentProfile.objects.filter(
        pfe_request__status=PFERequest.STATUS_ACCEPTED,
        jury_assignment__isnull=True,
    ).select_related("encadrant").order_by("encadrant__full_name", "full_name")

    return render(request, "soutenances/admin_jury_form.html", {
        "form": form,
        "title": "Créer un jury",
        "eligible_students": eligible_students,
    })


@login_required
@role_required(["admin"])
def admin_jury_quick_create(request):
    """Flux guidé et réellement contraint par les disponibilités :
    étudiant -> encadrant -> créneau réel -> membres réellement disponibles
    à ce créneau -> création atomique de Jury + JuryMember + JuryStudent +
    DefenseSchedule. Le filtrage n'est pas qu'une suggestion JS : le
    queryset du formulaire de membres est restreint côté serveur, et tout
    est revérifié juste avant la création."""

    eligible_students = StudentProfile.objects.filter(
        pfe_request__status=PFERequest.STATUS_ACCEPTED,
        jury_assignment__isnull=True,
    ).select_related("encadrant").order_by("encadrant__full_name", "full_name")

    context = {"eligible_students": eligible_students}

    student_id = request.POST.get("student_id") or request.GET.get("student_id")
    student = None

    if student_id:
        student = get_object_or_404(
            StudentProfile.objects.select_related("encadrant"),
            pk=student_id,
        )
        context["student"] = student

        if not student.encadrant_id:
            messages.error(request, "Cet étudiant n'a pas d'encadrant défini.")
            return render(request, "soutenances/admin_jury_quick_create.html", context)

    slot_key = request.POST.get("slot") or request.GET.get("slot")
    selected_date = None
    selected_time = None

    if student:
        slots = get_common_available_slots([student.encadrant])

        context["slots"] = [
            {
                "key": f"{defense_date.isoformat()}|{start_time.strftime('%H:%M')}",
                "date": defense_date,
                "start_time": start_time,
            }
            for defense_date, start_time in slots
        ]

        if not slots:
            context["no_slots"] = True

    if slot_key and student:
        try:
            date_raw, time_raw = slot_key.split("|")
            selected_date = datetime.strptime(date_raw, "%Y-%m-%d").date()
            selected_time = datetime.strptime(time_raw, "%H:%M").time()
        except ValueError:
            selected_date = None
            selected_time = None

    members_form = None
    other_available = None

    if student and selected_date and selected_time:
        # Revérification serveur : le créneau doit toujours être valide pour
        # l'encadrant au moment où on l'utilise, pas seulement quand la liste
        # a été affichée.
        encadrant_still_ok = (
            is_professor_available(student.encadrant, selected_date, selected_time)
            and not professor_has_conflict(student.encadrant, selected_date, selected_time)
        )

        if not encadrant_still_ok:
            messages.error(
                request,
                "Ce créneau n'est plus disponible pour l'encadrant. Choisissez-en un autre."
            )
            selected_date = None
            selected_time = None
        else:
            available_professors = get_available_professors_at_slot(selected_date, selected_time)
            other_available = ProfessorProfile.objects.filter(
                id__in=[p.id for p in available_professors]
            ).exclude(id=student.encadrant_id).order_by("full_name")

            context["selected_date"] = selected_date
            context["selected_time"] = selected_time
            context["slot_key"] = slot_key
            context["available_count"] = other_available.count()

            if other_available.count() < 2:
                context["not_enough_members"] = True
            else:
                if request.method == "POST" and request.POST.get("step") == "3":
                    members_form = JuryMembersForSlotForm(
                        request.POST, available_queryset=other_available
                    )
                else:
                    members_form = JuryMembersForSlotForm(available_queryset=other_available)

            context["members_form"] = members_form

    if (
        request.method == "POST"
        and request.POST.get("step") == "3"
        and members_form is not None
        and student
        and selected_date
        and selected_time
    ):
        if members_form.is_valid():
            try:
                with transaction.atomic():
                    if not (
                        is_professor_available(student.encadrant, selected_date, selected_time)
                        and not professor_has_conflict(student.encadrant, selected_date, selected_time)
                    ):
                        raise ValidationError(
                            "Le créneau n'est plus disponible pour l'encadrant."
                        )

                    chosen_others = list(members_form.cleaned_data["members"])

                    for professor in chosen_others:
                        if not (
                            is_professor_available(professor, selected_date, selected_time)
                            and not professor_has_conflict(professor, selected_date, selected_time)
                        ):
                            raise ValidationError(
                                f"{professor.full_name} n'est plus disponible à ce créneau."
                            )

                    jury = Jury.objects.create(
                        name=f"Jury {student.full_name} - {selected_date.strftime('%d/%m/%Y')}",
                        defense_date=selected_date,
                        is_validated=False,
                    )

                    members = [student.encadrant] + chosen_others

                    for professor in members:
                        JuryMember.objects.create(jury=jury, professor=professor)

                    president = choose_president_for_student(
                        student=student,
                        members=members,
                        defense_date=selected_date,
                    )

                    jury_student = JuryStudent.objects.create(
                        jury=jury,
                        student=student,
                        president=president,
                    )

                    DefenseSchedule.objects.create(
                        jury_student=jury_student,
                        start_time=selected_time,
                        duration_minutes=DEFENSE_DURATION_MINUTES,
                    )
            except ValidationError as exc:
                messages.error(
                    request,
                    "; ".join(exc.messages) if hasattr(exc, "messages") else str(exc)
                )
            else:
                messages.success(
                    request,
                    f"Jury créé pour {student.full_name}, le "
                    f"{selected_date.strftime('%d/%m/%Y')} à {selected_time.strftime('%H:%M')}."
                )
                return redirect("admin_jury_detail", pk=jury.pk)

    return render(request, "soutenances/admin_jury_quick_create.html", context)


def _timetable_data():
    """Programme (emploi du temps) : jurys planifiés groupés par jour, triés
    matin puis après-midi, avec pour chaque jury son numéro, sa salle, ses
    membres, le nombre d'étudiants et le détail des passages (heure, nom,
    matricule, président, encadrant)."""
    from collections import defaultdict

    juries = Jury.objects.prefetch_related(
        "members__professor",
        "students__student__encadrant",
        "students__president",
        "students__schedule",
    )
    by_day = defaultdict(list)
    for j in juries:
        rows = []
        for js in j.students.all():
            sch = getattr(js, "schedule", None)
            if not sch:
                continue
            rows.append({
                "start": sch.start_time,
                "end": sch.end_time or slot_end_time(j.defense_date, sch.start_time),
                "name": js.student.full_name or "(nom absent)",
                "matricule": js.student.matricule,
                "president": js.president.full_name if js.president else "—",
                "encadrant": js.student.encadrant.full_name if js.student.encadrant else "—",
            })
        if not rows:
            continue
        rows.sort(key=lambda r: r["start"])
        starts = [r["start"] for r in rows]
        ends = [r["end"] for r in rows]
        by_day[j.defense_date].append({
            "jury": j,
            "start": min(starts),
            "end": max(ends),
            "members": [m.professor.full_name for m in j.members.all()],
            "count": len(rows),
            "students": rows,
        })
    for d in by_day:
        by_day[d].sort(key=lambda e: (
            0 if _slot_label_at(d, e["start"]) == defense_slots.MORNING else 1,
            e["start"],
        ))
    return sorted(by_day.items())


@login_required
@role_required(["admin"])
def admin_timetable(request):
    """Emploi du temps résumé (par jour, créneaux 9h–19h) + export PDF/Excel
    avec le détail de chaque jury (horaires, nom, matricule, président,
    encadrant, membres)."""
    days = _timetable_data()
    all_days = [d for d, _ in days]

    # Filtre par date (optionnel).
    day_raw = (request.GET.get("day") or "").strip()
    selected_day = None
    if day_raw:
        try:
            selected_day = date_cls.fromisoformat(day_raw)
            days = [(d, e) for d, e in days if d == selected_day]
        except ValueError:
            selected_day = None

    fmt = (request.GET.get("format") or "").strip()

    def fmt_t(t):
        return t.strftime("%H:%M") if hasattr(t, "strftime") else str(t)

    def fmt_d(d):
        return d.strftime("%A %d/%m/%Y") if hasattr(d, "strftime") else str(d)

    if fmt == "xlsx":
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Résumé"
        ws.append(["Institut Supérieur de Génie Industriel (ISGI) — Département de l'IUP"])
        ws.append(["Emploi du temps des soutenances"])
        ws.append([])
        for d, entries in days:
            ws.append([fmt_d(d)])
            ws.append(["Horaire", "Jury", "Salle", "Nb étudiants", "Membres"])
            for e in entries:
                ws.append([
                    f"{fmt_t(e['start'])}-{fmt_t(e['end'])}",
                    e["jury"].name,
                    e["jury"].get_salle_display() if e["jury"].salle else "",
                    e["count"],
                    " / ".join(e["members"]),
                ])
            ws.append([])
        wd = wb.create_sheet("Détail")
        wd.append(["Jour", "Jury", "Salle", "Horaire", "Nom & Prénom", "Matricule", "Président", "Encadrant"])
        for d, entries in days:
            for e in entries:
                for r in e["students"]:
                    wd.append([
                        fmt_d(d), e["jury"].name,
                        e["jury"].get_salle_display() if e["jury"].salle else "",
                        f"{fmt_t(r['start'])}-{fmt_t(r['end'])}",
                        r["name"], r["matricule"], r["president"], r["encadrant"],
                    ])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="emploi_du_temps.xlsx"'
        wb.save(response)
        return response

    if fmt == "pdf":
        lines = [
            "Institut Supérieur de Génie Industriel (ISGI)",
            "Département de l'IUP",
            "", "Emploi du temps des soutenances", "",
        ]
        for d, entries in days:
            lines.append(f"=== {fmt_d(d)} ===")
            for e in entries:
                lines.append(
                    f"{e['jury'].name} · {fmt_t(e['start'])}-{fmt_t(e['end'])}"
                    + (f" · {e['jury'].get_salle_display()}" if e['jury'].salle else "")
                    + f" · {e['count']} etudiant(s)"
                )
                lines.append("   Membres : " + " / ".join(e["members"]))
                for r in e["students"]:
                    lines.append(
                        f"   {fmt_t(r['start'])}-{fmt_t(r['end'])}  {r['matricule']}  "
                        f"{r['name']}  (Pres. {r['president']} · Enc. {r['encadrant']})"
                    )
                lines.append("")
        if not days:
            lines.append("Aucune soutenance planifiee.")
        return simple_pdf_response("Emploi du temps", lines, "emploi_du_temps.pdf")

    return render(request, "soutenances/admin_timetable.html", {
        "days": days,
        "total_juries": sum(len(e) for _, e in days),
        "available_days": all_days,
        "selected_day": selected_day,
    })


@login_required
@role_required(["admin"])
def admin_jury_export(request, pk):
    """Télécharge le détail d'UN jury (salle, membres, étudiants avec matricule,
    horaires de passage, président, encadrant) en PDF ou Word."""
    jury = get_object_or_404(
        Jury.objects.prefetch_related(
            "members__professor", "students__student__encadrant",
            "students__president", "students__schedule",
        ), pk=pk,
    )
    rows = []
    for js in jury.students.all():
        sch = getattr(js, "schedule", None)
        rows.append({
            "start": sch.start_time if sch else None,
            "end": (sch.end_time or slot_end_time(jury.defense_date, sch.start_time)) if sch else None,
            "name": js.student.full_name or "(nom absent)",
            "matricule": js.student.matricule,
            "president": js.president.full_name if js.president else "—",
            "encadrant": js.student.encadrant.full_name if js.student.encadrant else "—",
        })
    rows.sort(key=lambda r: (r["start"] is None, r["start"] or time(23, 59)))
    members = [m.professor.full_name for m in jury.members.all()]
    salle = jury.get_salle_display() if jury.salle else "—"
    d = jury.defense_date.strftime("%d/%m/%Y") if jury.defense_date else "—"

    def ft(t):
        return t.strftime("%H:%M") if t else "—"

    fmt = (request.GET.get("format") or "pdf").strip()

    if fmt == "word":
        body = [
            f"<p><b>Date :</b> {d} &nbsp; <b>Salle :</b> {salle}<br>"
            f"<b>Membres :</b> {' / '.join(members)}</p>",
            "<table border='1' cellspacing='0' cellpadding='4'>"
            "<tr><th>Horaire</th><th>Matricule</th><th>Nom & Prénom</th>"
            "<th>Président</th><th>Encadrant</th></tr>",
        ]
        for r in rows:
            body.append(
                f"<tr><td>{ft(r['start'])}–{ft(r['end'])}</td><td>{r['matricule']}</td>"
                f"<td>{r['name']}</td><td>{r['president']}</td><td>{r['encadrant']}</td></tr>"
            )
        body.append("</table>")
        return _word_response_soutenances(
            f"{jury.name} — {d}", "".join(body),
            f"jury_{jury.pk}.doc",
        )

    lines = [
        "Institut Supérieur de Génie Industriel (ISGI)", "Département de l'IUP", "",
        f"{jury.name}", f"Date : {d}    Salle : {salle}",
        "Membres : " + " / ".join(members), "",
        "Passages :",
    ]
    for r in rows:
        lines.append(
            f"  {ft(r['start'])}-{ft(r['end'])}  {r['matricule']}  {r['name']}  "
            f"(Pres. {r['president']} · Enc. {r['encadrant']})"
        )
    return simple_pdf_response(jury.name, lines, f"jury_{jury.pk}.pdf")


def _word_response_soutenances(title, body_html, filename):
    html = (
        "<html><head><meta charset='utf-8'></head><body>"
        "<div style='text-align:center;'>"
        "<h2>Institut Supérieur de Génie Industriel (ISGI)</h2>"
        "<p><b>Département de l'IUP</b></p>"
        f"<h3>{title}</h3></div>{body_html}</body></html>"
    )
    resp = HttpResponse(html, content_type="application/msword")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@login_required
@role_required(["admin"])
def admin_jury_conflicts(request):
    """Outil global : détecte, pour TOUS les professeurs, les jurys dont les
    créneaux se chevauchent (membre doublement réservé). Affiche les jurys
    concernés, sans avoir à ouvrir chaque prof un par un."""
    from collections import defaultdict

    # Bloc horaire de chaque jury (min début → max fin) + date.
    sched = defaultdict(list)
    for s in DefenseSchedule.objects.values(
        "jury_student__jury_id", "start_time", "end_time"
    ):
        sched[s["jury_student__jury_id"]].append(
            (s["start_time"], s["end_time"] or s["start_time"])
        )
    jury_by_id = {j.id: j for j in Jury.objects.all()}
    blocks = {}
    for jid, lst in sched.items():
        jury = jury_by_id.get(jid)
        if not jury or not jury.defense_date:
            continue
        blocks[jid] = {
            "date": jury.defense_date,
            "start": min(a for a, _ in lst),
            "end": max(b for _, b in lst),
            "jury": jury,
        }

    prof_juries = defaultdict(list)
    for jm in JuryMember.objects.select_related("professor"):
        if jm.jury_id in blocks:
            prof_juries[jm.professor].append(jm.jury_id)

    conflicts = []
    for prof, jids in prof_juries.items():
        items = sorted(
            (blocks[j] for j in jids), key=lambda b: (b["date"], b["start"])
        )
        for i in range(len(items)):
            a = items[i]
            for k in range(i + 1, len(items)):
                b = items[k]
                if a["date"] != b["date"]:
                    break  # trié par date : plus rien ne peut chevaucher
                if b["start"] < a["end"]:  # chevauchement
                    conflicts.append({
                        "professor": prof,
                        "date": a["date"],
                        "jury_a": a["jury"], "a_start": a["start"], "a_end": a["end"],
                        "jury_b": b["jury"], "b_start": b["start"], "b_end": b["end"],
                    })

    conflicts.sort(key=lambda c: (c["date"], c["professor"].full_name.lower()))
    return render(request, "soutenances/admin_jury_conflicts.html", {
        "conflicts": conflicts,
        "professors_checked": len(prof_juries),
    })


def _mergeable_juries_for(jury, current_members):
    """Autres jurys du MÊME JOUR. Sépare ceux qui ont EXACTEMENT les mêmes
    membres (fusionnables) de ceux dont les membres diffèrent (non fusionnables,
    affichés pour expliquer)."""
    member_ids = set(m.professor_id for m in current_members)
    my_names = {m.professor.full_name for m in current_members}
    matching, others = [], []
    for other in Jury.objects.filter(
        defense_date=jury.defense_date
    ).exclude(pk=jury.pk).prefetch_related("members__professor"):
        ids = set(m.professor_id for m in other.members.all())
        first = DefenseSchedule.objects.filter(
            jury_student__jury=other
        ).order_by("start_time").first()
        row = {
            "jury": other,
            "start": first.start_time if first else None,
            "students": other.students.count(),
        }
        if ids == member_ids:
            matching.append(row)
        else:
            other_names = {m.professor.full_name for m in other.members.all()}
            row["diff"] = ", ".join(sorted(other_names ^ my_names)) or "composition différente"
            others.append(row)
    matching.sort(key=lambda r: (r["start"] is None, r["start"] or time(0, 0)))
    others.sort(key=lambda r: (r["start"] is None, r["start"] or time(0, 0)))
    return {"matching": matching, "others": others}


@login_required
@role_required(["admin"])
def admin_jury_update(request, pk):
    """
    Page Modifier jury — simplifiée.
    Le créneau réel est auto-détecté depuis les DefenseSchedule existants.
    Plus de sélection manuelle de créneau.
    """
    jury = get_object_or_404(
        Jury.objects.prefetch_related("members__professor"),
        pk=pk,
    )

    jury_students = list(
        jury.students.select_related("student__encadrant", "schedule")
        .order_by("schedule__start_time", "id")
    )

    # ── 1. Détection automatique du créneau réel (premier slot planifié) ──
    # On prend le DefenseSchedule avec le plus petit start_time du jury.
    first_schedule = (
        DefenseSchedule.objects.filter(jury_student__jury=jury)
        .order_by("start_time")
        .first()
    )
    real_slot_date  = jury.defense_date if first_schedule else None
    real_slot_start = first_schedule.start_time if first_schedule else None
    real_slot_end   = first_schedule.end_time   if first_schedule else None
    has_real_slot   = real_slot_date is not None and real_slot_start is not None

    # ── 2. Membres actuels avec logique can_remove ─────────────────────────
    current_members = list(
        jury.members.select_related("professor").order_by("professor__full_name")
    )

    # Rôles : expert (de la filière des étudiants du jury) et président.
    student_filieres = {js.student.filiere for js in jury_students if js.student.filiere}
    expert_ids = set(
        FiliereExpert.objects.filter(filiere__in=student_filieres)
        .values_list("professor_id", flat=True)
    ) if student_filieres else set()
    president_ids = {js.president_id for js in jury_students if js.president_id}

    member_rows = []

    for member in current_members:
        professor = member.professor
        supervised_here = [
            js.student for js in jury_students
            if js.student.encadrant_id == professor.id
        ]

        is_available = None
        if has_real_slot:
            is_available = (
                is_professor_available(professor, real_slot_date, real_slot_start)
                and not professor_has_conflict(professor, real_slot_date, real_slot_start)
            )

        # Bloquer le retrait si ce professeur est l'encadrant d'étudiants dans ce jury
        # (chaque étudiant n'a qu'un seul encadrant → retrait toujours bloqué si encadrant actif)
        can_remove = len(supervised_here) == 0
        cannot_remove_reason = (
            "Impossible de retirer ce professeur : il est le seul encadrant "
            f"des étudiants affectés ({', '.join(s.full_name for s in supervised_here)}). "
            "Utilisez « Supprimer le jury » si nécessaire."
            if not can_remove else None
        )

        member_rows.append({
            "member": member,
            "professor": professor,
            "supervised_here": supervised_here,
            "is_available": is_available,
            "can_remove": can_remove,
            "cannot_remove_reason": cannot_remove_reason,
            "is_encadrant": len(supervised_here) > 0,
            "is_expert": professor.id in expert_ids,
            "is_president": professor.id in president_ids,
            "is_priority": professor.is_priority,
        })

    # ── 3. Formulaire d'ajout filtré sur le créneau réel ──────────────────
    add_member_form = None
    addable_count = 0

    if has_real_slot and len(current_members) < 3:
        available_at_slot = get_available_professors_at_slot(real_slot_date, real_slot_start)
        current_member_ids = {m.professor_id for m in current_members}
        addable_qs = ProfessorProfile.objects.filter(
            id__in=[p.id for p in available_at_slot if p.id not in current_member_ids]
        ).order_by("full_name")
        addable_count = addable_qs.count()

        if request.method == "POST" and request.POST.get("action") == "add_member":
            add_member_form = JuryAddMemberForm(request.POST, selectable_queryset=addable_qs)
        else:
            add_member_form = JuryAddMemberForm(selectable_queryset=addable_qs)

    # ── 3bis. Candidats au remplacement d'un membre (ex. encadrant indisponible) ──
    # TOUS les professeurs hors jury sont proposés ; la disponibilité au
    # créneau réel est indiquée dans la liste. Le remplacement par un
    # professeur sans disponibilité déclarée reste possible (avertissement).
    current_member_ids = {m.professor_id for m in current_members}
    replacement_candidates = []
    for p in ProfessorProfile.objects.order_by("full_name"):
        if p.id in current_member_ids:
            continue
        status = professor_slot_status(p, real_slot_date, real_slot_start) if has_real_slot else {"free": None, "label": ""}
        replacement_candidates.append({
            "id": p.id,
            "full_name": p.full_name,
            "available": status["free"],
            "status": status["label"],
        })
    # Les libres d'abord, puis "libre l'autre demi-journée", puis les autres.
    replacement_candidates.sort(
        key=lambda r: (0 if r["available"] else 1, r["full_name"].lower())
    )

    # ── 3ter. Candidats à l'ÉCHANGE : membres des AUTRES jurys (numéro de
    #     jury affiché ; possible même si les jurys sont publiés).
    current_member_ids_all = {m.professor_id for m in current_members}
    swap_candidates = []
    for jm in JuryMember.objects.exclude(jury=jury).select_related(
        "professor", "jury"
    ).order_by("jury__defense_date", "jury__name", "professor__full_name"):
        if jm.professor_id in current_member_ids_all:
            continue
        swap_candidates.append({
            "value": f"{jm.jury_id}:{jm.professor_id}",
            "label": (
                f"{jm.professor.full_name} — {jm.jury.name} "
                f"({jm.jury.defense_date.strftime('%d/%m/%Y')})"
                f"{' · publié' if jm.jury.is_validated else ''}"
            ),
        })

    context = {
        "jury": jury,
        "jury_students": jury_students,
        "member_rows": member_rows,
        "current_members_count": len(current_members),
        "real_slot_date": real_slot_date,
        "real_slot_start": real_slot_start,
        "real_slot_end": real_slot_end,
        "has_real_slot": has_real_slot,
        "add_member_form": add_member_form,
        "addable_count": addable_count,
        "replacement_candidates": replacement_candidates,
        "swap_candidates": swap_candidates,
        # Jurys cibles pour déplacer un étudiant (tous sauf celui-ci).
        "move_target_juries": Jury.objects.exclude(pk=jury.pk).order_by(
            "defense_date", "name"
        ),
        # Créneaux proposés pour reprogrammer (dispo commune des 3 membres),
        # dimensionnés sur les étudiants NON notés (les notés ne bougent pas).
        "reschedule_options": reschedule_options_for_jury(
            jury,
            [m.professor for m in current_members],
            sum(
                1 for js in jury_students
                if not (_has_real_grades(js))
            ),
        ) if len(current_members) >= 3 else [],
        "graded_count": sum(
            1 for js in jury_students
            if _has_real_grades(js)
        ),
        # Jurys fusionnables : MÊMES membres + MÊME date (pour coller deux
        # jurys du même jury et supprimer le vide).
        "mergeable_juries": _mergeable_juries_for(jury, current_members),
    }

    # ── 4. POST : retirer un membre ────────────────────────────────────────
    if request.method == "POST" and request.POST.get("action") == "remove_member":
        try:
            professor_id = int(request.POST.get("professor_id", ""))
        except ValueError:
            professor_id = None

        member_to_remove = (
            jury.members.filter(professor_id=professor_id).first()
            if professor_id is not None else None
        )

        if not member_to_remove:
            messages.error(request, "Ce professeur n'est pas membre de ce jury.")
        else:
            supervised = [
                js for js in jury_students
                if js.student.encadrant_id == professor_id
            ]
            if supervised:
                names = ", ".join(js.student.full_name for js in supervised)
                messages.error(
                    request,
                    f"Impossible de retirer ce professeur : il est le seul encadrant "
                    f"des étudiants affectés à ce jury ({names}). "
                    "Utilisez « Supprimer le jury » si le jury entier doit être dissous."
                )
            else:
                prof_name = member_to_remove.professor.full_name

                # Vérifier si ce professeur est président d'une soutenance dans ce jury
                presided = [js for js in jury_students if js.president_id == professor_id]
                if presided:
                    # Chercher un remplaçant parmi les membres restants
                    # (doit être membre du jury et différent de l'encadrant de chaque étudiant)
                    remaining = [
                        m.professor for m in current_members
                        if m.professor_id != professor_id
                    ]
                    replacements = {}  # jury_student.pk -> nouveau président
                    blocked = False

                    for js in presided:
                        new_pres = next(
                            (p for p in remaining if p.id != js.student.encadrant_id),
                            None,
                        )
                        if new_pres is None:
                            blocked = True
                            messages.error(
                                request,
                                f"Impossible de retirer {prof_name} : il est président de la "
                                f"soutenance de {js.student.full_name} et aucun autre membre "
                                "valide ne peut le remplacer comme président."
                            )
                            break
                        replacements[js.pk] = new_pres

                    if not blocked:
                        # Appliquer les remplacements via update() pour éviter full_clean
                        # (le membre n'est pas encore retiré, la cohérence est préservée)
                        for js in presided:
                            JuryStudent.objects.filter(pk=js.pk).update(
                                president=replacements[js.pk]
                            )
                        member_to_remove.delete()
                        new_pres_name = next(iter(replacements.values())).full_name
                        presided_names = ", ".join(js.student.full_name for js in presided)
                        messages.success(
                            request,
                            f"{prof_name} retiré. Président remplacé par {new_pres_name} "
                            f"pour : {presided_names}."
                        )
                else:
                    member_to_remove.delete()
                    messages.success(request, f"{prof_name} retiré du jury.")

        return redirect("admin_jury_update", pk=jury.pk)

    # ── 5. POST : ajouter un membre (avec vérification backend) ───────────
    if (
        request.method == "POST"
        and request.POST.get("action") == "add_member"
        and add_member_form is not None
    ):
        if add_member_form.is_valid():
            professor = add_member_form.cleaned_data["professor"]

            # Vérification backend : disponible ET sans conflit au créneau réel
            still_ok = has_real_slot and (
                is_professor_available(professor, real_slot_date, real_slot_start)
                and not professor_has_conflict(professor, real_slot_date, real_slot_start)
            )

            if not still_ok:
                messages.error(
                    request,
                    f"{professor.full_name} n'est pas disponible au créneau réel de ce jury "
                    f"({real_slot_date} à {real_slot_start})."
                )
            elif jury.members.count() >= 3:
                messages.error(request, "Ce jury contient déjà 3 professeurs.")
            else:
                try:
                    JuryMember.objects.create(jury=jury, professor=professor)
                    messages.success(request, f"{professor.full_name} ajouté au jury.")
                except ValidationError as exc:
                    messages.error(request, "; ".join(exc.messages))
        else:
            messages.error(request, "Sélection de professeur invalide ou indisponible.")

        return redirect("admin_jury_update", pk=jury.pk)

    # ── 6. POST : remplacer un membre (ex. encadrant devenu indisponible) ─────
    if request.method == "POST" and request.POST.get("action") == "replace_member":
        try:
            old_id = int(request.POST.get("old_professor_id", ""))
            new_id = int(request.POST.get("new_professor_id", ""))
        except (ValueError, TypeError):
            messages.error(request, "Sélection invalide.")
            return redirect("admin_jury_update", pk=jury.pk)

        old_member = jury.members.filter(professor_id=old_id).first()
        new_prof = ProfessorProfile.objects.filter(id=new_id).first()

        # Conflit dur : le nouveau prof est-il déjà dans un AUTRE jury sur un
        # créneau de CE jury ? (évite le double-booking sur des jurys qui se
        # chevauchent).
        conflict_here = False
        if new_prof and has_real_slot:
            my_starts = DefenseSchedule.objects.filter(
                jury_student__jury=jury
            ).values_list("start_time", flat=True)
            for st in my_starts:
                if professor_has_conflict(new_prof, real_slot_date, st):
                    conflict_here = True
                    break

        if not old_member:
            messages.error(request, "Le membre à remplacer n'appartient pas à ce jury.")
        elif not new_prof:
            messages.error(request, "Choisissez un professeur remplaçant.")
        elif jury.members.filter(professor_id=new_id).exists():
            messages.error(request, f"{new_prof.full_name} est déjà membre de ce jury.")
        elif conflict_here:
            messages.error(
                request,
                f"Impossible : {new_prof.full_name} est déjà membre d'un autre jury "
                f"sur un créneau de ce jury (chevauchement). Choisissez un prof libre "
                f"ou reprogrammez l'un des jurys."
            )
        else:
            old_name = old_member.professor.full_name
            with transaction.atomic():
                # Reprendre la présidence des soutenances présidées par l'ancien membre
                JuryStudent.objects.filter(
                    jury=jury, president_id=old_id
                ).update(president=new_prof)
                # Retirer l'ancien AVANT d'ajouter le nouveau (max 3 membres)
                old_member.delete()
                JuryMember.objects.create(jury=jury, professor=new_prof)
            messages.success(
                request,
                f"{old_name} a été remplacé par {new_prof.full_name} dans ce jury."
            )
            # Le remplacement par un professeur sans disponibilité déclarée
            # reste possible : simple avertissement (décision de l'admin).
            if has_real_slot and not (
                is_professor_available(new_prof, real_slot_date, real_slot_start)
                and not professor_has_conflict(new_prof, real_slot_date, real_slot_start)
            ):
                messages.warning(
                    request,
                    f"Attention : {new_prof.full_name} n'a pas de disponibilité "
                    f"déclarée au créneau de ce jury "
                    f"({real_slot_date.strftime('%d/%m/%Y')} à "
                    f"{real_slot_start.strftime('%H:%M')})."
                )
        return redirect("admin_jury_update", pk=jury.pk)

    # ── 6bis. POST : ÉCHANGER un membre avec un membre d'un AUTRE jury ──────
    #     (possible même si les jurys sont déjà publiés).
    if request.method == "POST" and request.POST.get("action") == "swap_member":
        try:
            old_id = int(request.POST.get("old_professor_id", ""))
            target_jury_id, target_prof_id = (
                request.POST.get("swap_target", "").split(":")
            )
            target_jury_id = int(target_jury_id)
            target_prof_id = int(target_prof_id)
        except (ValueError, TypeError):
            messages.error(request, "Sélection d'échange invalide.")
            return redirect("admin_jury_update", pk=jury.pk)

        old_member = jury.members.select_related("professor").filter(
            professor_id=old_id
        ).first()
        other_jury = Jury.objects.filter(pk=target_jury_id).first()
        other_member = (
            other_jury.members.select_related("professor").filter(
                professor_id=target_prof_id
            ).first() if other_jury else None
        )

        def _encadres_dans(prof_id, jury_obj):
            """Étudiants de jury_obj encadrés par prof_id qui comptaient sur sa
            présence (encadrant_absent=False)."""
            return list(
                JuryStudent.objects.filter(
                    jury=jury_obj, student__encadrant_id=prof_id,
                    encadrant_absent=False,
                ).select_related("student")
            )

        if not old_member or not other_jury or not other_member:
            messages.error(request, "Membre ou jury cible introuvable.")
        elif other_jury.pk == jury.pk:
            messages.error(request, "Choisissez un membre d'un AUTRE jury.")
        elif jury.members.filter(professor_id=target_prof_id).exists():
            messages.error(
                request,
                f"{other_member.professor.full_name} est déjà membre de ce jury."
            )
        elif other_jury.members.filter(professor_id=old_id).exists():
            messages.error(
                request,
                f"{old_member.professor.full_name} est déjà membre du jury cible."
            )
        else:
            prof_a = old_member.professor      # part vers l'autre jury
            prof_b = other_member.professor    # arrive dans ce jury

            # Étudiants qui vont soutenir SANS leur encadrant après l'échange
            # (leur encadrant quitte leur jury). Collectés AVANT le swap.
            affected = (
                [(prof_a, js) for js in _encadres_dans(old_id, jury)]
                + [(prof_b, js) for js in _encadres_dans(target_prof_id, other_jury)]
            )

            with transaction.atomic():
                # Retirer les deux, puis recréer croisés (max 3 respecté).
                old_member.delete()
                other_member.delete()
                JuryMember.objects.create(jury=jury, professor=prof_b)
                JuryMember.objects.create(jury=other_jury, professor=prof_a)

                # Reprendre les présidences : l'arrivant reprend celles du
                # partant (sauf pour ses propres encadrés → autre membre).
                def _reassign(jury_obj, leaving_id, incoming):
                    members_now = [
                        m.professor for m in
                        jury_obj.members.select_related("professor").all()
                    ]
                    for js in JuryStudent.objects.filter(
                        jury=jury_obj, president_id=leaving_id
                    ).select_related("student"):
                        if js.student.encadrant_id != incoming.id:
                            new_president = incoming
                        else:
                            new_president = next(
                                (m for m in members_now
                                 if m.id != js.student.encadrant_id),
                                None,
                            )
                        JuryStudent.objects.filter(pk=js.pk).update(
                            president=new_president
                        )

                _reassign(jury, old_id, prof_b)
                _reassign(other_jury, target_prof_id, prof_a)

                # Recalcule « encadrant absent » selon la nouvelle composition :
                # un encadrant qui part -> ses étudiants passent sans lui ; un
                # encadrant qui arrive -> ses étudiants le retrouvent.
                def _normalize_encadrant_absent(jury_obj):
                    member_ids = set(
                        jury_obj.members.values_list("professor_id", flat=True)
                    )
                    for js in JuryStudent.objects.filter(
                        jury=jury_obj
                    ).select_related("student"):
                        absent = js.student.encadrant_id not in member_ids
                        if js.encadrant_absent != absent:
                            JuryStudent.objects.filter(pk=js.pk).update(
                                encadrant_absent=absent
                            )

                _normalize_encadrant_absent(jury)
                _normalize_encadrant_absent(other_jury)

            messages.success(
                request,
                f"Échange effectué : {prof_b.full_name} rejoint ce jury ; "
                f"{prof_a.full_name} rejoint « {other_jury.name} » du "
                f"{other_jury.defense_date.strftime('%d/%m/%Y')}."
            )
            if affected:
                details = " ; ".join(
                    f"{js.student.full_name} ({js.student.matricule}) — encadrant "
                    f"{prof.full_name}"
                    for prof, js in affected
                )
                messages.warning(
                    request,
                    "Ces étudiants soutiendront SANS leur encadrant (désormais "
                    f"marqués « encadrant absent ») suite à l'échange : {details}."
                )
            # Avertissements de disponibilité (l'échange reste possible).
            if has_real_slot and not (
                is_professor_available(prof_b, real_slot_date, real_slot_start)
                and not professor_has_conflict(prof_b, real_slot_date, real_slot_start)
            ):
                messages.warning(
                    request,
                    f"Attention : {prof_b.full_name} n'a pas de disponibilité "
                    f"déclarée au créneau de ce jury."
                )
            other_sched = DefenseSchedule.objects.filter(
                jury_student__jury=other_jury
            ).order_by("start_time").first()
            if other_sched and not (
                is_professor_available(prof_a, other_jury.defense_date, other_sched.start_time)
                and not professor_has_conflict(prof_a, other_jury.defense_date, other_sched.start_time)
            ):
                messages.warning(
                    request,
                    f"Attention : {prof_a.full_name} n'a pas de disponibilité "
                    f"déclarée au créneau du jury « {other_jury.name} »."
                )
            # Notifier les deux professeurs de leur nouveau jury.
            for prof, jry in ((prof_b, jury), (prof_a, other_jury)):
                if jry.is_validated:
                    notify(
                        getattr(prof, "user", None),
                        "Changement de jury",
                        f"Vous êtes désormais membre du jury « {jry.name} » du "
                        f"{jry.defense_date.strftime('%d/%m/%Y')}.",
                        "/professors/juries/",
                        category=Notification.CATEGORY_JURY,
                    )
        return redirect("admin_jury_update", pk=jury.pk)

    # ── 6ter. POST : RETIRER un étudiant du jury ─────────────────────────────
    if request.method == "POST" and request.POST.get("action") == "remove_student":
        try:
            js_id = int(request.POST.get("jury_student_id", ""))
        except (ValueError, TypeError):
            messages.error(request, "Sélection invalide.")
            return redirect("admin_jury_update", pk=jury.pk)

        js = JuryStudent.objects.select_related("student").filter(
            pk=js_id, jury=jury
        ).first()
        if not js:
            messages.error(request, "Étudiant introuvable dans ce jury.")
        elif _has_real_grades(js):
            messages.error(
                request,
                f"Impossible de retirer {js.student.full_name} : des notes ou un "
                f"résultat existent déjà pour cette soutenance."
            )
        else:
            name = js.student.full_name
            js.delete()
            recompact_jury_schedule(jury)
            refresh_jury_name_count(jury)
            messages.success(
                request,
                f"{name} a été retiré du jury. Les horaires ont été resserrés "
                f"(pas de trou). Il redevient affectable à un autre jury."
            )
        return redirect("admin_jury_update", pk=jury.pk)

    # ── 6quater. POST : DÉPLACER un étudiant vers un AUTRE jury ─────────────
    #     (avertissement si son encadrant n'est pas membre du jury cible).
    if request.method == "POST" and request.POST.get("action") == "move_student":
        try:
            js_id = int(request.POST.get("jury_student_id", ""))
            target_id = int(request.POST.get("target_jury_id", ""))
        except (ValueError, TypeError):
            messages.error(request, "Sélection invalide.")
            return redirect("admin_jury_update", pk=jury.pk)

        js = JuryStudent.objects.select_related(
            "student", "student__encadrant"
        ).filter(pk=js_id, jury=jury).first()
        target = Jury.objects.filter(pk=target_id).first()

        if not js or not target:
            messages.error(request, "Étudiant ou jury cible introuvable.")
        elif target.pk == jury.pk:
            messages.error(request, "Choisissez un AUTRE jury.")
        elif _has_real_grades(js):
            messages.error(
                request,
                f"Impossible de déplacer {js.student.full_name} : des notes ou "
                f"un résultat existent déjà pour cette soutenance."
            )
        elif target.members.count() < 3:
            messages.error(
                request,
                f"Le jury cible « {target.name} » est incomplet "
                f"({target.members.count()}/3 membres)."
            )
        else:
            student = js.student
            target_members = [
                m.professor
                for m in target.members.select_related("professor")
            ]
            encadrant_in = any(
                m.id == student.encadrant_id for m in target_members
            )
            president = choose_president_for_student(
                student=student,
                members=target_members,
                defense_date=target.defense_date,
            )
            next_start = calculate_next_defense_slot_for_jury(target, target_members)
            schedule_warning = None
            if next_start is None:
                # Repli admin : après le dernier passage du jury cible (ou en
                # début de matinée), même sans disponibilité déclarée.
                last = DefenseSchedule.objects.filter(
                    jury_student__jury=target
                ).order_by("-start_time").first()
                if last:
                    next_start = last.end_time or slot_end_time(
                        target.defense_date,
                        last.start_time,
                        last.duration_minutes or DEFENSE_DURATION_MINUTES,
                    )
                else:
                    next_start = defense_slots.morning_slot(target.defense_date)[0]
                schedule_warning = (
                    "Horaire placé sans disponibilité déclarée des membres — "
                    "à vérifier."
                )

            source_jury = jury
            with transaction.atomic():
                js.delete()
                new_js = JuryStudent.objects.create(
                    jury=target,
                    student=student,
                    president=president,
                    encadrant_absent=not encadrant_in,
                )
                # bulk_create : ne bloque pas si les membres n'ont pas déclaré
                # de disponibilité (décision de l'admin, avertissement).
                # end_time est renseigné explicitement (bulk_create ne passe
                # pas par save()).
                DefenseSchedule.objects.bulk_create([
                    DefenseSchedule(
                        jury_student=new_js,
                        start_time=next_start,
                        end_time=slot_end_time(
                            target.defense_date, next_start,
                            DEFENSE_DURATION_MINUTES,
                        ),
                        duration_minutes=DEFENSE_DURATION_MINUTES,
                    )
                ])
                recompact_jury_schedule(source_jury)
                recompact_jury_schedule(target)
                refresh_jury_name_count(source_jury)
                refresh_jury_name_count(target)

            messages.success(
                request,
                f"{student.full_name} a été déplacé vers « {target.name} » du "
                f"{target.defense_date.strftime('%d/%m/%Y')} "
                f"(passage à {next_start.strftime('%H:%M')})."
            )
            if not encadrant_in:
                enc_name = (
                    student.encadrant.full_name
                    if student.encadrant else "(encadrant inconnu)"
                )
                messages.warning(
                    request,
                    f"Attention : l'encadrant de {student.full_name} — "
                    f"{enc_name} — n'est PAS membre du jury « {target.name} ». "
                    f"L'affectation est marquée « encadrant absent »."
                )
            if schedule_warning:
                messages.warning(request, schedule_warning)
            if target.is_validated:
                notify(
                    getattr(student, "user", None),
                    "Soutenance déplacée",
                    f"Votre soutenance est désormais prévue le "
                    f"{target.defense_date.strftime('%d/%m/%Y')} à "
                    f"{next_start.strftime('%H:%M')} (jury « {target.name} »).",
                    "/student-dashboard/",
                    category=Notification.CATEGORY_JURY,
                )
        return redirect("admin_jury_update", pk=jury.pk)

    # ── 6quinquies. POST : resserrer les horaires (supprimer les trous) ──────
    if request.method == "POST" and request.POST.get("action") == "recompact":
        recompact_jury_schedule(jury)
        messages.success(
            request,
            "Horaires resserrés : les passages sont désormais contigus (sans trou)."
        )
        return redirect("admin_jury_update", pk=jury.pk)

    # ── 6quinquies-ter. POST : changer l'HEURE DE DÉBUT du jury (même jour) ──
    #     Décale TOUS les passages à partir de la nouvelle heure, contigus.
    if request.method == "POST" and request.POST.get("action") == "set_start_time":
        raw = (request.POST.get("start_time") or "").strip()
        try:
            new_begin = datetime.strptime(raw, "%H:%M").time()
        except (ValueError, TypeError):
            messages.error(request, "Heure de début invalide (format HH:MM).")
            return redirect("admin_jury_update", pk=jury.pk)

        scheds = list(
            DefenseSchedule.objects.filter(jury_student__jury=jury)
            .select_related("jury_student__student__user")
            .order_by("start_time", "id")
        )
        if not scheds:
            messages.error(request, "Ce jury n'a pas encore de passage planifié.")
            return redirect("admin_jury_update", pk=jury.pk)

        old_begin = scheds[0].start_time
        cursor = datetime.combine(jury.defense_date, new_begin)
        changed = []
        for sched in scheds:
            minutes = sched.duration_minutes or DEFENSE_DURATION_MINUTES
            new_start = cursor.time()
            new_end = (cursor + timedelta(minutes=minutes)).time()
            if sched.start_time != new_start or sched.end_time != new_end:
                DefenseSchedule.objects.filter(pk=sched.pk).update(
                    start_time=new_start, end_time=new_end,
                )
                changed.append((sched.jury_student, new_start, new_end))
            cursor += timedelta(minutes=minutes)

        # Prévenir les étudiants si le jury est déjà publié.
        if jury.is_validated and changed:
            for js, ns, ne in changed:
                notify(
                    getattr(js.student, "user", None),
                    "Horaire de soutenance modifié",
                    f"Votre passage est désormais prévu à "
                    f"{ns.strftime('%H:%M')}–{ne.strftime('%H:%M')} le "
                    f"{jury.defense_date.strftime('%d/%m/%Y')} "
                    f"(jury « {jury.name} »).",
                    "/student-dashboard/",
                    category=Notification.CATEGORY_JURY,
                )
        messages.success(
            request,
            f"Heure de début changée de {old_begin.strftime('%H:%M')} à "
            f"{new_begin.strftime('%H:%M')} : les passages ont été décalés."
        )
        return redirect("admin_jury_update", pk=jury.pk)

    # ── 6quinquies-bis. POST : FUSIONNER un autre jury (mêmes membres) dans
    #     celui-ci et coller les passages (supprime le vide).
    if request.method == "POST" and request.POST.get("action") == "merge_jury":
        try:
            other_id = int(request.POST.get("merge_jury_id", ""))
        except (ValueError, TypeError):
            messages.error(request, "Sélection invalide.")
            return redirect("admin_jury_update", pk=jury.pk)
        other = Jury.objects.filter(pk=other_id).first()
        my_ids = sorted(m.professor_id for m in current_members)
        other_ids = sorted(m.professor_id for m in other.members.all()) if other else []
        graded_here = any(
            _has_real_grades(js) for js in jury_students
        )
        graded_other = other and any(
            _has_real_grades(js)
            for js in other.students.all()
        )
        if not other or other.pk == jury.pk:
            messages.error(request, "Jury à fusionner introuvable.")
        elif other_ids != my_ids:
            messages.error(request, "Fusion impossible : les deux jurys n'ont pas les mêmes membres.")
        elif other.defense_date != jury.defense_date:
            messages.error(request, "Fusion impossible : les deux jurys ne sont pas le même jour.")
        elif graded_here or graded_other:
            messages.error(request, "Fusion impossible : un des jurys contient déjà des étudiants notés.")
        else:
            other_name = other.name
            moved = other.students.count()
            with transaction.atomic():
                # Rattacher les étudiants de l'autre jury à celui-ci.
                for js in list(other.students.all()):
                    JuryStudent.objects.filter(pk=js.pk).update(jury=jury)
                # Supprimer l'autre jury (désormais vide) puis coller les passages.
                Jury.objects.filter(pk=other.pk).delete()
                recompact_jury_schedule(jury)
                refresh_jury_name_count(jury)
                renumber_all_juries()
            messages.success(
                request,
                f"« {other_name} » a été fusionné dans ce jury ({moved} étudiant(s) "
                f"ajouté(s)) et les passages ont été collés (sans vide)."
            )
        return redirect("admin_jury_update", pk=jury.pk)

    # ── 6sexies. POST : REPROGRAMMER le jury entier à une autre date ─────────
    #     (report suite à coupure d'électricité, etc.) — composition, salle,
    #     étudiants et ordre de passage conservés.
    if request.method == "POST" and request.POST.get("action") == "reschedule_date":
        members = [m.professor for m in jury.members.select_related("professor")]
        if len(members) < 3:
            messages.error(request, "Le jury doit avoir 3 membres avant d'être reprogrammé.")
            return redirect("admin_jury_update", pk=jury.pk)

        js_list = list(
            JuryStudent.objects.filter(jury=jury)
            .select_related("student__user").order_by("schedule__start_time", "id")
        )
        # Les étudiants DÉJÀ NOTÉS (évaluation ou résultat) restent sur leur
        # date d'origine : on ne reprogramme QUE les non notés.
        graded = [
            js for js in js_list
            if _has_real_grades(js)
        ]
        to_move = [js for js in js_list if js not in graded]
        n = len(to_move)

        if n == 0:
            messages.info(
                request,
                "Aucun étudiant à reprogrammer : tous les étudiants de ce jury "
                "sont déjà notés."
            )
            return redirect("admin_jury_update", pk=jury.pk)

        # L'admin choisit un CRÉNEAU PROPOSÉ (date|HH:MM) où les 3 membres sont
        # disponibles simultanément — pas une date au hasard.
        slot_raw = (request.POST.get("reschedule_slot") or "").strip()
        try:
            date_part, time_part = slot_raw.split("|")
            new_date = date_cls.fromisoformat(date_part)
            chosen_start = datetime.strptime(time_part, "%H:%M").time()
        except (ValueError, AttributeError):
            messages.error(request, "Choisissez un créneau proposé dans la liste.")
            return redirect("admin_jury_update", pk=jury.pk)

        if new_date < timezone.localdate():
            messages.error(request, "La nouvelle date doit être aujourd'hui ou ultérieure.")
            return redirect("admin_jury_update", pk=jury.pk)

        # Revérification serveur : le créneau doit toujours convenir aux 3
        # membres (dispo commune) avec une salle libre.
        start_times, salle = build_block_from_start(
            members, new_date, chosen_start, n, preferred_salle=jury.salle
        )
        if start_times is None:
            messages.error(
                request,
                f"Ce créneau n'est plus disponible pour les 3 membres le "
                f"{new_date.strftime('%d/%m/%Y')}. Choisissez-en un autre dans la liste."
            )
            return redirect("admin_jury_update", pk=jury.pk)

        target_jury = jury
        split = bool(graded)
        with transaction.atomic():
            if split:
                # Certains étudiants sont déjà notés → ils RESTENT dans le jury
                # d'origine (date d'origine). On crée un NOUVEAU jury (mêmes
                # membres) à la nouvelle date pour les non notés.
                target_jury = Jury.objects.create(
                    name=build_grouped_jury_name(
                        [js.student for js in to_move], new_date
                    ),
                    defense_date=new_date,
                    salle=salle,
                    is_validated=jury.is_validated,
                )
                for m in members:
                    JuryMember.objects.create(jury=target_jury, professor=m)
                for js, t in zip(to_move, start_times):
                    JuryStudent.objects.filter(pk=js.pk).update(jury=target_jury)
                    DefenseSchedule.objects.filter(jury_student=js).update(
                        start_time=t, end_time=slot_end_time(new_date, t),
                    )
                # Le jury d'origine ne garde que les notés : on rafraîchit son
                # nom, sans toucher aux horaires déjà passés.
                refresh_jury_name_count(jury)
                refresh_jury_name_count(target_jury)
            else:
                # Aucun étudiant noté → on déplace tout le jury.
                Jury.objects.filter(pk=jury.pk).update(
                    defense_date=new_date, salle=salle
                )
                target_jury.defense_date = new_date
                target_jury.salle = salle
                for js, t in zip(to_move, start_times):
                    DefenseSchedule.objects.filter(jury_student=js).update(
                        start_time=t, end_time=slot_end_time(new_date, t),
                    )

        # Notifier les non notés (déplacés) + les membres si publié.
        if target_jury.is_validated:
            for js, t in zip(to_move, start_times):
                notify(
                    getattr(js.student, "user", None),
                    "Soutenance reportée",
                    f"Votre soutenance est reportée au "
                    f"{new_date.strftime('%d/%m/%Y')} à {t.strftime('%H:%M')} "
                    f"(jury « {target_jury.name} », salle "
                    f"{target_jury.get_salle_display() or '—'}).",
                    "/student-dashboard/",
                    category=Notification.CATEGORY_JURY,
                )
            for m in members:
                notify(
                    getattr(m, "user", None),
                    "Jury reporté",
                    f"Le jury « {target_jury.name} » est prévu le "
                    f"{new_date.strftime('%d/%m/%Y')}.",
                    "/professors/juries/",
                    category=Notification.CATEGORY_JURY,
                )

        créneau = (
            f" — passages de {start_times[0].strftime('%H:%M')} à "
            f"{slot_end_time(new_date, start_times[-1]).strftime('%H:%M')} "
            f"en {target_jury.get_salle_display()}"
        )
        if split:
            messages.success(
                request,
                f"{n} étudiant(s) non noté(s) reprogrammé(s) au "
                f"{new_date.strftime('%d/%m/%Y')}{créneau}. Les "
                f"{len(graded)} étudiant(s) déjà noté(s) restent sur leur date "
                f"d'origine (jury « {jury.name} »)."
            )
        else:
            messages.success(
                request,
                f"Jury reprogrammé au {new_date.strftime('%d/%m/%Y')}{créneau} "
                f"(disponibilité commune des 3 membres vérifiée)."
            )
        return redirect("admin_jury_update", pk=target_jury.pk)

    # ── 7. POST : désigner le président d'une soutenance (≠ encadrant) ────────
    if request.method == "POST" and request.POST.get("action") == "set_president":
        try:
            js_id = int(request.POST.get("jury_student_id", ""))
            pres_id = int(request.POST.get("president_id", ""))
        except (ValueError, TypeError):
            messages.error(request, "Sélection invalide.")
            return redirect("admin_jury_update", pk=jury.pk)

        js = JuryStudent.objects.select_related("student").filter(
            pk=js_id, jury=jury
        ).first()

        if not js:
            messages.error(request, "Étudiant introuvable dans ce jury.")
        elif not jury.members.filter(professor_id=pres_id).exists():
            messages.error(request, "Le président doit être membre du jury.")
        elif js.student.encadrant_id == pres_id:
            messages.error(
                request,
                "L'encadrant de l'étudiant ne peut pas être président de sa soutenance."
            )
        else:
            JuryStudent.objects.filter(pk=js.pk).update(president_id=pres_id)
            messages.success(
                request, f"Président défini pour {js.student.full_name}."
            )
        return redirect("admin_jury_update", pk=jury.pk)

    if request.method == "POST" and request.POST.get("action") == "set_salle":
        salle = (request.POST.get("salle") or "").strip()
        valid = {choice for choice, _ in Jury.SALLE_CHOICES}
        if salle not in valid:
            messages.error(request, "Salle invalide.")
        else:
            Jury.objects.filter(pk=jury.pk).update(salle=salle)
            messages.success(request, "Salle mise à jour.")
        return redirect("admin_jury_update", pk=jury.pk)

    context["salle_choices"] = Jury.SALLE_CHOICES
    return render(request, "soutenances/admin_jury_update.html", context)


@login_required
@role_required(["admin"])
def admin_jury_delete(request, pk):
    """Suppression d'un jury en préservant les étudiants DÉJÀ NOTÉS.

    - Les étudiants non notés sont retirés en masse (redeviennent « sans jury »,
      prêts à être régénérés).
    - Les étudiants déjà notés (évaluation ou résultat) et leurs notes sont
      CONSERVÉS : le jury est gardé mais ne contient plus qu'eux.
    - S'il ne reste aucun étudiant noté, le jury est supprimé entièrement.
    """
    jury = get_object_or_404(Jury, pk=pk)

    if request.method != "POST":
        return redirect("admin_jury_update", pk=pk)

    js_list = list(JuryStudent.objects.filter(jury=jury).select_related("student"))
    graded = [
        js for js in js_list
        if _has_real_grades(js)
    ]
    ungraded = [js for js in js_list if js not in graded]

    if not graded:
        # Aucune note → suppression complète (comportement classique).
        jury_name = jury.name
        jury.delete()
        renumber_all_juries()
        messages.success(request, f"Le jury « {jury_name} » a été supprimé (numéros mis à jour).")
        return redirect("admin_jury_list")

    # Il reste des notes → on libère les non notés et on garde le jury pour les
    # notés (données intactes).
    freed = len(ungraded)
    with transaction.atomic():
        JuryStudent.objects.filter(
            pk__in=[js.pk for js in ungraded]
        ).delete()
        recompact_jury_schedule(jury)
        refresh_jury_name_count(jury)

    messages.success(
        request,
        f"{freed} étudiant(s) non noté(s) libéré(s) (redevenus « sans jury » — "
        f"régénérez-les via « Générer automatiquement », « Placer dans les jurys "
        f"de l'encadrant » ou « Ajouter un jury manuel »). Le jury « {jury.name} » "
        f"est conservé avec les {len(graded)} étudiant(s) déjà noté(s) — notes "
        f"intactes."
    )
    return redirect("admin_jury_list")


@login_required
@role_required(["admin"])
def admin_delete_draft_juries(request):
    """Supprime tous les jurys en brouillon (non validés) en une fois. Ceux qui
    ont déjà des évaluations ou résultats sont conservés."""
    if request.method != "POST":
        return redirect("admin_jury_list")

    drafts = Jury.objects.filter(is_validated=False)
    deleted = 0
    kept = 0
    for jury in drafts:
        has_eval = Evaluation.objects.filter(jury_student__jury=jury).exists()
        has_result = Result.objects.filter(jury_student__jury=jury).exists()
        if has_eval or has_result:
            kept += 1
            continue
        jury.delete()  # CASCADE : membres, JuryStudent, créneaux
        deleted += 1

    if deleted:
        renumber_all_juries()
        messages.success(request, f"{deleted} jury(s) brouillon supprimé(s) (numéros mis à jour).")
    if kept:
        messages.warning(
            request,
            f"{kept} brouillon(s) conservé(s) car des évaluations/résultats existent."
        )
    if not deleted and not kept:
        messages.info(request, "Aucun jury brouillon à supprimer.")
    return redirect("admin_jury_list")


@transaction.atomic
def save_jury_with_members(form):
    jury = form.save()

    JuryMember.objects.filter(jury=jury).delete()

    for professor in form.cleaned_data["members"]:
        JuryMember.objects.create(
            jury=jury,
            professor=professor,
        )

    return jury


@login_required
@role_required(["admin"])
def admin_jury_detail(request, pk):
    jury = get_object_or_404(
        Jury.objects.prefetch_related(
            "members__professor",
            "students__student",
            "students__student__encadrant",
            "students__president",
            "students__schedule",
        ),
        pk=pk,
    )

    form = JuryStudentAssignForm(jury=jury)

    # Rôles des membres pour l'affichage coloré : encadrant / expert /
    # prioritaire / président (au sein de ce jury).
    from collections import defaultdict
    experts_by_filiere = defaultdict(set)
    for entry in FiliereExpert.objects.all():
        experts_by_filiere[entry.filiere].add(entry.professor_id)

    jury_students = list(jury.students.all())
    encadrant_ids = {js.student.encadrant_id for js in jury_students}
    president_ids = {js.president_id for js in jury_students if js.president_id}
    jury_expert_ids = set()
    for js in jury_students:
        jury_expert_ids |= experts_by_filiere.get(js.student.filiere or "", set())

    member_roles = []
    for member in jury.members.all():
        prof = member.professor
        member_roles.append({
            "professor": prof,
            "is_president": prof.id in president_ids,
            "is_encadrant": prof.id in encadrant_ids,
            "is_expert": prof.id in jury_expert_ids,
            "is_priority": prof.is_priority,
        })

    ordered_students = list(
        jury.students.select_related("student__encadrant", "president", "schedule")
        .order_by("schedule__start_time", "id")
    )
    return render(request, "soutenances/admin_jury_detail.html", {
        "jury": jury,
        "form": form,
        "member_roles": member_roles,
        "ordered_students": ordered_students,
    })


@login_required
@role_required(["admin"])
def admin_jury_add_student(request, pk):
    jury = get_object_or_404(
        Jury.objects.prefetch_related("members__professor"),
        pk=pk,
    )

    if request.method == "POST":
        form = JuryStudentAssignForm(request.POST, jury=jury)

        if form.is_valid():
            try:
                student = form.cleaned_data["student"]

                members = [
                    member.professor
                    for member in jury.members.select_related("professor")
                ]

                # L'encadrant de l'étudiant est-il membre de ce jury ?
                encadrant_in_jury = any(
                    m.id == student.encadrant_id for m in members
                )

                president = choose_president_for_student(
                    student=student,
                    members=members,
                    defense_date=jury.defense_date,
                )

                # Horaire PERSONNALISÉ (hors programme) fourni par l'admin ?
                custom_raw = (request.POST.get("custom_start") or "").strip()
                schedule_warning = None
                next_start = None
                if custom_raw:
                    try:
                        next_start = datetime.strptime(custom_raw, "%H:%M").time()
                    except ValueError:
                        messages.error(request, "Heure personnalisée invalide (format HH:MM).")
                        return redirect("admin_jury_detail", pk=jury.pk)
                    warn_bits = []
                    if not (time(9, 0) <= next_start <= time(19, 0)):
                        warn_bits.append("hors de la plage 9h–19h")
                    busy = [
                        m.full_name for m in members
                        if professor_has_conflict(m, jury.defense_date, next_start)
                    ]
                    if busy:
                        warn_bits.append("membre(s) déjà occupé(s) : " + ", ".join(busy))
                    if jury.salle and _salle_occupee(
                        jury.defense_date, next_start,
                        slot_end_time(jury.defense_date, next_start), jury.salle
                    ):
                        warn_bits.append(f"salle {jury.get_salle_display()} occupée")
                    schedule_warning = (
                        f"Horaire personnalisé {custom_raw} placé hors programme"
                        + (" — " + " ; ".join(warn_bits) if warn_bits else "")
                        + "."
                    )

                if next_start is None and not custom_raw:
                    # Calculate the next available 20-min slot for this student
                    next_start = calculate_next_defense_slot_for_jury(jury, members)

                if next_start is None:
                    # Repli admin : placer à la suite du dernier passage du
                    # jury, même sans disponibilité déclarée de tous les
                    # membres (avertissement). On reste dans la demi-journée
                    # du jury et on saute les conflits/salle occupée.
                    last = DefenseSchedule.objects.filter(
                        jury_student__jury=jury
                    ).order_by("-start_time").first()
                    if last:
                        candidate = last.end_time or slot_end_time(
                            jury.defense_date, last.start_time,
                            last.duration_minutes or DEFENSE_DURATION_MINUTES,
                        )
                    else:
                        candidate = defense_slots.morning_slot(jury.defense_date)[0]

                    jury_label = _slot_label_at(
                        jury.defense_date,
                        last.start_time if last else candidate,
                    )

                    def _fits(t):
                        if t is None:
                            return False
                        label = _slot_label_at(jury.defense_date, t)
                        if label is None or label != jury_label:
                            return False
                        _, slot_end = defense_slots.slot_bounds(
                            jury.defense_date, label
                        )
                        return slot_end_time(jury.defense_date, t) <= slot_end

                    while _fits(candidate) and (
                        any(
                            professor_has_conflict(m, jury.defense_date, candidate)
                            for m in members
                        )
                        or (
                            jury.salle and _salle_occupee(
                                jury.defense_date, candidate,
                                slot_end_time(jury.defense_date, candidate),
                                jury.salle,
                            )
                        )
                    ):
                        candidate = slot_end_time(jury.defense_date, candidate)

                    if not _fits(candidate):
                        # Demi-journée pleine : sur décision de l'admin, on place
                        # quand même l'étudiant à la suite du dernier passage, au-delà
                        # de l'heure de fin du créneau (avertissement, pas de blocage).
                        fin = defense_slots.slot_bounds(
                            jury.defense_date, jury_label
                        )[1] if jury_label else None
                        schedule_warning = (
                            "Passage placé à la suite du dernier, au-delà de la fin "
                            "de la demi-journée"
                            + (f" ({fin.strftime('%H:%M')})" if fin else "")
                            + " — à vérifier."
                        )
                    else:
                        schedule_warning = (
                            "Horaire placé sans disponibilité déclarée de tous les "
                            "membres — à vérifier."
                        )

                    next_start = candidate

                with transaction.atomic():
                    assignment = JuryStudent.objects.create(
                        jury=jury,
                        student=student,
                        president=president,
                        # Ajout forcé par l'admin : l'encadrant n'est pas dans
                        # le jury → marqué « encadrant absent ».
                        encadrant_absent=not encadrant_in_jury,
                    )
                    if schedule_warning:
                        # bulk_create : ne bloque pas sur la validation de
                        # disponibilité des membres (décision de l'admin).
                        DefenseSchedule.objects.bulk_create([
                            DefenseSchedule(
                                jury_student=assignment,
                                start_time=next_start,
                                end_time=slot_end_time(
                                    jury.defense_date, next_start,
                                    DEFENSE_DURATION_MINUTES,
                                ),
                                duration_minutes=DEFENSE_DURATION_MINUTES,
                            )
                        ])
                    else:
                        DefenseSchedule.objects.create(
                            jury_student=assignment,
                            start_time=next_start,
                            duration_minutes=DEFENSE_DURATION_MINUTES,
                        )
                    # Un horaire PERSONNALISÉ (hors programme) ne doit pas être
                    # recompacté (sinon il serait ramené à la suite des autres).
                    if not custom_raw:
                        recompact_jury_schedule(jury)
                    refresh_jury_name_count(jury)

            except ValidationError as exc:
                messages.error(request, "; ".join(exc.messages))
            else:
                messages.success(
                    request,
                    f"{student.full_name} a été ajouté au jury « {jury.name} » — "
                    f"passage à {next_start.strftime('%H:%M')}."
                )
                if schedule_warning:
                    messages.warning(request, schedule_warning)
                if not encadrant_in_jury:
                    enc_name = (
                        student.encadrant.full_name
                        if student.encadrant else "(encadrant inconnu)"
                    )
                    messages.warning(
                        request,
                        f"Attention : l'encadrant de {student.full_name} — "
                        f"{enc_name} — n'est PAS membre de ce jury. "
                        f"L'affectation est marquée « encadrant absent »."
                    )

            return redirect("admin_jury_detail", pk=jury.pk)

        messages.error(request, "Impossible d'ajouter cet étudiant au jury.")

        return render(request, "soutenances/admin_jury_detail.html", {
            "jury": jury,
            "form": form,
        })

    return redirect("admin_jury_detail", pk=jury.pk)


@login_required
@role_required(["admin"])
def admin_jury_remove_student(request, pk, assignment_pk):
    jury = get_object_or_404(Jury, pk=pk)

    assignment = get_object_or_404(
        JuryStudent,
        pk=assignment_pk,
        jury=jury,
    )

    if request.method == "POST":
        if _has_real_grades(assignment):
            messages.error(
                request,
                "Impossible de retirer cet étudiant : des notes ou un résultat "
                "existent déjà pour cette soutenance."
            )
            return redirect("admin_jury_detail", pk=jury.pk)
        assignment.delete()
        recompact_jury_schedule(jury)
        refresh_jury_name_count(jury)
        messages.success(
            request,
            "L'affectation a été supprimée. Les horaires ont été resserrés "
            "(pas de trou)."
        )

    return redirect("admin_jury_detail", pk=jury.pk)


def _notify_jury_published(jury):
    """Prévient les étudiants (date + horaire) et les membres du jury publié."""
    jury_students = JuryStudent.objects.filter(jury=jury).select_related(
        "student", "student__user"
    )
    for js in jury_students:
        schedule = get_assignment_schedule(js)
        horaire = ""
        if schedule:
            horaire = (
                f" à {format_time(schedule.start_time)}"
                f"–{format_time(schedule.end_time)}"
            )
        salle = f" Salle : {jury.get_salle_display()}." if jury.salle else ""
        notify(
            getattr(js.student, "user", None),
            "Votre soutenance est planifiée",
            f"Date : {format_date(jury.defense_date)}{horaire}.{salle} Jury : {jury.name}.",
            "/student-dashboard/",
            category=Notification.CATEGORY_JURY,
        )

    for member in JuryMember.objects.filter(jury=jury).select_related(
        "professor", "professor__user"
    ):
        notify(
            getattr(member.professor, "user", None),
            "Vous êtes membre d'un jury",
            f"Jury « {jury.name} » — {format_date(jury.defense_date)}.",
            "/professors/juries/",
            category=Notification.CATEGORY_JURY,
        )


@login_required
@role_required(["admin"])
def admin_jury_publish(request, pk):
    jury = get_object_or_404(Jury, pk=pk)

    if request.method == "POST":
        already_validated = jury.is_validated
        jury.is_validated = True
        jury.save(update_fields=["is_validated"])

        if not already_validated:
            _notify_jury_published(jury)

        messages.success(request, "Le jury a été publié. Étudiants et professeurs peuvent le voir.")
        next_url = request.POST.get("next") or request.META.get("HTTP_REFERER")
        if next_url:
            return redirect(next_url)
        return redirect("admin_jury_list")

    return redirect("admin_jury_list")


@login_required
@role_required(["admin"])
def admin_planning(request):
    # Planning : uniquement les soutenances À VENIR (date >= aujourd'hui),
    # publiées ou non. Les jurys passés restent consultables via l'onglet
    # « Passés » de la liste des jurys.
    schedules = DefenseSchedule.objects.select_related(
        "jury_student__student",
        "jury_student__student__encadrant",
        "jury_student__president",
        "jury_student__jury",
    ).prefetch_related(
        "jury_student__jury__members__professor",
    ).filter(
        jury_student__jury__defense_date__gte=timezone.localdate(),
    ).order_by(
        "jury_student__jury__defense_date",
        "start_time",
        "jury_student__jury__name",
    )

    # Experts par filière (pour étiqueter le rôle des membres par étudiant).
    from collections import defaultdict
    experts_by_filiere = defaultdict(set)
    for entry in FiliereExpert.objects.all():
        experts_by_filiere[entry.filiere].add(entry.professor_id)

    rows = []
    for schedule in schedules:
        js = schedule.jury_student
        student = js.student
        student_experts = experts_by_filiere.get(student.filiere or "", set())
        members = []
        for m in js.jury.members.all():
            prof = m.professor
            members.append({
                "professor": prof,
                "is_president": prof.id == js.president_id,
                "is_encadrant": prof.id == student.encadrant_id,
                "is_expert": prof.id in student_experts,
                "is_priority": prof.is_priority,
            })
        rows.append({"schedule": schedule, "members": members})

    return render(request, "soutenances/admin_planning.html", {
        "schedules": schedules,
        "planning_rows": rows,
        "generation_form": PlanningGenerationForm(),
        "duration_minutes": DEFENSE_DURATION_MINUTES,
    })


@login_required
@role_required(["admin"])
def admin_generate_planning(request):
    if request.method == "POST":
        form = PlanningGenerationForm(request.POST)

        if form.is_valid():
            result = generate_planning_for_date(
                form.cleaned_data["defense_date"],
                form.cleaned_data["overwrite_existing"],
            )

            if result["created"]:
                messages.success(
                    request,
                    f"{result['created']} horaire(s) généré(s) en créneaux de 20 minutes."
                )

            if result["errors"]:
                messages.warning(
                    request,
                    f"{len(result['errors'])} affectation(s) sans horaire disponible."
                )

            return render(request, "soutenances/admin_planning_generate.html", {
                "form": PlanningGenerationForm(initial={
                    "defense_date": form.cleaned_data["defense_date"],
                }),
                "result": result,
                "duration_minutes": DEFENSE_DURATION_MINUTES,
            })

    else:
        first_jury = Jury.objects.order_by("-defense_date").first()

        initial = {
            "defense_date": first_jury.defense_date
        } if first_jury else {}

        form = PlanningGenerationForm(initial=initial)

    return render(request, "soutenances/admin_planning_generate.html", {
        "form": form,
        "duration_minutes": DEFENSE_DURATION_MINUTES,
    })


@transaction.atomic
def generate_planning_for_date(defense_date, overwrite_existing=False):
    if overwrite_existing:
        DefenseSchedule.objects.filter(
            jury_student__jury__defense_date=defense_date
        ).delete()

    assignments = JuryStudent.objects.filter(
        jury__defense_date=defense_date,
        schedule__isnull=True,
    ).select_related(
        "student",
        "student__encadrant",
        "president",
        "jury",
    ).order_by(
        "student__encadrant__full_name",
        "jury__name",
        "student__full_name",
    )

    slots = build_slots(defense_date)

    result = {
        "created": 0,
        "errors": [],
    }

    for assignment in assignments:
        created = False
        last_error = None

        for slot in slots:
            if jury_slot_capacity_reached(defense_date, slot):
                continue

            schedule = DefenseSchedule(
                jury_student=assignment,
                start_time=slot,
                duration_minutes=DEFENSE_DURATION_MINUTES,
            )

            try:
                schedule.save()
            except ValidationError as exc:
                last_error = exc
                continue
            else:
                result["created"] += 1
                created = True
                break

        if not created:
            result["errors"].append({
                "assignment": assignment,
                "message": (
                    last_error.messages[0]
                    if last_error and hasattr(last_error, "messages")
                    else "Aucun créneau compatible."
                ),
            })

    return result


def build_smart_jury_name(student, index):
    return f"Jury {index}"


def build_future_slots_for_professor(professor):
    availabilities = ProfessorAvailability.objects.filter(
        professor=professor,
        date__gte=timezone.localdate(),
    ).order_by(
        "date",
        "start_time",
    )

    slots = []
    seen = set()

    for availability in availabilities:
        for defense_date, start_time in build_slots_from_availability(availability):
            key = (defense_date, start_time)

            if key in seen:
                continue

            seen.add(key)
            slots.append(key)

    return slots


def build_slots_from_availability(availability):
    slots = []

    cursor = datetime.combine(
        availability.date,
        availability.start_time,
    )

    limit = datetime.combine(
        availability.date,
        availability.end_time,
    )

    while cursor + timedelta(minutes=DEFENSE_DURATION_MINUTES) <= limit:
        slots.append((
            availability.date,
            cursor.time(),
        ))
        cursor += timedelta(minutes=DEFENSE_DURATION_MINUTES)

    return slots


@login_required
@role_required(["admin"])
def admin_fill_unassigned_into_juries(request):
    """Place les étudiants sans jury dans un jury EXISTANT de leur encadrant
    qui a encore du temps libre (matin fini avant 14h ou après-midi avant 19h),
    avec les 3 membres disponibles et la salle libre. Les horaires sont
    resserrés ; les étudiants des jurys publiés sont prévenus."""
    if request.method != "POST":
        return redirect("admin_scheduling_diagnostic")

    today = timezone.localdate()
    unassigned = (
        StudentProfile.objects.filter(
            pfe_request__status=PFERequest.STATUS_ACCEPTED,
            jury_assignment__isnull=True,
            encadrant__isnull=False,
        ).select_related("encadrant").order_by("encadrant__full_name", "full_name")
    )

    placed = 0
    for student in unassigned:
        candidate_juries = (
            Jury.objects.filter(
                members__professor=student.encadrant,
                defense_date__gte=today,
            ).order_by("defense_date", "id").distinct()
        )
        for jury in candidate_juries:
            members = [m.professor for m in jury.members.select_related("professor")]
            start = find_free_slot_in_jury(jury, members)
            if start is None:
                continue
            president = choose_president_for_student(
                student=student, members=members, defense_date=jury.defense_date,
            )
            try:
                with transaction.atomic():
                    js = JuryStudent.objects.create(
                        jury=jury, student=student, president=president,
                        encadrant_absent=False,
                    )
                    DefenseSchedule.objects.create(
                        jury_student=js, start_time=start,
                        duration_minutes=DEFENSE_DURATION_MINUTES,
                    )
                    refresh_jury_name_count(jury)
            except ValidationError:
                continue
            if jury.is_validated:
                notify(
                    getattr(student, "user", None),
                    "Soutenance planifiée",
                    f"Votre soutenance est prévue le "
                    f"{jury.defense_date.strftime('%d/%m/%Y')} à "
                    f"{start.strftime('%H:%M')} (jury « {jury.name} »).",
                    "/student-dashboard/",
                    category=Notification.CATEGORY_JURY,
                )
            placed += 1
            break

    remaining = unassigned.filter(jury_assignment__isnull=True).count()
    if placed:
        messages.success(
            request,
            f"{placed} étudiant(s) placé(s) dans un jury de leur encadrant "
            f"(temps libre). Restants : {remaining}."
        )
    else:
        messages.info(
            request,
            "Aucun étudiant n'a pu être placé dans un jury existant de son "
            "encadrant (pas de temps libre commun). Utilisez « Générer "
            "automatiquement » ou « Ajouter un jury manuel »."
        )
    return redirect("admin_scheduling_diagnostic")


@login_required
@role_required(["admin"])
def admin_scheduling_diagnostic(request):
    """Diagnostic de programmation : étudiants ACCEPTÉS non programmés,
    groupés par encadrant avec la cause précise (pas de dispo / dispo
    consommée / salles saturées / créneaux encore utilisables), et
    professeurs disponibles jamais mobilisés."""
    from collections import defaultdict

    today = timezone.localdate()

    accepted = StudentProfile.objects.filter(
        pfe_request__status=PFERequest.STATUS_ACCEPTED,
    )
    total_accepted = accepted.count()
    unscheduled = list(
        accepted.filter(jury_assignment__isnull=True)
        .select_related("encadrant", "user")
        .order_by("encadrant__full_name", "full_name")
    )

    groups = {}
    for s in unscheduled:
        groups.setdefault(s.encadrant_id, []).append(s)

    rows = []
    for enc_id, students in groups.items():
        enc = students[0].encadrant
        filieres = sorted({s.filiere or "?" for s in students})

        if enc is None:
            rows.append({
                "encadrant": None,
                "students": students,
                "count": len(students),
                "used": 0, "total": 0, "free": 0, "usable": [],
                "diagnosis": "Encadrant inconnu.",
                "action": "Corriger l'encadrant de ces étudiants.",
                "experts": [],
            })
            continue

        used, total, free_slots = feasible_priority_usage(enc, today)
        # Créneaux réellement libres : sans conflit avec un jury où il siège
        # (couvre aussi les horaires non alignés sur la grille de 20 min).
        free_slots = [
            (d, t) for (d, t) in free_slots
            if not professor_has_conflict(enc, d, t)
        ]
        used = total - len(free_slots)

        # Créneaux libres de l'encadrant réellement UTILISABLES
        # (capacité de jurys simultanés non atteinte + une salle libre).
        usable = []
        for (d, t) in free_slots:
            if jury_slot_capacity_reached(d, t):
                continue
            end = slot_end_time(d, t)
            if _choisir_salle_libre(d, t, end) is None:
                continue
            usable.append((d, t))

        # Experts définis pour les filières de ces étudiants (2e passe).
        experts = list(
            FiliereExpert.objects.filter(filiere__in=filieres)
            .select_related("professor")
            .values_list("professor__full_name", flat=True)
        )

        if total == 0:
            diagnosis = "Aucune disponibilité déclarée par l'encadrant."
            action = (
                "Demander des disponibilités à l'encadrant, OU laisser la "
                "génération former un jury avec un expert de la filière "
                f"({', '.join(experts) if experts else 'AUCUN EXPERT DÉFINI — à cocher dans Experts par filière'}), "
                "OU affecter manuellement (marqué « encadrant absent »)."
            )
        elif not free_slots:
            diagnosis = (
                f"Disponibilité entièrement consommée : {used}/{total} créneaux "
                f"déjà occupés par les jurys où il siège."
            )
            action = (
                "Demander des disponibilités supplémentaires, retirer/échanger "
                "l'encadrant d'un jury, ou déplacer les étudiants vers un autre "
                "jury (marqués « encadrant absent »)."
            )
        elif not usable:
            diagnosis = (
                f"{len(free_slots)} créneau(x) libres chez l'encadrant, mais "
                f"salles ou capacité saturées sur ces créneaux."
            )
            action = "Libérer une salle ou déplacer un jury sur ces créneaux."
        else:
            diagnosis = (
                f"{len(usable)} créneau(x) encore utilisables chez l'encadrant."
            )
            action = (
                "Relancer « Générer automatiquement » avec une fenêtre couvrant "
                "ces créneaux : ces étudiants seront placés."
            )

        rows.append({
            "encadrant": enc,
            "students": students,
            "count": len(students),
            "used": used, "total": total,
            "free": len(free_slots),
            "usable": usable[:8],
            "usable_count": len(usable),
            "diagnosis": diagnosis,
            "action": action,
            "experts": experts,
            "filieres": filieres,
        })

    rows.sort(key=lambda r: -r["count"])

    # Professeurs disponibles mais jamais mobilisés dans un jury à venir.
    idle_profs = []
    busy_ids = set(
        JuryMember.objects.filter(
            jury__defense_date__gte=today
        ).values_list("professor_id", flat=True)
    )
    for p in ProfessorProfile.objects.order_by("full_name"):
        if p.id in busy_ids:
            continue
        if p.availabilities.filter(date__gte=today).exists():
            idle_profs.append(p)

    # ── Recherche : matricule/nom d'étudiant OU nom de professeur ────────────
    query = (request.GET.get("q") or "").strip()
    student_results = []
    prof_results = []
    if query:
        from django.db.models import Q
        students_found = StudentProfile.objects.filter(
            Q(matricule__icontains=query) | Q(full_name__icontains=query)
        ).select_related("encadrant")[:15]
        for s in students_found:
            js = JuryStudent.objects.filter(student=s).select_related(
                "jury", "president"
            ).first()
            schedule = (
                DefenseSchedule.objects.filter(jury_student=js).first()
                if js else None
            )
            demande = getattr(s, "pfe_request", None)
            student_results.append({
                "student": s,
                "js": js,
                "jury": js.jury if js else None,
                "schedule": schedule,
                "members": (
                    [m.professor.full_name for m in js.jury.members.select_related("professor")]
                    if js else []
                ),
                "demande_status": demande.get_status_display() if demande else "Aucune demande",
            })

        for p in ProfessorProfile.objects.filter(
            full_name__icontains=query
        ).order_by("full_name")[:10]:
            memberships = []
            for jm in JuryMember.objects.filter(professor=p).select_related(
                "jury"
            ).order_by("jury__defense_date"):
                scheds = DefenseSchedule.objects.filter(
                    jury_student__jury=jm.jury
                ).order_by("start_time")
                first = scheds.first()
                last = scheds.last()
                memberships.append({
                    "jury": jm.jury,
                    "students_count": jm.jury.students.count(),
                    "start": first.start_time if first else None,
                    "end": (last.end_time or last.start_time) if last else None,
                })
            prof_results.append({"professor": p, "memberships": memberships})

    return render(request, "soutenances/admin_scheduling_diagnostic.html", {
        "rows": rows,
        "unscheduled_count": len(unscheduled),
        "total_accepted": total_accepted,
        "scheduled_count": total_accepted - len(unscheduled),
        "idle_profs": idle_profs,
        "query": query,
        "student_results": student_results,
        "prof_results": prof_results,
    })


def feasible_priority_usage(prof, today=None):
    """Utilisation d'un professeur sur sa disponibilité FAISABLE.

    Règle stricte matin OU après-midi : si un prof déclare les deux
    demi-journées d'un même jour, une seule est réellement utilisable. Par
    jour, le dénominateur retient la demi-journée où il est programmé (s'il
    l'est), sinon la plus grande déclarée.

    Retourne (créneaux utilisés, créneaux faisables, créneaux libres triés).
    """
    from collections import defaultdict

    if today is None:
        today = timezone.localdate()

    day_slots = defaultdict(lambda: defaultdict(set))
    for availability in prof.availabilities.filter(date__gte=today):
        for d, t in build_slots_from_availability(availability):
            label = _slot_label_at(d, t)
            if label:
                day_slots[d][label].add(t)

    sched_by_day = defaultdict(set)
    rows = DefenseSchedule.objects.filter(
        jury_student__jury__members__professor=prof,
        jury_student__jury__defense_date__gte=today,
    ).values_list("jury_student__jury__defense_date", "start_time")
    for d, t in rows:
        sched_by_day[d].add(t)

    used = total = 0
    free_slots = []
    for d, slots_map in day_slots.items():
        used_labels = {_slot_label_at(d, t) for t in sched_by_day.get(d, set())}
        used_labels.discard(None)
        if used_labels:
            label = sorted(used_labels)[0]
        else:
            label = max(slots_map, key=lambda k: len(slots_map[k]))
        avail = slots_map.get(label, set())
        day_used = sched_by_day.get(d, set()) & avail
        used += len(day_used)
        total += len(avail)
        free_slots.extend((d, t) for t in sorted(avail - day_used))
    free_slots.sort()
    return used, total, free_slots


@login_required
@role_required(["admin"])
def admin_priority_professors_report(request):
    """Sélection des profs prioritaires (cases à cocher, comme les experts) et
    rapport du taux d'utilisation de leurs disponibilités (créneaux de 20 min à
    venir) : combien sont déjà occupés par un jury, combien restent libres."""
    professors = list(ProfessorProfile.objects.order_by("full_name"))

    if request.method == "POST":
        posted_ids = {
            int(pid) for pid in request.POST.getlist("priority") if pid.isdigit()
        }
        ProfessorProfile.objects.filter(is_priority=True).exclude(
            id__in=posted_ids
        ).update(is_priority=False)
        ProfessorProfile.objects.filter(id__in=posted_ids).update(is_priority=True)
        messages.success(request, "Liste des profs prioritaires mise à jour.")
        return redirect("admin_priority_professors_report")

    today = timezone.localdate()
    priority_profs = [p for p in professors if p.is_priority]

    rows = []
    for prof in priority_profs:
        # Taux calculé sur la disponibilité FAISABLE : au plus une demi-journée
        # par jour (règle matin OU après-midi).
        used, total, free_slots = feasible_priority_usage(prof, today)
        rows.append({
            "professor": prof,
            "total": total,
            "used": used,
            "free": total - used,
            "pct": round(100 * used / total) if total else 0,
            "free_slots": free_slots,
            "president_count": JuryStudent.objects.filter(president=prof).count(),
        })

    return render(request, "soutenances/admin_priority_report.html", {
        "rows": rows,
        "professors": professors,
        "priority_count": len(priority_profs),
    })


def build_slots(defense_date):
    """Créneaux de 20 min disponibles ce jour-là, à l'intérieur des deux
    créneaux matin/après-midi (exception vendredi)."""
    slots = []

    for start, end in defense_slots_for(defense_date):
        cursor = datetime.combine(defense_date, start)
        limit = datetime.combine(defense_date, end)

        while cursor + timedelta(minutes=DEFENSE_DURATION_MINUTES) <= limit:
            slots.append(cursor.time())
            cursor += timedelta(minutes=DEFENSE_DURATION_MINUTES)

    return slots


def slot_end_time(defense_date, start_time, duration_minutes=DEFENSE_DURATION_MINUTES):
    start_datetime = datetime.combine(defense_date, start_time)
    end_datetime = start_datetime + timedelta(minutes=duration_minutes)
    return end_datetime.time()


def is_professor_available(professor, defense_date, start_time, duration_minutes=DEFENSE_DURATION_MINUTES):
    end_time = slot_end_time(defense_date, start_time, duration_minutes)

    return ProfessorAvailability.objects.filter(
        professor=professor,
        date=defense_date,
        start_time__lte=start_time,
        end_time__gte=end_time,
    ).exists()


def professor_has_conflict(professor, defense_date, start_time, duration_minutes=DEFENSE_DURATION_MINUTES):
    end_time = slot_end_time(defense_date, start_time, duration_minutes)

    return DefenseSchedule.objects.filter(
        jury_student__jury__defense_date=defense_date,
        start_time__lt=end_time,
        end_time__gt=start_time,
        jury_student__jury__members__professor=professor,
    ).distinct().exists()


def professor_busy_other_slot(professor, defense_date, current_slot):
    """Règle STRICTE : un membre de jury siège SOIT le matin SOIT l'après-midi
    d'une même journée, jamais les deux. Vrai si le professeur a déjà une
    soutenance programmée ce jour-là dans l'AUTRE créneau (vérifié en base,
    donc valable aussi pour les jurys créés lors de générations précédentes)."""
    if not current_slot:
        return False
    other = (
        defense_slots.MORNING
        if current_slot == defense_slots.AFTERNOON
        else defense_slots.AFTERNOON
    )
    o_start, o_end = defense_slots.slot_bounds(defense_date, other)
    return DefenseSchedule.objects.filter(
        jury_student__jury__defense_date=defense_date,
        jury_student__jury__members__professor=professor,
        start_time__lt=o_end,
        end_time__gt=o_start,
    ).exists()


def professor_slot_status(professor, defense_date, start_time):
    """Statut d'un professeur pour un créneau donné :
    - free=True s'il est libre ici (dispo déclarée + pas de conflit + pas déjà
      en jury l'autre demi-journée) ;
    - sinon un libellé expliquant pourquoi et s'il est libre l'autre
      demi-journée (ex. « Pris ce créneau · libre l'après-midi »)."""
    if not (defense_date and start_time):
        return {"free": None, "label": ""}
    label = _slot_label_at(defense_date, start_time)
    other = (
        defense_slots.MORNING if label == defense_slots.AFTERNOON
        else defense_slots.AFTERNOON
    )
    other_fr = "le matin" if other == defense_slots.MORNING else "l'après-midi"
    avail_here = is_professor_available(professor, defense_date, start_time)
    conflict_here = professor_has_conflict(professor, defense_date, start_time)
    busy_other = professor_busy_other_slot(professor, defense_date, label)

    if avail_here and not conflict_here and not busy_other:
        return {"free": True, "label": "Libre à ce créneau"}

    o_start, _ = defense_slots.slot_bounds(defense_date, other)
    avail_other = is_professor_available(professor, defense_date, o_start)

    if busy_other:
        return {"free": False, "label": f"Déjà en jury {other_fr}"}
    if conflict_here:
        if avail_other:
            return {"free": False, "label": f"Pris ce créneau · libre {other_fr}"}
        return {"free": False, "label": "Déjà en jury à ce créneau"}
    if not avail_here:
        if avail_other:
            return {"free": False, "label": f"Dispo {other_fr} seulement"}
        return {"free": False, "label": "Aucune disponibilité ce jour"}
    return {"free": False, "label": "Indisponible"}


def juries_count_at_slot(defense_date, start_time, duration_minutes=DEFENSE_DURATION_MINUTES):
    end_time = slot_end_time(defense_date, start_time, duration_minutes)

    return DefenseSchedule.objects.filter(
        jury_student__jury__defense_date=defense_date,
        start_time__lt=end_time,
        end_time__gt=start_time,
    ).values_list("jury_student__jury_id", flat=True).distinct().count()


def jury_slot_capacity_reached(defense_date, start_time, duration_minutes=DEFENSE_DURATION_MINUTES, max_simultaneous=None):
    cap = max_simultaneous or MAX_SIMULTANEOUS_JURIES
    return juries_count_at_slot(defense_date, start_time, duration_minutes) >= cap


def professor_load_on_date(professor, defense_date):
    return JuryMember.objects.filter(
        professor=professor,
        jury__defense_date=defense_date,
    ).count()


def professor_total_scheduled_load(professor):
    return DefenseSchedule.objects.filter(
        jury_student__jury__members__professor=professor,
    ).distinct().count()


def supervised_students_count(professor):
    try:
        return professor.students.count()
    except Exception:
        return 0


def compute_criteria_averages(assignment):
    """Moyennes des notes des membres du jury par critère (Rapport, Présentation,
    Questions) et note finale. Robuste aux changements de composition : ne compte
    que les évaluations envoyées par les membres ACTUELS du jury. « complete » est
    vrai quand tous les membres actuels ont envoyé leur note."""
    from decimal import Decimal

    member_ids = set(assignment.jury.members.values_list("professor_id", flat=True))
    submitted = [
        e for e in assignment.evaluations.all()
        if e.is_submitted and e.professor_id in member_ids
    ]
    members_count = len(member_ids)
    n = len(submitted)
    complete = members_count > 0 and n >= members_count

    data = {
        "submitted_count": n,
        "members_count": members_count,
        "complete": complete,
        "avg_rapport": None,
        "avg_presentation": None,
        "avg_questions": None,
        "avg_finale": None,
        "raw_avg_finale": None,
        "gap": None,
        "gap_alert": False,
        "submitted": submitted,
        "breakdown": None,
    }

    if n == 0:
        return data

    # Note corrigée critère par critère (écart >= 3 -> membre aberrant écarté)
    # ET moyennes BRUTES (simple moyenne des membres).
    bd = corrected_breakdown(submitted)

    def _mean(field):
        total = sum((getattr(e, field) for e in submitted), Decimal("0"))
        return (total / Decimal(n)).quantize(Decimal("0.01"))

    # On affiche CE QUI A ÉTÉ CALCULÉ : l'affichage doit correspondre à la note
    # STOCKÉE. Un résultat publié AVANT la règle de correction (moyenne simple)
    # s'affiche en BRUT ; un résultat corrigé par critère s'affiche corrigé.
    result = getattr(assignment, "result", None)
    stored = result.average if result else None
    use_raw = (
        stored is not None
        and abs(stored - bd["raw_avg_finale"]) <= Decimal("0.01")
        and abs(stored - bd["avg_finale"]) > Decimal("0.01")
    )

    if use_raw:
        data["avg_rapport"] = _mean("rapport_note")
        data["avg_presentation"] = _mean("presentation_note")
        data["avg_questions"] = _mean("questions_note")
        data["avg_finale"] = bd["raw_avg_finale"]
        data["gap"] = bd["gap"]
        data["gap_alert"] = False
        data["breakdown"] = None
    else:
        data["avg_rapport"] = bd["avg_rapport"]
        data["avg_presentation"] = bd["avg_presentation"]
        data["avg_questions"] = bd["avg_questions"]
        data["avg_finale"] = bd["avg_finale"]
        data["gap"] = bd["gap"]
        data["gap_alert"] = bd["gap_alert"]
        data["breakdown"] = bd
    data["raw_avg_finale"] = bd["raw_avg_finale"]
    return data


@login_required
@role_required(["admin"])
def admin_results(request):
    assignments = JuryStudent.objects.select_related(
        "student",
        "jury",
        "president",
        "result",
    ).prefetch_related(
        "evaluations__professor",
        "jury__members",
    ).order_by(
        "jury__defense_date",
        "student__full_name",
    )

    items = []

    for assignment in assignments:
        result = getattr(assignment, "result", None)
        avgs = compute_criteria_averages(assignment)
        ready = avgs["complete"]

        computed_average = avgs["avg_finale"] if ready else None
        computed_gap = avgs["gap"] if ready else None
        computed_gap_alert = avgs["gap_alert"] if ready else False

        mention_average = result.average if (result and result.average is not None) else computed_average
        mention = mention_for_average(mention_average) if mention_average is not None else None

        # Une alerte est conservée dans l'historique même après publication.
        is_alert = bool(result.has_note_gap_alert) if result else bool(computed_gap_alert)

        # Récence : dernière note envoyée (sinon publication) — pour trier de la
        # note la plus récente à la plus ancienne.
        sub_times = [e.submitted_at for e in avgs["submitted"] if e.submitted_at]
        recency = max(sub_times) if sub_times else (result.published_at if result else None)

        items.append({
            "assignment": assignment,
            "evaluations": avgs["submitted"],
            "result": result,
            "ready": ready,
            "computed_average": computed_average,
            "computed_gap": computed_gap,
            "computed_gap_alert": computed_gap_alert,
            "avg_rapport": avgs["avg_rapport"] if ready else None,
            "avg_presentation": avgs["avg_presentation"] if ready else None,
            "avg_questions": avgs["avg_questions"] if ready else None,
            "raw_average": avgs["raw_avg_finale"] if ready else None,
            "breakdown": avgs["breakdown"] if ready else None,
            "is_alert": is_alert,
            "recency": recency,
            "mention": mention,
        })

    # Ordre : de la note la plus récente à la plus ancienne (les alertes aussi).
    items.sort(key=lambda it: it["recency"].timestamp() if it["recency"] else 0.0,
               reverse=True)

    alert_count = sum(1 for it in items if it["is_alert"])

    return render(request, "soutenances/admin_results.html", {
        "items": items,
        "alert_count": alert_count,
    })


@login_required
@role_required(["admin"])
def admin_publish_result(request, pk):
    assignment = get_object_or_404(JuryStudent, pk=pk)

    if request.method == "POST":
        # Ne pas créer de Result vide tant que les 3 notes ne sont pas là
        # (un Result vide bloquerait ensuite le retrait de l'étudiant).
        if assignment.evaluations.filter(is_submitted=True).count() != 3:
            messages.error(request, "Les 3 évaluations ne sont pas encore saisies.")
            return redirect("admin_results")

        result, _ = Result.objects.get_or_create(
            jury_student=assignment,
        )

        try:
            # On (re)calcule avec la règle en vigueur (note corrigée critère par
            # critère en cas d'écart >= 3) avant de publier.
            result.calculate_average()
            result.publish()
        except ValidationError as exc:
            messages.error(request, "; ".join(exc.messages))
        else:
            notify(
                getattr(assignment.student, "user", None),
                "Résultat de soutenance publié",
                "Votre note finale est disponible dans votre espace.",
                "/student-dashboard/",
                category=Notification.CATEGORY_RESULT,
            )
            messages.success(request, "Résultat publié.")

    return redirect("admin_results")


@login_required
@role_required(["admin"])
def admin_apply_gap_rule(request, pk):
    """Applique la règle de correction d'écart à UN étudiant (cas par cas) :
    recalcule la note en écartant, critère par critère, la note aberrante
    (écart >= 3), puis publie le résultat. Sert notamment aux résultats en
    alerte restés non publiés."""
    assignment = get_object_or_404(JuryStudent, pk=pk)

    if request.method == "POST":
        if assignment.evaluations.filter(is_submitted=True).count() != 3:
            messages.error(request, "Les 3 évaluations ne sont pas encore saisies.")
            return redirect("admin_results")

        result, _ = Result.objects.get_or_create(jury_student=assignment)
        try:
            result.calculate_average()  # force la règle par critère
            result.publish()
        except ValidationError as exc:
            messages.error(request, "; ".join(exc.messages))
            return redirect("admin_results")

        notify(
            getattr(assignment.student, "user", None),
            "Résultat de soutenance publié",
            "Votre note finale est disponible dans votre espace.",
            "/student-dashboard/",
            category=Notification.CATEGORY_RESULT,
        )
        if result.has_note_gap_alert:
            messages.success(
                request,
                f"Règle appliquée : {assignment.student.full_name} — note corrigée "
                f"{result.average}/20 (écart {result.note_gap_value}, membre aberrant "
                f"écarté par critère). Résultat publié."
            )
        else:
            messages.success(
                request,
                f"Résultat publié : {assignment.student.full_name} — moyenne "
                f"{result.average}/20 (aucun écart)."
            )

    return redirect("admin_results")


@login_required
@role_required(["admin"])
def admin_publish_all_results(request):
    if request.method == "POST":
        count = 0

        for assignment in JuryStudent.objects.prefetch_related("evaluations").select_related("student__user"):
            if assignment.evaluations.filter(is_submitted=True).count() == 3:
                result, _ = Result.objects.get_or_create(
                    jury_student=assignment,
                )
                result.calculate_average()  # applique la règle par critère
                result.publish()
                notify(
                    getattr(assignment.student, "user", None),
                    "Résultat de soutenance publié",
                    "Votre note finale est disponible dans votre espace.",
                    "/student-dashboard/",
                    category=Notification.CATEGORY_RESULT,
                )
                count += 1

        messages.success(request, f"{count} résultat(s) publié(s).")

    return redirect("admin_results")


@login_required
@role_required(["admin"])
def admin_grade_sheet(request):
    """Fiche de notes : TOUS les étudiants notés depuis le début (résultat
    calculé, publié ou non), triés par MATRICULE croissant. Vue HTML
    imprimable, export PDF et Excel. En-tête ISGI — Département de l'IUP."""
    header_l1 = "Institut Supérieur de Génie Industriel (ISGI)"
    header_l2 = "Département de l'IUP"

    results = (
        Result.objects.filter(average__isnull=False)
        .select_related("jury_student__student")
    )
    rows = []
    for r in results:
        s = r.jury_student.student
        rows.append({
            "matricule": s.matricule or "",
            "name": s.full_name or "(nom absent)",
            "filiere": s.filiere or "—",
            "average": r.average,
            "mention": mention_for_average(r.average),
            "published": r.is_published,
        })
    rows.sort(key=lambda x: (x["matricule"] or "").upper())

    fmt = (request.GET.get("format") or "").strip()

    if fmt == "xlsx":
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Fiche de notes"
        ws.append([header_l1])
        ws.append([header_l2])
        ws.append([f"Fiche de notes — {len(rows)} étudiant(s) noté(s)"])
        ws.append([])
        ws.append(["Matricule", "Nom & Prénom", "Filière", "Note finale", "Mention", "Statut"])
        for row in rows:
            ws.append([
                row["matricule"], row["name"], row["filiere"],
                float(row["average"]) if row["average"] is not None else "",
                row["mention"], "Publié" if row["published"] else "Non publié",
            ])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="fiche_de_notes.xlsx"'
        wb.save(response)
        return response

    if fmt == "pdf":
        lines = [header_l1, header_l2, "",
                 f"Fiche de notes ({len(rows)} etudiant(s) note(s)) — par matricule", ""]
        for row in rows:
            note = f"{row['average']}" if row["average"] is not None else "—"
            statut = "" if row["published"] else "  [non publie]"
            lines.append(
                f"{row['matricule']}  {row['name']}  ({row['filiere']})  :  "
                f"{note}  ({row['mention']}){statut}"
            )
        if not rows:
            lines.append("Aucun etudiant note pour le moment.")
        return simple_pdf_response("Fiche de notes", lines, "fiche_de_notes.pdf")

    return render(request, "soutenances/admin_grade_sheet.html", {
        "rows": rows,
        "header_l1": header_l1,
        "header_l2": header_l2,
        "total": len(rows),
    })


@login_required
@role_required(["admin"])
def admin_results_by_filiere(request):
    """Rapport des notes par filière (résultats PUBLIÉS uniquement).
    Se remplit au fur et à mesure des publications. Vue HTML imprimable,
    export PDF et Excel. En-tête ISGI — Département de l'IUP."""
    from collections import OrderedDict

    header_l1 = "Institut Supérieur de Génie Industriel (ISGI)"
    header_l2 = "Département de l'IUP"

    results = (
        Result.objects.filter(is_published=True)
        .select_related("jury_student__student")
        .order_by("jury_student__student__filiere", "jury_student__student__matricule")
    )
    groups = OrderedDict()
    for r in results:
        s = r.jury_student.student
        # Regroupement INSENSIBLE À LA CASSE : « FinTech » et « FINTECH » sont la
        # même filière (on normalise en majuscules).
        fil = (s.filiere or "").strip().upper() or "(FILIÈRE NON RENSEIGNÉE)"
        groups.setdefault(fil, []).append({
            "name": s.full_name or "(nom absent)",
            "matricule": s.matricule,
            "average": r.average,
        })
    # Filières triées alphabétiquement ; dans chaque filière : matricule, nom, note.
    groups = OrderedDict(sorted(groups.items(), key=lambda kv: kv[0]))
    for fil in groups:
        groups[fil].sort(key=lambda x: (
            (x["matricule"] or "").upper(),
            (x["name"] or "").upper(),
            x["average"] if x["average"] is not None else Decimal("0"),
        ))

    fmt = (request.GET.get("format") or "").strip()

    if fmt == "xlsx":
        from openpyxl import Workbook
        wb = Workbook()
        first = True
        for fil, rows in groups.items():
            ws = wb.active if first else wb.create_sheet()
            ws.title = (fil[:28] or "Filiere")
            first = False
            ws.append([header_l1])
            ws.append([header_l2])
            ws.append([f"Résultats — Filière {fil}"])
            ws.append([])
            ws.append(["Matricule", "Nom & Prénom", "Note finale"])
            for row in rows:
                ws.append([row["matricule"], row["name"],
                           float(row["average"]) if row["average"] is not None else ""])
        if first:  # aucun résultat
            wb.active.append([header_l1]); wb.active.append([header_l2])
            wb.active.append(["Aucun résultat publié pour le moment."])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="notes_par_filiere.xlsx"'
        wb.save(response)
        return response

    if fmt == "pdf":
        lines = [header_l1, header_l2, "", "Résultats de soutenance par filière", ""]
        for fil, rows in groups.items():
            lines.append(f"— Filière {fil} ({len(rows)}) —")
            for row in rows:
                note = f"{row['average']}" if row["average"] is not None else "—"
                lines.append(f"{row['matricule']}  {row['name']}  :  {note}")
            lines.append("")
        if not groups:
            lines.append("Aucun résultat publié pour le moment.")
        return simple_pdf_response(
            "Notes par filière", lines, "notes_par_filiere.pdf"
        )

    if fmt == "word":
        # En-tête identique à celui de l'AUTORISATION DE SOUTENANCE.
        head_l1 = "Institut Supérieur de Génie Industriel"
        head_l2 = "Département des Formations IUP"
        body = [
            "<html><head><meta charset='utf-8'></head><body>",
            "<div style='text-align:center;'>",
            f"<h2>{head_l1}</h2><p><b>{head_l2}</b></p>",
            "<h3>Résultats de soutenance par filière</h3></div>",
        ]
        if not groups:
            body.append("<p>Aucun résultat publié pour le moment.</p>")
        for fil, rows in groups.items():
            body.append(f"<h4>Filière {fil} — {len(rows)} étudiant(s)</h4>")
            body.append(
                "<table border='1' cellspacing='0' cellpadding='4'>"
                "<tr><th>Matricule</th><th>Nom &amp; Prénom</th><th>Note finale</th></tr>"
            )
            for row in rows:
                note = f"{row['average']}" if row["average"] is not None else "—"
                body.append(
                    f"<tr><td>{row['matricule']}</td><td>{row['name']}</td>"
                    f"<td>{note}</td></tr>"
                )
            body.append("</table><br>")
        body.append("</body></html>")
        resp = HttpResponse("".join(body), content_type="application/msword")
        resp["Content-Disposition"] = 'attachment; filename="notes_par_filiere.doc"'
        return resp

    return render(request, "soutenances/admin_results_by_filiere.html", {
        "groups": groups,
        "header_l1": header_l1,
        "header_l2": header_l2,
        "total": sum(len(v) for v in groups.values()),
    })


@login_required
@role_required(["admin"])
def admin_unlock_evaluation(request, pk):
    evaluation = get_object_or_404(Evaluation, pk=pk)

    if request.method == "POST":
        evaluation.unlock(request.user)

        result = getattr(evaluation.jury_student, "result", None)

        if result:
            result.is_published = False
            result.published_at = None
            result.average = None
            result.note_gap_value = None
            result.has_note_gap_alert = False
            result.save()

        messages.success(request, "Évaluation déverrouillée.")

    return redirect("admin_results")


@login_required
@role_required(["admin"])
def export_juries_pdf(request):
    lines = []

    juries = Jury.objects.prefetch_related(
        "members__professor",
        "students__student",
        "students__student__encadrant",
        "students__president",
    ).order_by(
        "defense_date",
        "name",
    )

    for jury in juries:
        lines.append("JURY :")
        lines.append(f"Nom du jury : {jury.name}")
        lines.append(f"Date de soutenance : {format_date(jury.defense_date)}")
        lines.append(f"Jury validé : {yes_no(jury.is_validated)}")
        lines.append("Membres :")

        for member in jury.members.all():
            lines.append(f"- {member.professor.full_name}")

        lines.append("Étudiants affectés :")

        if jury.students.exists():
            for assignment in jury.students.all():
                president_name = (
                    assignment.president.full_name
                    if assignment.president
                    else "Non défini"
                )

                lines.append(
                    f"- {assignment.student.matricule} | "
                    f"{assignment.student.full_name} | "
                    f"{assignment.student.filiere or '-'} | "
                    f"Encadrant : {assignment.student.encadrant.full_name} | "
                    f"Président : {president_name}"
                )
        else:
            lines.append("- Aucun étudiant affecté")

        lines.append("-----")

    return simple_pdf_response(
        "Liste des jurys",
        lines,
        "liste-jurys.pdf",
    )


@login_required
@role_required(["admin"])
def export_planning_pdf(request):
    lines = []

    schedules = DefenseSchedule.objects.select_related(
        "jury_student__student",
        "jury_student__president",
        "jury_student__jury",
    ).order_by(
        "jury_student__jury__defense_date",
        "start_time",
    )

    for schedule in schedules:
        president_name = (
            schedule.jury_student.president.full_name
            if schedule.jury_student.president
            else "Non défini"
        )

        lines.append(
            f"{format_date(schedule.jury_student.jury.defense_date)} | "
            f"{format_time(schedule.start_time)} - {format_time(schedule.end_time)} | "
            f"{schedule.jury_student.jury.name} | "
            f"{schedule.jury_student.student.matricule} | "
            f"{schedule.jury_student.student.full_name} | "
            f"Président : {president_name}"
        )

    return simple_pdf_response(
        "Planning des soutenances",
        lines,
        "planning-soutenances.pdf",
    )


@login_required
@role_required(["admin"])
def export_results_pdf(request):
    lines = []

    results = Result.objects.select_related(
        "jury_student__student",
        "jury_student__president",
        "jury_student__jury",
    ).order_by(
        "jury_student__student__full_name",
    )

    for result in results:
        status = "publié" if result.is_published else "non publié"
        alert = "oui" if result.has_note_gap_alert else "non"

        lines.append(
            f"{result.jury_student.student.matricule} | "
            f"{result.jury_student.student.full_name} | "
            f"{result.jury_student.jury.name} | "
            f"Moyenne : {decimal_text(result.average)} /20 | "
            f"Écart : {decimal_text(result.note_gap_value)} | "
            f"Alerte : {alert} | "
            f"Publication : {status}"
        )

    return simple_pdf_response(
        "Liste des résultats",
        lines,
        "résultats.pdf",
    )


@login_required
@role_required(["admin"])
def export_student_pv_pdf(request, pk):
    assignment = get_object_or_404(
        JuryStudent.objects.select_related(
            "student",
            "student__user",
            "student__encadrant",
            "president",
            "jury",
            "result",
        ).prefetch_related(
            "jury__members__professor",
            "evaluations__professor",
        ),
        pk=pk,
    )

    student = assignment.student
    jury = assignment.jury
    result = getattr(assignment, "result", None)
    schedule = get_assignment_schedule(assignment)
    pfe_request = PFERequest.objects.filter(student=student).first()

    members = [
        member.professor.full_name
        for member in jury.members.all()
    ]

    president_name = (
        assignment.president.full_name
        if assignment.president
        else "Non défini"
    )

    evaluations = list(
        assignment.evaluations.select_related("professor").order_by(
            "professor__full_name"
        )
    )

    lines = []

    lines.append("INFORMATIONS ÉTUDIANT :")
    lines.append(f"Nom complet : {student.full_name}")
    lines.append(f"Matricule : {student.matricule}")
    lines.append(f"Filière : {student.filiere or '-'}")
    lines.append(f"Téléphone : {student.user.phone_number or '-'}")
    lines.append(f"Encadrant : {student.encadrant.full_name}")
    lines.append(f"Président de soutenance : {president_name}")
    lines.append("")

    lines.append("DEMANDE DE SOUTENANCE :")
    if pfe_request:
        lines.append(f"Statut de la demande : {pfe_request.get_status_display()}")
        lines.append(f"Date de dépôt : {format_datetime(pfe_request.submitted_at)}")
        lines.append(f"Validation encadrant : {format_datetime(pfe_request.professor_reviewed_at)}")
        lines.append(f"Validation département de l'IUP : {format_datetime(pfe_request.admin_reviewed_at)}")

        if pfe_request.professor_comment:
            lines.append(f"Commentaire encadrant : {pfe_request.professor_comment}")

        if pfe_request.admin_comment:
            lines.append(f"Commentaire département de l'IUP : {pfe_request.admin_comment}")
    else:
        lines.append("Aucune demande trouvée.")
    lines.append("")

    lines.append("JURY ET PLANNING :")
    lines.append(f"Jury : {jury.name}")
    lines.append(f"Date de soutenance : {format_date(jury.defense_date)}")

    if schedule:
        lines.append(
            f"Horaire : {format_time(schedule.start_time)} - {format_time(schedule.end_time)}"
        )
    else:
        lines.append("Horaire : non planifié")

    lines.append("Membres du jury :")
    for member_name in members:
        lines.append(f"- {member_name}")
    lines.append("")

    lines.append("ÉVALUATIONS :")
    if evaluations:
        for evaluation in evaluations:
            statut = "envoyée" if evaluation.is_submitted else "brouillon"

            lines.append(f"Professeur : {evaluation.professor.full_name}")
            lines.append(f"  Rapport : {decimal_text(evaluation.rapport_note)} /20 | Coef. 0.30")
            lines.append(f"  Présentation : {decimal_text(evaluation.presentation_note)} /20 | Coef. 0.30")
            lines.append(f"  Réponses aux questions : {decimal_text(evaluation.questions_note)} /20 | Coef. 0.40")
            lines.append(f"  Note finale professeur : {decimal_text(evaluation.final_note)} /20")
            lines.append(f"  Statut : {statut}")
    else:
        lines.append("Aucune évaluation enregistrée.")
    lines.append("")

    lines.append("RÉSULTAT FINAL :")
    if result:
        avgs = compute_criteria_averages(assignment)
        bd = avgs.get("breakdown")
        if bd and bd.get("any_correction"):
            lines.append(
                "Écart >= 3 détecté : note recalculée critère par critère "
                "(membre aberrant écarté)."
            )
            lines.append(
                f"Note d'origine (moyenne des 3 membres) : "
                f"{decimal_text(bd.get('raw_avg_finale'))} /20"
            )
            for crit in bd["criteria"]:
                suffix = ""
                if crit["excluded"] is not None:
                    suffix = f" (note de {crit['excluded'].full_name} écartée, écart {decimal_text(crit['spread'])})"
                lines.append(
                    f"  {crit['label']} retenu : {decimal_text(crit['adjusted'])} /20{suffix}"
                )
        lines.append(f"Moyenne finale : {decimal_text(result.average)} /20")
        lines.append(f"Écart entre notes : {decimal_text(result.note_gap_value)}")
        lines.append(f"Note recalculée (écart >= 3) : {yes_no(result.has_note_gap_alert)}")
        lines.append(f"Résultat publié : {yes_no(result.is_published)}")
        lines.append(f"Date de publication : {format_datetime(result.published_at)}")
    else:
        lines.append("Résultat non calculé.")
    lines.append("")

    lines.append("DÉCISION :")
    if result and result.is_published:
        lines.append("Décision : résultat validé et publié par le département de l'IUP.")
    else:
        lines.append("Décision : en attente de publication par le département de l'IUP.")
    lines.append("")

    lines.append("SIGNATURES :")
    lines.append(f"Président de soutenance ({president_name}) : ______________________________")
    lines.append("Membre du jury : _________________________________")
    lines.append("Membre du jury : _________________________________")
    lines.append("Département de l'IUP : _________________________________")

    return simple_pdf_response(
        f"PV de soutenance - {student.full_name}",
        lines,
        f"pv-{student.matricule}.pdf",
    )


@login_required
@role_required(["admin", "professor"])
def export_evaluation_fiche_pdf(request, pk):
    """Fiche d'Évaluation de Stage de Fin d'études (Word), identique au document
    officiel et pré-remplie avec les moyennes du jury (Rapport ×0,30,
    Présentation ×0,30, Questions ×0,40, Note finale). Accessible au président du
    jury concerné et au chef de département."""
    from .fiche import build_fiche_docx, DOCX_CONTENT_TYPE

    assignment = get_object_or_404(
        JuryStudent.objects.select_related(
            "student", "student__user", "student__encadrant", "jury", "president"
        ).prefetch_related("jury__members__professor", "evaluations"),
        pk=pk,
    )

    # Contrôle d'accès : admin OU président de ce jury.
    is_admin = getattr(request.user, "role", None) == "admin"
    if not is_admin:
        professor = getattr(request.user, "professor_profile", None)
        if not professor or assignment.president_id != professor.id:
            messages.error(request, "Seul le président du jury ou le département peut accéder à cette fiche.")
            return redirect("professor_my_juries")

    avgs = compute_criteria_averages(assignment)

    # La fiche n'est imprimable que lorsque les 3 membres ont tous noté.
    if not avgs["complete"]:
        messages.warning(
            request,
            "La fiche d'évaluation sera disponible une fois les 3 notes saisies "
            f"({avgs['submitted_count']}/{avgs['members_count']} pour le moment)."
        )
        return redirect("admin_results" if is_admin else "professor_president_results")

    document = build_fiche_docx(assignment, avgs)

    filename = f"fiche-evaluation-{assignment.student.matricule}.docx"
    response = HttpResponse(document, content_type=DOCX_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def get_assignment_schedule(assignment):
    try:
        return assignment.schedule
    except Exception:
        return None


def format_date(value):
    if not value:
        return "-"

    return value.strftime("%d/%m/%Y")


def format_time(value):
    if not value:
        return "-"

    return value.strftime("%H:%M")


def format_datetime(value):
    if not value:
        return "-"

    try:
        value = timezone.localtime(value)
    except Exception:
        pass

    return value.strftime("%d/%m/%Y %H:%M")


def decimal_text(value):
    if value is None:
        return "-"

    return str(value).replace(".", ",")


def yes_no(value):
    return "oui" if value else "non"


def get_common_available_slots(encadrants):
    """Real, server-side source of truth: future slots where every given
    encadrant is both declared-available and conflict-free. Used by the
    AJAX hint endpoint and by the enforced quick-create flow."""

    common_slots = None

    for encadrant in encadrants:
        professor_slots = set(build_future_slots_for_professor(encadrant))
        professor_slots = {
            (defense_date, start_time)
            for defense_date, start_time in professor_slots
            if is_professor_available(encadrant, defense_date, start_time)
            and not professor_has_conflict(encadrant, defense_date, start_time)
        }

        if common_slots is None:
            common_slots = professor_slots
        else:
            common_slots &= professor_slots

    return sorted(common_slots or [])


def get_available_professors_at_slot(defense_date, start_time):
    """Real, server-side source of truth: professors actually available
    (declared availability, no conflict) at a given date/time."""

    return [
        professor for professor in ProfessorProfile.objects.order_by("full_name")
        if is_professor_available(professor, defense_date, start_time)
        and not professor_has_conflict(professor, defense_date, start_time)
    ]


@login_required
@role_required(["admin"])
def admin_jury_helper_slots(request):
    """AJAX helper (informational only, used by the free-form manual jury
    creation page): given one or more student ids, return the future slots
    where every selected student's encadrant is available."""

    student_ids = request.GET.getlist("student_id")

    if not student_ids:
        return JsonResponse({"slots": [], "encadrants": []})

    students = StudentProfile.objects.filter(
        id__in=student_ids
    ).select_related("encadrant")

    encadrants = {
        student.encadrant for student in students if student.encadrant_id
    }

    if not encadrants:
        return JsonResponse({"slots": [], "encadrants": []})

    common_slots = get_common_available_slots(encadrants)

    return JsonResponse({
        "slots": [
            {
                "date": defense_date.isoformat(),
                "start_time": start_time.strftime("%H:%M"),
                "label": f"{defense_date.strftime('%d/%m/%Y')} à {start_time.strftime('%H:%M')}",
            }
            for defense_date, start_time in common_slots
        ],
        "encadrants": [
            {"id": encadrant.id, "name": encadrant.full_name}
            for encadrant in encadrants
        ],
    })


@login_required
@role_required(["admin"])
def admin_jury_helper_members(request):
    """AJAX helper (informational only): given a date and a start time,
    return the professors who are actually available at that slot."""

    date_raw = request.GET.get("date")
    start_time_raw = request.GET.get("start_time")

    if not date_raw or not start_time_raw:
        return JsonResponse({"available_professor_ids": [], "all_count": 0})

    try:
        defense_date = datetime.strptime(date_raw, "%Y-%m-%d").date()
        start_time = datetime.strptime(start_time_raw, "%H:%M").time()
    except ValueError:
        return JsonResponse({"error": "Date ou heure invalide."}, status=400)

    available_ids = [
        professor.id for professor in get_available_professors_at_slot(defense_date, start_time)
    ]

    return JsonResponse({
        "available_professor_ids": available_ids,
        "all_count": ProfessorProfile.objects.count(),
    })
