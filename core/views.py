import csv
import io
import re
import tempfile
import unicodedata
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.management import call_command
from django.db import transaction
from django.db.models import Count, Q
from django.shortcuts import redirect, render

from accounts.decorators import role_required
from accounts.models import CustomUser
from documents.models import DocumentTemplate
from professors.models import ProfessorProfile
from students.models import StudentProfile, StudentReference, normalize_matricule
from soutenances.models import (
    Deadline,
    DefenseSchedule,
    Evaluation,
    Jury,
    JuryStudent,
    PFERequest,
    Result,
)
from .forms import ImportPeopleForm, ImportStudentReferencesForm


DEFAULT_IMPORT_PASSWORD = "iup2026"


def dashboard_redirect(request):
    if not request.user.is_authenticated:
        return redirect("login")

    if request.user.role == "admin":
        return redirect("admin_dashboard")

    if request.user.role == "professor":
        return redirect("professor_dashboard")

    if request.user.role == "student":
        return redirect("student_dashboard")

    return redirect("login")


@login_required
@role_required(["admin"])
def admin_dashboard(request):
    stats = {
        "students": StudentProfile.objects.count(),
        "professors": ProfessorProfile.objects.count(),
        "total_requests": PFERequest.objects.count(),
        "pending_professor_requests": PFERequest.objects.filter(
            status=PFERequest.STATUS_PENDING_PROFESSOR
        ).count(),
        "pending_admin_requests": PFERequest.objects.filter(
            status=PFERequest.STATUS_PENDING_ADMIN
        ).count(),
        "accepted_requests": PFERequest.objects.filter(
            status=PFERequest.STATUS_ACCEPTED
        ).count(),
        "refused_requests": PFERequest.objects.filter(
            status__in=[
                PFERequest.STATUS_REFUSED_BY_PROFESSOR,
                PFERequest.STATUS_REFUSED_BY_ADMIN,
            ]
        ).count(),
        "juries": Jury.objects.count(),
        "published_juries": Jury.objects.filter(is_validated=True).count(),
        "draft_juries": Jury.objects.filter(is_validated=False).count(),
        "planned_defenses": DefenseSchedule.objects.count(),
        "published_results": Result.objects.filter(is_published=True).count(),
        "unpublished_results": Result.objects.filter(
            average__isnull=False,
            is_published=False
        ).count(),
        "note_gap_alerts": Result.objects.filter(has_note_gap_alert=True).count(),
        "score_16_plus": Result.objects.filter(average__gte=16).count(),
        "score_14_1599": Result.objects.filter(average__gte=14, average__lt=16).count(),
        "score_12_1399": Result.objects.filter(average__gte=12, average__lt=14).count(),
        "score_10_1199": Result.objects.filter(average__gte=10, average__lt=12).count(),
        "score_below_10": Result.objects.filter(average__lt=10).count(),
    }

    recent_requests = PFERequest.objects.select_related(
        "student",
        "student__encadrant",
    ).order_by("-submitted_at")[:6]

    upcoming_schedules = DefenseSchedule.objects.select_related(
        "jury_student__student",
        "jury_student__jury",
    ).order_by("jury_student__jury__defense_date", "start_time")[:6]

    gap_alerts = Result.objects.filter(
        has_note_gap_alert=True
    ).select_related("jury_student__student", "jury_student__jury")[:5]

    draft_juries = Jury.objects.filter(
        is_validated=False
    ).order_by("-created_at")[:5]

    results_to_publish = Result.objects.filter(
        average__isnull=False,
        is_published=False,
    ).select_related("jury_student__student", "jury_student__jury")[:5]

    return render(request, "core/admin_dashboard.html", {
        "stats": stats,
        "recent_requests": recent_requests,
        "upcoming_schedules": upcoming_schedules,
        "gap_alerts": gap_alerts,
        "draft_juries": draft_juries,
        "results_to_publish": results_to_publish,
    })


@login_required
@role_required(["professor"])
def professor_dashboard(request):
    professor = getattr(request.user, "professor_profile", None)
    if not professor:
        messages.warning(request, "Votre profil professeur n'est pas encore configuré.")
        return render(request, "core/professor_dashboard.html", {
            "professor": None,
        })

    supervised_students = StudentProfile.objects.filter(
        encadrant=professor
    ).select_related("pfe_request")

    jury_students = JuryStudent.objects.filter(
        jury__members__professor=professor,
        jury__is_validated=True,
    ).select_related(
        "student",
        "student__encadrant",
        "jury",
    ).distinct()

    pending_requests_count = PFERequest.objects.filter(
        student__encadrant=professor,
        status=PFERequest.STATUS_PENDING_PROFESSOR,
    ).count()
    submitted_notes = professor.given_evaluations.filter(is_submitted=True).count()
    pending_evaluations = 0
    for assignment in jury_students:
        evaluation = assignment.evaluations.filter(professor=professor).first()
        if evaluation is None or not evaluation.is_submitted:
            pending_evaluations += 1

    availability_blocks = professor.availabilities.order_by("date", "start_time")[:5]
    availability_count = professor.availabilities.count()

    to_start_count = JuryStudent.objects.filter(
        president=professor,
        presentation_started=False,
        jury__is_validated=True,
    ).count()

    # ── Compteurs des étudiants encadrés (base : liste officielle, inclut les
    #    non-inscrits ; le rattachement se fait par le nom de l'encadrant). ──
    official_refs = StudentReference.objects.filter(
        encadrant_name__iexact=professor.full_name
    )
    official_total = official_refs.count()
    registered_mats = {
        normalize_matricule(m)
        for m in StudentProfile.objects.values_list("matricule", flat=True)
    }
    inscrits_count = sum(
        1 for ref in official_refs
        if normalize_matricule(ref.matricule) in registered_mats
    )
    non_inscrits_count = official_total - inscrits_count

    # Statuts (sur les étudiants encadrés inscrits, via le lien encadrant).
    accepted_dept_count = PFERequest.objects.filter(
        student__encadrant=professor,
        status=PFERequest.STATUS_ACCEPTED,
    ).count()
    soutenu_count = (
        Result.objects.filter(
            is_published=True,
            jury_student__student__encadrant=professor,
        )
        .values_list("jury_student__student_id", flat=True)
        .distinct()
        .count()
    )

    # Étudiants où ce prof a été membre de jury ET les a NOTÉS (compte réel :
    # s'il a été remplacé avant de noter certains, ceux-là ne comptent pas).
    jury_graded_count = (
        Evaluation.objects.filter(professor=professor, is_submitted=True)
        .values_list("jury_student__student_id", flat=True)
        .distinct()
        .count()
    )

    return render(request, "core/professor_dashboard.html", {
        "professor": professor,
        "supervised_students_count": supervised_students.count(),
        "pending_requests_count": pending_requests_count,
        "juries_count": Jury.objects.filter(members__professor=professor, is_validated=True).distinct().count(),
        "pending_evaluations": pending_evaluations,
        "submitted_notes": submitted_notes,
        "availability_blocks": availability_blocks,
        "availability_count": availability_count,
        "to_start_count": to_start_count,
        "supervised_students": supervised_students[:6],
        "jury_graded_count": jury_graded_count,
        # Compteurs par statut des étudiants encadrés
        "students_official_total": official_total,
        "students_inscrits_count": inscrits_count,
        "students_non_inscrits_count": non_inscrits_count,
        "students_pending_prof_count": pending_requests_count,
        "students_accepted_dept_count": accepted_dept_count,
        "students_soutenu_count": soutenu_count,
    })


@login_required
@role_required(["student"])
def student_dashboard(request):
    student = getattr(request.user, "student_profile", None)
    pfe_request = None
    jury_assignment = None
    schedule = None
    result = None
    result_breakdown = None

    if student:
        pfe_request = getattr(student, "pfe_request", None)
        jury_assignment = getattr(student, "jury_assignment", None)
        if jury_assignment and not jury_assignment.jury.is_validated:
            jury_assignment = None
        if jury_assignment:
            schedule = getattr(jury_assignment, "schedule", None)
            candidate_result = getattr(jury_assignment, "result", None)
            if candidate_result and candidate_result.is_published:
                result = candidate_result

                # Répartition de la note APRÈS correction d'écart : l'étudiant
                # voit le détail Rapport / Présentation / Questions tel qu'utilisé
                # pour sa note finale (note aberrante écartée par critère si un
                # écart >= 3 a été détecté).
                from soutenances.models import corrected_breakdown

                evaluations = list(
                    jury_assignment.evaluations.filter(is_submitted=True)
                )
                if evaluations:
                    bd = corrected_breakdown(evaluations)
                    result_breakdown = {
                        "rapport": bd["avg_rapport"],
                        "presentation": bd["avg_presentation"],
                        "questions": bd["avg_questions"],
                        "final": bd["avg_finale"],
                        "raw_final": bd["raw_avg_finale"],
                        "corrected": bd["any_correction"],
                    }

    # Dossier incomplet : pièces obligatoires manquantes sur une demande déjà
    # déposée (ex. rapport non joint). L'étudiant pourra la compléter.
    dossier_missing = []
    if pfe_request:
        if not pfe_request.authorization_document:
            dossier_missing.append("l'autorisation de soutenance")
        if not pfe_request.attestation_stage:
            dossier_missing.append("l'attestation de stage")
        if not pfe_request.rapport_stage:
            dossier_missing.append("le rapport de stage")

    deadline = Deadline.objects.filter(
        is_active=True
    ).order_by("-deadline_date").first()

    document_templates = DocumentTemplate.objects.filter(
        is_active=True,
        template_type=DocumentTemplate.TYPE_STUDENT_REQUEST,
    ).order_by("-uploaded_at")
    official_template = document_templates.first()
    deadline_closed = deadline.is_closed() if deadline else False

    return render(request, "core/student_dashboard.html", {
        "student": student,
        "pfe_request": pfe_request,
        "deadline": deadline,
        "deadline_closed": deadline_closed,
        "document_templates": document_templates,
        "official_template": official_template,
        "jury_assignment": jury_assignment,
        "schedule": schedule,
        "result": result,
        "result_breakdown": result_breakdown,
        "dossier_missing": dossier_missing,
    })


@login_required
@role_required(["admin"])
def admin_student_list(request):
    students = StudentProfile.objects.select_related(
        "user",
        "encadrant",
    ).annotate(
        has_request=Count("pfe_request")
    ).order_by("filiere", "full_name")

    return render(request, "core/admin_students.html", {
        "students": students,
    })


def _build_students_overview(filters):
    """Construit la liste unifiée des étudiants (liste officielle + comptes
    inscrits), avec leur statut : inscrit, accepté, soutenu. Applique les
    filtres fournis (filiere, inscrit, accepte, soutenu : '1', '0' ou '')."""
    accepted_ids = set(
        PFERequest.objects.filter(
            status=PFERequest.STATUS_ACCEPTED
        ).values_list("student_id", flat=True)
    )
    request_ids = set(
        PFERequest.objects.values_list("student_id", flat=True)
    )
    defended_ids = set(
        Result.objects.filter(is_published=True).values_list(
            "jury_student__student_id", flat=True
        )
    )

    profiles = StudentProfile.objects.select_related("user", "encadrant")
    profiles_by_mat = {
        normalize_matricule(profile.matricule): profile for profile in profiles
    }

    rows = []
    seen = set()

    # 1) Base = liste officielle
    for reference in StudentReference.objects.all():
        key = normalize_matricule(reference.matricule)
        seen.add(key)
        profile = profiles_by_mat.get(key)
        rows.append(_overview_row(reference, profile, accepted_ids, defended_ids, request_ids))

    # 2) Comptes inscrits absents de la liste officielle (cas limite)
    for key, profile in profiles_by_mat.items():
        if key in seen:
            continue
        rows.append(_overview_row(None, profile, accepted_ids, defended_ids, request_ids))

    # Filtres
    f_filiere = (filters.get("filiere") or "").strip()
    f_inscrit = (filters.get("inscrit") or "").strip()
    f_accepte = (filters.get("accepte") or "").strip()
    f_soutenu = (filters.get("soutenu") or "").strip()
    f_demande = (filters.get("demande") or "").strip()

    def keep(row):
        if f_filiere and row["filiere"] != f_filiere:
            return False
        if f_inscrit == "1" and not row["inscrit"]:
            return False
        if f_inscrit == "0" and row["inscrit"]:
            return False
        if f_accepte == "1" and not row["accepte"]:
            return False
        if f_accepte == "0" and row["accepte"]:
            return False
        if f_soutenu == "1" and not row["soutenu"]:
            return False
        if f_soutenu == "0" and row["soutenu"]:
            return False
        if f_demande == "1" and not row["demande"]:
            return False
        if f_demande == "0" and row["demande"]:
            return False
        return True

    rows = [row for row in rows if keep(row)]
    rows.sort(key=lambda r: (r["filiere"] or "~", r["full_name"].lower()))
    return rows


def _overview_row(reference, profile, accepted_ids, defended_ids, request_ids):
    matricule = (profile.matricule if profile else reference.matricule)
    full_name = (profile.full_name if profile else reference.full_name)
    filiere = (
        (profile.filiere if profile and profile.filiere else None)
        or (reference.filiere if reference else "")
        or ""
    )
    if profile and profile.encadrant_id:
        encadrant = profile.encadrant.full_name
    elif reference and reference.encadrant_name:
        encadrant = reference.encadrant_name
    else:
        encadrant = ""

    return {
        "matricule": matricule,
        "full_name": full_name,
        "filiere": filiere,
        "encadrant": encadrant,
        "entreprise": profile.entreprise if profile else "",
        "telephone": (profile.user.phone_number if profile and profile.user else "") or "",
        "inscrit": profile is not None,
        "demande": bool(profile and profile.id in request_ids),
        "accepte": bool(profile and profile.id in accepted_ids),
        "soutenu": bool(profile and profile.id in defended_ids),
    }


@login_required
@role_required(["admin"])
def admin_students_overview(request):
    rows = _build_students_overview(request.GET)

    # Compteurs récap calculés sur la base de la filière seule (sans les filtres
    # inscrit/accepté/soutenu), afin qu'ils restent significatifs. Chaque
    # compteur suit le sens choisi : « Oui » → positifs, « Non » → négatifs.
    base = _build_students_overview({"filiere": request.GET.get("filiere", "")})

    def directional(rows_, key, flag):
        if flag == "0":
            return sum(1 for r in rows_ if not r[key])
        return sum(1 for r in rows_ if r[key])

    inscrit_flag = request.GET.get("inscrit", "")
    demande_flag = request.GET.get("demande", "")
    accepte_flag = request.GET.get("accepte", "")
    soutenu_flag = request.GET.get("soutenu", "")

    stats = {
        "total": len(base),
        "inscrit_label": "Non inscrits" if inscrit_flag == "0" else "Inscrits",
        "inscrit_count": directional(base, "inscrit", inscrit_flag),
        "demande_label": "Sans demande" if demande_flag == "0" else "Demande déposée",
        "demande_count": directional(base, "demande", demande_flag),
        "accepte_label": "Non acceptés" if accepte_flag == "0" else "Acceptés",
        "accepte_count": directional(base, "accepte", accepte_flag),
        "soutenu_label": "Non soutenus" if soutenu_flag == "0" else "Soutenus",
        "soutenu_count": directional(base, "soutenu", soutenu_flag),
    }

    return render(request, "core/admin_students_overview.html", {
        "rows": rows,
        "stats": stats,
        "filiere_choices": [c for c in StudentProfile.FILIERE_CHOICES if c[0]],
        "selected": {
            "filiere": request.GET.get("filiere", ""),
            "inscrit": request.GET.get("inscrit", ""),
            "demande": request.GET.get("demande", ""),
            "accepte": request.GET.get("accepte", ""),
            "soutenu": request.GET.get("soutenu", ""),
        },
        "querystring": request.GET.urlencode(),
    })


@login_required
@role_required(["admin"])
def admin_students_overview_export(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    rows = _build_students_overview(request.GET)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Etudiants"

    headers = [
        "Matricule", "Nom complet", "Filiere", "Encadrant",
        "Inscrit", "Demande deposee", "Accepte", "Soutenu", "Entreprise", "Telephone",
    ]
    sheet.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F766E")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    def yn(value):
        return "Oui" if value else "Non"

    for row in rows:
        sheet.append([
            row["matricule"], row["full_name"], row["filiere"], row["encadrant"],
            yn(row["inscrit"]), yn(row["demande"]), yn(row["accepte"]), yn(row["soutenu"]),
            row["entreprise"], row["telephone"],
        ])

    widths = [14, 30, 10, 26, 9, 15, 9, 9, 24, 14]
    for i, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + i)].width = width

    from django.http import HttpResponse
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="etudiants.xlsx"'
    workbook.save(response)
    return response


@login_required
@role_required(["admin"])
def admin_check_official_list(request):
    """Vérifie si un étudiant figure sur la liste officielle (StudentReference)
    par matricule ou nom, et affiche ses informations même s'il n'est pas encore
    inscrit sur la plateforme."""
    query = (request.GET.get("q") or "").strip()
    results = []
    searched = bool(query)

    if query:
        references = list(StudentReference.objects.filter(
            Q(matricule__iexact=query) | Q(full_name__icontains=query)
        ).order_by("full_name")[:50])

        # Repli tolérant sur le matricule (espaces insécables / caractères
        # invisibles / casse) si la recherche directe ne donne rien.
        if not references:
            target = normalize_matricule(query)
            if target:
                for candidate in StudentReference.objects.all():
                    if normalize_matricule(candidate.matricule) == target:
                        references = [candidate]
                        break

        for reference in references:
            profile = StudentProfile.objects.filter(
                matricule__iexact=reference.matricule
            ).select_related("user").first()

            results.append({
                "reference": reference,
                "is_registered": profile is not None,
                "profile": profile,
                "phone": profile.user.phone_number if profile and profile.user else None,
            })

    return render(request, "core/admin_check_official.html", {
        "query": query,
        "searched": searched,
        "results": results,
    })


@login_required
@role_required(["admin"])
def admin_rename_professor(request, pk):
    """Corrige le nom d'un professeur (ex. faute d'orthographe corrigée dans la
    liste officielle) afin qu'il corresponde de nouveau."""
    from django.shortcuts import get_object_or_404

    professor = get_object_or_404(ProfessorProfile, pk=pk)
    if request.method == "POST":
        new_name = (request.POST.get("full_name") or "").strip()
        if not new_name:
            messages.error(request, "Le nom ne peut pas être vide.")
        else:
            old_name = professor.full_name
            professor.full_name = new_name
            professor.save(update_fields=["full_name"])
            messages.success(
                request, f"Professeur renommé : « {old_name} » → « {new_name} »."
            )
    return redirect("admin_professor_list")


@login_required
@role_required(["admin"])
def admin_professor_list(request):
    professors = ProfessorProfile.objects.select_related("user").annotate(
        supervised_count=Count("students", distinct=True),
        jury_count=Count("jury_memberships__jury", distinct=True),
        availability_count=Count("availabilities", distinct=True),
    ).order_by("full_name")

    # Nombre d'étudiants réellement NOTÉS par chaque prof (membre de jury qui a
    # évalué) — s'il a été remplacé avant de noter, ceux-là ne comptent pas.
    graded_counts = {}
    for row in (
        Evaluation.objects.filter(is_submitted=True)
        .values("professor_id")
        .annotate(c=Count("jury_student__student_id", distinct=True))
    ):
        graded_counts[row["professor_id"]] = row["c"]
    for p in professors:
        p.graded_count = graded_counts.get(p.id, 0)

    return render(request, "core/admin_professors.html", {
        "professors": professors,
    })


ISGI_L1 = "Institut Supérieur de Génie Industriel (ISGI)"
ISGI_L2 = "Département de l'IUP"


def _word_response(title, body_html, filename):
    """Génère un document ouvrable dans Word (HTML servi en application/msword),
    sans dépendance externe. En-tête ISGI inclus."""
    from django.http import HttpResponse
    html = (
        "<html><head><meta charset='utf-8'></head><body>"
        f"<div style='text-align:center;'><h2>{ISGI_L1}</h2>"
        f"<p><b>{ISGI_L2}</b></p><h3>{title}</h3></div>"
        f"{body_html}</body></html>"
    )
    resp = HttpResponse(html, content_type="application/msword")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _professors_recap_rows():
    """Une ligne par prof : nom, tél, email, nb encadrés, nb de jury (= nombre
    d'étudiants réellement notés par le prof dans un jury)."""
    profs = list(ProfessorProfile.objects.select_related("user").order_by("full_name"))
    graded = {
        r["professor_id"]: r["c"]
        for r in Evaluation.objects.filter(is_submitted=True)
        .values("professor_id")
        .annotate(c=Count("jury_student__student_id", distinct=True))
    }
    enc = {
        r["encadrant_id"]: r["c"]
        for r in StudentProfile.objects.filter(encadrant__isnull=False)
        .values("encadrant_id").annotate(c=Count("id", distinct=True))
    }
    rows = []
    for p in profs:
        rows.append({
            "prof": p,
            "nom": p.full_name,
            "tel": (getattr(p.user, "phone_number", None) if p.user else None) or p.phone or "",
            "email": (p.user.email if p.user else "") or (p.user.username if p.user else ""),
            "encadres": enc.get(p.id, 0),
            "jury": graded.get(p.id, 0),
        })
    return rows


def _professors_details():
    """Chaque prof avec la liste de SES étudiants encadrés (nom, matricule,
    filière). Pour la comptabilité de la direction."""
    profs = list(ProfessorProfile.objects.select_related("user").order_by("full_name"))
    students_by_enc = {}
    for s in StudentProfile.objects.filter(encadrant__isnull=False).order_by("filiere", "full_name"):
        students_by_enc.setdefault(s.encadrant_id, []).append(s)
    result = []
    for p in profs:
        result.append({"prof": p, "students": students_by_enc.get(p.id, [])})
    return result


@login_required
@role_required(["admin"])
def admin_professors_recap(request):
    """Récapitulatif des profs (nom, tél, email, nb encadrés, nb de jury) en
    Excel ou Word."""
    rows = _professors_recap_rows()
    fmt = (request.GET.get("format") or "xlsx").strip()

    if fmt == "word":
        body = ["<table border='1' cellspacing='0' cellpadding='4'>",
                "<tr><th>Nom</th><th>Téléphone</th><th>Email</th>"
                "<th>Étudiants encadrés</th><th>Nombre de jury</th></tr>"]
        for r in rows:
            body.append(
                f"<tr><td>{r['nom']}</td><td>{r['tel']}</td><td>{r['email']}</td>"
                f"<td>{r['encadres']}</td><td>{r['jury']}</td></tr>"
            )
        body.append("</table>")
        return _word_response("Récapitulatif des professeurs", "".join(body),
                              "recap_professeurs.doc")

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Récap professeurs"
    ws.append([ISGI_L1]); ws.append([ISGI_L2]); ws.append(["Récapitulatif des professeurs"]); ws.append([])
    ws.append(["Nom", "Téléphone", "Email", "Étudiants encadrés", "Nombre de jury"])
    for r in rows:
        ws.append([r["nom"], r["tel"], r["email"], r["encadres"], r["jury"]])
    from django.http import HttpResponse
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="recap_professeurs.xlsx"'
    wb.save(resp)
    return resp


@login_required
@role_required(["admin"])
def admin_professors_details(request):
    """Détails : tous les profs et leurs étudiants (nom, matricule, filière),
    en Excel ou Word — pour la comptabilité de la direction."""
    data = _professors_details()
    fmt = (request.GET.get("format") or "xlsx").strip()

    if fmt == "word":
        body = []
        for entry in data:
            p = entry["prof"]
            body.append(f"<h4>{p.full_name} — {len(entry['students'])} étudiant(s)</h4>")
            body.append("<table border='1' cellspacing='0' cellpadding='4'>"
                        "<tr><th>Matricule</th><th>Nom & Prénom</th><th>Filière</th></tr>")
            for s in entry["students"]:
                body.append(f"<tr><td>{s.matricule}</td><td>{s.full_name}</td><td>{s.filiere or ''}</td></tr>")
            body.append("</table><br>")
        return _word_response("Professeurs et leurs étudiants", "".join(body),
                              "professeurs_details.doc")

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Profs et étudiants"
    ws.append([ISGI_L1]); ws.append([ISGI_L2]); ws.append(["Professeurs et leurs étudiants encadrés"]); ws.append([])
    ws.append(["Professeur", "Matricule", "Nom & Prénom étudiant", "Filière"])
    for entry in data:
        p = entry["prof"]
        if not entry["students"]:
            ws.append([p.full_name, "—", "(aucun étudiant encadré)", ""])
            continue
        for s in entry["students"]:
            ws.append([p.full_name, s.matricule, s.full_name, s.filiere or ""])
    from django.http import HttpResponse
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="professeurs_details.xlsx"'
    wb.save(resp)
    return resp


@login_required
@role_required(["professor"])
def professor_my_recap(request):
    """Fiche du prof connecté (Excel/Word) : son récap + ses étudiants."""
    professor = getattr(request.user, "professor_profile", None)
    if not professor:
        return redirect("professor_dashboard")
    encadres = list(
        StudentProfile.objects.filter(encadrant=professor).order_by("filiere", "full_name")
    )
    jury_graded = (
        Evaluation.objects.filter(professor=professor, is_submitted=True)
        .values_list("jury_student__student_id", flat=True).distinct().count()
    )
    tel = (getattr(professor.user, "phone_number", None) if professor.user else None) or professor.phone or ""
    email = (professor.user.email if professor.user else "") or ""
    fmt = (request.GET.get("format") or "xlsx").strip()

    if fmt == "word":
        body = [
            f"<p><b>Nom :</b> {professor.full_name}<br>"
            f"<b>Téléphone :</b> {tel}<br><b>Email :</b> {email}<br>"
            f"<b>Étudiants encadrés :</b> {len(encadres)}<br>"
            f"<b>Nombre de jury (étudiants notés) :</b> {jury_graded}</p>",
            "<h4>Mes étudiants encadrés</h4>",
            "<table border='1' cellspacing='0' cellpadding='4'>"
            "<tr><th>Matricule</th><th>Nom & Prénom</th><th>Filière</th></tr>",
        ]
        for s in encadres:
            body.append(f"<tr><td>{s.matricule}</td><td>{s.full_name}</td><td>{s.filiere or ''}</td></tr>")
        body.append("</table>")
        return _word_response(f"Fiche — {professor.full_name}", "".join(body), "ma_fiche.doc")

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Ma fiche"
    ws.append([ISGI_L1]); ws.append([ISGI_L2]); ws.append([f"Fiche — {professor.full_name}"]); ws.append([])
    ws.append(["Nom", professor.full_name])
    ws.append(["Téléphone", tel])
    ws.append(["Email", email])
    ws.append(["Étudiants encadrés", len(encadres)])
    ws.append(["Nombre de jury (étudiants notés)", jury_graded])
    ws.append([])
    ws.append(["Matricule", "Nom & Prénom", "Filière"])
    for s in encadres:
        ws.append([s.matricule, s.full_name, s.filiere or ""])
    from django.http import HttpResponse
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="ma_fiche.xlsx"'
    wb.save(resp)
    return resp


@login_required
@role_required(["admin"])
def admin_toggle_priority_professor(request, pk):
    """Active/désactive le statut « prof prioritaire » (dispos à utiliser au max
    + président automatique)."""
    professor = ProfessorProfile.objects.filter(pk=pk).first()
    if not professor:
        messages.error(request, "Professeur introuvable.")
        return redirect("admin_professor_list")

    if request.method == "POST":
        professor.is_priority = not professor.is_priority
        professor.save(update_fields=["is_priority"])
        etat = "prioritaire" if professor.is_priority else "non prioritaire"
        messages.success(request, f"{professor.full_name} est désormais {etat}.")

    return redirect("admin_professor_list")


@login_required
@role_required(["admin"])
def admin_reset_student_password(request, pk):
    """Réinitialise le mot de passe d'un étudiant : génère un mot de passe
    temporaire affiché à l'administration, à communiquer à l'étudiant."""
    student = StudentProfile.objects.select_related("user").filter(pk=pk).first()
    if not student:
        messages.error(request, "Étudiant introuvable.")
        return redirect("admin_student_list")

    if request.method != "POST":
        return redirect("admin_student_list")

    if not student.user:
        messages.error(request, "Cet étudiant n'a pas encore de compte.")
        return redirect("admin_student_list")

    from django.utils.crypto import get_random_string
    temp_password = get_random_string(8, "ABCDEFGHJKLMNPQRSTUVWXYZ23456789")
    student.user.set_password(temp_password)
    student.user.save(update_fields=["password"])

    messages.success(
        request,
        f"Mot de passe de {student.full_name} réinitialisé. "
        f"Mot de passe temporaire : {temp_password} — communiquez-le à l'étudiant, "
        f"qui pourra le changer après connexion."
    )
    return redirect("admin_student_list")


@login_required
@role_required(["admin"])
def admin_reset_professor_password(request, pk):
    """Réinitialise le mot de passe d'un professeur (mot de passe temporaire
    affiché à l'administration)."""
    professor = ProfessorProfile.objects.select_related("user").filter(pk=pk).first()
    if not professor:
        messages.error(request, "Professeur introuvable.")
        return redirect("admin_professor_list")

    if request.method != "POST":
        return redirect("admin_professor_list")

    if not professor.user:
        messages.error(request, "Ce professeur n'a pas encore de compte.")
        return redirect("admin_professor_list")

    from django.utils.crypto import get_random_string
    temp_password = get_random_string(8, "ABCDEFGHJKLMNPQRSTUVWXYZ23456789")
    professor.user.set_password(temp_password)
    professor.user.save(update_fields=["password"])

    messages.success(
        request,
        f"Mot de passe de {professor.full_name} réinitialisé. "
        f"Mot de passe temporaire : {temp_password} — communiquez-le au professeur, "
        f"qui pourra le changer après connexion."
    )
    return redirect("admin_professor_list")


@login_required
@role_required(["admin"])
def admin_professor_students(request):
    """Liste des étudiants encadrés par un professeur (source : liste officielle),
    avec nom, matricule, filière et statut d'inscription."""
    name = (request.GET.get("name") or "").strip()
    students = []
    if name:
        refs = StudentReference.objects.filter(
            encadrant_name__iexact=name
        ).order_by("filiere", "full_name")

        registered = {
            (m or "").strip().upper()
            for m in StudentProfile.objects.values_list("matricule", flat=True)
        }
        for ref in refs:
            students.append({
                "full_name": ref.full_name,
                "matricule": ref.matricule,
                "filiere": ref.filiere,
                "inscrit": (ref.matricule or "").strip().upper() in registered,
            })

    return render(request, "core/admin_professor_students.html", {
        "encadrant_name": name,
        "students": students,
        "count": len(students),
    })


@login_required
@role_required(["admin"])
def admin_import_people(request):
    if request.method == "POST":
        form = ImportPeopleForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                rows = read_import_rows(form.cleaned_data["data_file"])
                result = import_people_rows(rows)
            except Exception as exc:
                messages.error(request, f"Import impossible : {exc}")
            else:
                messages.success(
                    request,
                    "Import terminé : "
                    f"{result['created_students']} étudiants créés, "
                    f"{result['updated_students']} mis à jour, "
                    f"{result['created_professors']} professeurs créés."
                )
                if result["errors"]:
                    messages.warning(
                        request,
                        f"{len(result['errors'])} ligne(s) ignorée(s)."
                    )
                return render(request, "core/admin_import.html", {
                    "form": ImportPeopleForm(),
                    "result": result,
                })
    else:
        form = ImportPeopleForm()

    return render(request, "core/admin_import.html", {
        "form": form,
    })


@login_required
@role_required(["admin"])
def admin_import_student_references(request):
    if request.method == "POST":
        form = ImportStudentReferencesForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.cleaned_data["data_file"]
            extension = uploaded_file.name.rsplit(".", 1)[-1].lower()
            temporary_path = None

            try:
                with tempfile.NamedTemporaryFile(
                    suffix=f".{extension}",
                    delete=False,
                ) as temporary_file:
                    for chunk in uploaded_file.chunks():
                        temporary_file.write(chunk)
                    temporary_path = temporary_file.name

                output = io.StringIO()
                call_command(
                    "import_student_references",
                    temporary_path,
                    stdout=output,
                )
            except Exception as exc:
                messages.error(request, f"Import impossible : {exc}")
            else:
                messages.success(
                    request,
                    "Liste officielle importée avec succès. "
                    "Les professeurs manquants ont été créés automatiquement."
                )
                return render(request, "core/admin_import_references.html", {
                    "form": ImportStudentReferencesForm(),
                    "report": output.getvalue(),
                })
            finally:
                if temporary_path:
                    try:
                        import os
                        os.unlink(temporary_path)
                    except Exception:
                        pass
    else:
        form = ImportStudentReferencesForm()

    return render(request, "core/admin_import_references.html", {
        "form": form,
    })


def read_import_rows(uploaded_file):
    extension = uploaded_file.name.rsplit(".", 1)[-1].lower()
    if extension == "csv":
        return read_csv_rows(uploaded_file)
    if extension == "xlsx":
        return read_xlsx_rows(uploaded_file)
    raise ValueError("Format non pris en charge.")


def read_csv_rows(uploaded_file):
    raw = uploaded_file.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("cp1252")

    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"

    reader = csv.reader(io.StringIO(text), dialect)
    return normalize_table_rows(list(reader))


def read_xlsx_rows(uploaded_file):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "Le support Excel necessite openpyxl. Installez requirements.txt."
        ) from exc

    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    return normalize_table_rows(list(sheet.iter_rows(values_only=True)))


def normalize_table_rows(raw_rows):
    rows = [list(row) for row in raw_rows if any(cell for cell in row)]
    if not rows:
        return []

    headers = [_header_key(cell) for cell in rows[0]]
    normalized_rows = []
    for index, row in enumerate(rows[1:], start=2):
        data = {}
        for position, value in enumerate(row):
            if position < len(headers) and headers[position]:
                data[headers[position]] = (str(value).strip() if value is not None else "")
        data["_line"] = index
        normalized_rows.append(data)
    return normalized_rows


def _header_key(value):
    aliases = {
        "matricule": "matricule",
        "nomcomplet": "full_name",
        "nometprenom": "full_name",
        "fullname": "full_name",
        "etudiant": "full_name",
        "filiere": "filiere",
        "encadrant": "encadrant",
        "superviseur": "encadrant",
    }
    key = _ascii_slug(value).replace("_", "")
    return aliases.get(key, key)


def _ascii_slug(value):
    value = unicodedata.normalize("NFKD", str(value or ""))
    value = value.encode("ascii", "ignore").decode("ascii").lower()
    value = re.sub(r"[^a-z0-9]+", "_", value).strip("_")
    return value


def normalize_filiere(value):
    raw = str(value or "").strip().upper()
    compact = re.sub(r"[^A-Z0-9]+", "", raw)
    mapping = {
        "FINTECH": "FINTECH",
        "FIN": "FINTECH",
        "DS": "DS",
        "DATASCIENCE": "DS",
        "MAN": "MAN",
        "MANAGEMENT": "MAN",
        "LGTR": "LGTR",
        "RXTL": "RXTL",
        "MAEF": "MAEF",
    }
    return mapping.get(compact)


@transaction.atomic
def import_people_rows(rows):
    result = {
        "created_students": 0,
        "updated_students": 0,
        "created_professors": 0,
        "errors": [],
    }

    for row in rows:
        matricule = row.get("matricule", "").strip()
        full_name = row.get("full_name", "").strip()
        filiere = normalize_filiere(row.get("filiere", ""))
        encadrant_name = row.get("encadrant", "").strip()

        if not all([matricule, full_name, filiere, encadrant_name]):
            result["errors"].append({
                "line": row.get("_line"),
                "message": "matricule, nom complet, filière ou encadrant manquant",
            })
            continue

        professor, professor_created = get_or_create_professor(encadrant_name)
        if professor_created:
            result["created_professors"] += 1

        user, _ = get_or_create_user(
            username_base=matricule,
            role=CustomUser.ROLE_STUDENT,
            full_name=full_name,
        )

        student, created = StudentProfile.objects.get_or_create(
            matricule=matricule,
            defaults={
                "user": user,
                "full_name": full_name,
                "filiere": filiere,
                "encadrant": professor,
            }
        )

        if created:
            result["created_students"] += 1
        else:
            student.user = student.user or user
            student.full_name = full_name
            student.filiere = filiere
            student.encadrant = professor
            student.save()
            result["updated_students"] += 1

    return result


def get_or_create_professor(full_name):
    professor = ProfessorProfile.objects.filter(full_name__iexact=full_name).first()
    if professor:
        return professor, False

    return ProfessorProfile.objects.create(full_name=full_name), True


def get_or_create_user(username_base, role, full_name):
    username = unique_username(username_base)
    user = CustomUser.objects.filter(username=username_base).first()
    if user:
        user.role = role
        user.save(update_fields=["role"])
        return user, False

    user = CustomUser(username=username, role=role, is_active=True)
    parts = full_name.split(" ", 1)
    user.first_name = parts[0]
    user.last_name = parts[1] if len(parts) > 1 else ""
    user.set_password(DEFAULT_IMPORT_PASSWORD)
    user.save()
    return user, True


def unique_username(username_base):
    base = _ascii_slug(username_base) or "user"
    base = base[:145]
    candidate = base
    suffix = 1
    while CustomUser.objects.filter(username=candidate).exists():
        suffix += 1
        candidate = f"{base}_{suffix}"[:150]
    return candidate


# ── Notifications in-app ──────────────────────────────────────────────────────

@login_required
def notifications_feed(request):
    """Sondage JSON : nombre de non-lus + dernières notifications (pour la cloche
    et le déclenchement de l'alerte sonore côté client)."""
    from django.http import JsonResponse
    from .models import Notification

    queryset = Notification.objects.filter(recipient=request.user)
    unread = queryset.filter(is_read=False).count()
    items = []
    for notification in queryset[:12]:
        items.append({
            "id": notification.id,
            "title": notification.title,
            "message": notification.message,
            "url": notification.url,
            "category": notification.category,
            "is_read": notification.is_read,
            "created_at": notification.created_at.strftime("%d/%m/%Y %H:%M"),
        })
    latest_id = items[0]["id"] if items else 0
    return JsonResponse({"unread": unread, "latest_id": latest_id, "items": items})


@login_required
def notification_open(request, pk):
    """Marque une notification comme lue puis redirige vers sa page cible."""
    from django.shortcuts import get_object_or_404
    from .models import Notification

    notification = get_object_or_404(Notification, pk=pk, recipient=request.user)
    if not notification.is_read:
        notification.is_read = True
        notification.save(update_fields=["is_read"])
    return redirect(notification.url or "dashboard")


@login_required
def notifications_mark_all_read(request):
    """Marque toutes les notifications de l'utilisateur comme lues."""
    from .models import Notification

    if request.method == "POST":
        Notification.objects.filter(
            recipient=request.user, is_read=False
        ).update(is_read=True)
    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER") or "dashboard"
    return redirect(next_url)
