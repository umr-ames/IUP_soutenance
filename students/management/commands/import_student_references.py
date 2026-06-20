import csv
import os
import re
import unicodedata
from collections import Counter, defaultdict

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from accounts.models import CustomUser
from professors.models import ProfessorAvailability, ProfessorProfile
from soutenances.models import Evaluation, JuryMember, JuryStudent, Note, PFERequest
from students.models import StudentProfile, StudentReference


COLUMN_ALIASES = {
    "matricule": "matricule",
    "full_name": "full_name",
    "nom complet": "full_name",
    "nom": "full_name",
    "filiere": "filiere",
    "specialite": "filiere",
    "encadrant_name": "encadrant_name",
    "encadrant": "encadrant_name",
}

HEADER_KEYWORDS = set(COLUMN_ALIASES)


def normalize_text(value):
    value = re.sub(r"\s+", " ", str(value or "").strip().lower())
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def normalize_person_name(name):
    return normalize_text(name)


class Command(BaseCommand):
    help = (
        "Importe la liste officielle des etudiants (matricule, full_name, filiere, "
        "encadrant_name) depuis un fichier CSV ou XLSX vers StudentReference. "
        "Pour les fichiers XLSX, la ligne d'en-tete est detectee automatiquement "
        "(des lignes de titre peuvent preceder l'en-tete). "
        "Colonnes acceptees (insensibles a la casse et aux accents) : "
        "matricule/Matricule, full_name/'Nom complet', filiere/Filiere, "
        "encadrant_name/Encadrant."
    )

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)

    def build_canonical_encadrants(self, rows):
        """Construit, pour chaque nom d'encadrant normalise, l'orthographe
        canonique a conserver : la variante la plus frequente du fichier, et a
        egalite celle qui comporte une majuscule (orthographe propre)."""
        counts = Counter()
        for _line_number, raw_row in rows:
            name = (raw_row.get("encadrant_name") or "").strip()
            if name:
                counts[name] += 1

        groups = defaultdict(list)
        for name, count in counts.items():
            groups[normalize_person_name(name)].append((name, count))

        canonical = {}
        for key, variants in groups.items():
            canonical[key] = sorted(
                variants,
                key=lambda v: (v[1], any(c.isupper() for c in v[0])),
            )[-1][0]
        return canonical

    @transaction.atomic
    def handle(self, *args, **options):
        file_path = options["file_path"]
        extension = os.path.splitext(file_path)[1].lower()

        if extension == ".xlsx":
            rows = self.read_xlsx_rows(file_path)
        elif extension == ".csv":
            rows = self.read_csv_rows(file_path)
        else:
            raise CommandError(
                f"Format non supporte : '{extension}'. Utilisez un fichier .csv ou .xlsx."
            )

        rows = list(rows)

        # Fusion des variantes d'orthographe d'un meme encadrant (casse, espaces,
        # accents) vers une orthographe canonique : evite de compter un meme
        # professeur plusieurs fois (ex. "Yahya marega" / "Yahya Marega").
        canonical_encadrant = self.build_canonical_encadrants(rows)

        old_matricules = set(
            StudentReference.objects.values_list("matricule", flat=True)
        )
        old_encadrant_names = set(
            name
            for name in StudentReference.objects.values_list(
                "encadrant_name", flat=True
            ).distinct()
            if name
        )

        total_rows = 0
        created = 0
        updated = 0
        ignored = 0
        errors = []
        seen_matricules = {}
        duplicate_matricules = []
        empty_name = []
        empty_filiere = []
        empty_encadrant = []
        new_encadrant_names = set()
        new_matricules = set()

        for line_number, raw_row in rows:
            total_rows += 1

            if not any((value or "").strip() for value in raw_row.values()):
                ignored += 1
                continue

            matricule = (raw_row.get("matricule") or "").strip()

            if not matricule:
                ignored += 1
                continue

            full_name = (raw_row.get("full_name") or "").strip()
            filiere = (raw_row.get("filiere") or "").strip()
            encadrant_name = (raw_row.get("encadrant_name") or "").strip()

            if matricule in seen_matricules:
                duplicate_matricules.append(matricule)
            seen_matricules[matricule] = line_number

            if not full_name:
                empty_name.append(matricule)

            if not filiere:
                empty_filiere.append(matricule)

            if not encadrant_name:
                empty_encadrant.append(matricule)
            else:
                encadrant_name = canonical_encadrant.get(
                    normalize_person_name(encadrant_name), encadrant_name
                )
                new_encadrant_names.add(encadrant_name)

            new_matricules.add(matricule)

            try:
                _, was_created = StudentReference.objects.update_or_create(
                    matricule=matricule,
                    defaults={
                        "full_name": full_name,
                        "filiere": filiere,
                        "encadrant_name": encadrant_name,
                    },
                )
            except Exception as exc:
                errors.append(f"Ligne {line_number} (matricule={matricule}) : {exc}")
                continue

            if was_created:
                created += 1
            else:
                updated += 1

        existing_professor_names = {
            normalize_person_name(name)
            for name in ProfessorProfile.objects.values_list("full_name", flat=True)
        }

        professors_created = []
        professors_existing = 0

        for encadrant_name in sorted(new_encadrant_names):
            if normalize_person_name(encadrant_name) in existing_professor_names:
                professors_existing += 1
                continue

            ProfessorProfile.objects.create(full_name=encadrant_name)
            professors_created.append(encadrant_name)
            existing_professor_names.add(normalize_person_name(encadrant_name))

        missing_matricules = sorted(old_matricules - new_matricules)
        missing_linked = []
        missing_unlinked = []

        for matricule in missing_matricules:
            is_linked = (
                StudentProfile.objects.filter(matricule=matricule).exists()
                or PFERequest.objects.filter(student__matricule=matricule).exists()
            )

            if is_linked:
                missing_linked.append(matricule)
            else:
                missing_unlinked.append(matricule)

        deleted_missing_unlinked, _ = StudentReference.objects.filter(
            matricule__in=missing_unlinked
        ).delete()

        missing_encadrants = sorted(old_encadrant_names - new_encadrant_names)
        professor_cleanup = self.sync_professors_with_latest_list(new_encadrant_names)

        total_in_db = StudentReference.objects.count()
        distinct_encadrants_in_db = StudentReference.objects.exclude(
            encadrant_name=""
        ).values_list("encadrant_name", flat=True).distinct().count()

        self.stdout.write(self.style.SUCCESS("Import termine."))
        self.stdout.write(f"Lignes lues dans le fichier : {total_rows}")
        self.stdout.write(f"StudentReference crees : {created}")
        self.stdout.write(f"StudentReference mis a jour : {updated}")
        self.stdout.write(f"Lignes ignorees (vides ou sans matricule) : {ignored}")
        self.stdout.write(f"Erreurs : {len(errors)}")
        self.stdout.write(f"Matricules dupliques dans le fichier : {len(duplicate_matricules)}")
        if duplicate_matricules:
            self.stdout.write("  -> " + ", ".join(duplicate_matricules))
        self.stdout.write(f"Matricules avec nom vide : {len(empty_name)}")
        if empty_name:
            self.stdout.write("  -> " + ", ".join(empty_name))
        self.stdout.write(f"Matricules avec filiere vide : {len(empty_filiere)}")
        if empty_filiere:
            self.stdout.write("  -> " + ", ".join(empty_filiere))
        self.stdout.write(f"Matricules avec encadrant vide : {len(empty_encadrant)}")
        if empty_encadrant:
            self.stdout.write("  -> " + ", ".join(empty_encadrant))

        self.stdout.write("")
        self.stdout.write(f"Total StudentReference en base apres import : {total_in_db}")
        self.stdout.write(f"Encadrants distincts en base : {distinct_encadrants_in_db}")
        self.stdout.write(f"ProfessorProfile crees : {len(professors_created)}")
        if professors_created:
            self.stdout.write("  -> " + ", ".join(professors_created))
        self.stdout.write(f"ProfessorProfile deja existants (reutilises) : {professors_existing}")

        self.stdout.write("")
        self.stdout.write(
            "References presentes avant l'import mais absentes du nouveau fichier : "
            f"{len(missing_matricules)}"
        )
        self.stdout.write(
            "  -> liees a un compte/demande (NON supprimees, signalees) : "
            f"{len(missing_linked)}"
        )
        if missing_linked:
            self.stdout.write("     " + ", ".join(missing_linked))
        self.stdout.write(
            "  -> non liees, supprimees automatiquement : "
            f"{deleted_missing_unlinked}"
        )
        if missing_unlinked:
            self.stdout.write("     " + ", ".join(missing_unlinked))

        self.stdout.write(
            "Anciens encadrants absents de la nouvelle liste : "
            f"{len(missing_encadrants)}"
        )
        if missing_encadrants:
            self.stdout.write("  -> " + ", ".join(missing_encadrants))

        self.stdout.write(
            "ProfessorProfile absents de la nouvelle liste supprimes : "
            f"{len(professor_cleanup['removed'])}"
        )
        if professor_cleanup["removed"]:
            self.stdout.write("  -> " + ", ".join(professor_cleanup["removed"]))
        self.stdout.write(
            "Comptes professeurs supprimes avec ces anciens encadrants : "
            f"{professor_cleanup['removed_users']}"
        )
        self.stdout.write(
            "ProfessorProfile hors nouvelle liste mais conserves car lies a des donnees actives : "
            f"{len(professor_cleanup['blocked'])}"
        )
        for entry in professor_cleanup["blocked"]:
            self.stdout.write(
                f"  -> {entry['name']} : "
                f"{entry['jurys']} jury(s), "
                f"{entry['presidences']} presidence(s), "
                f"{entry['evaluations']} evaluation(s), "
                f"{entry['notes']} note(s), "
                f"{entry['students']} etudiant(s) encadre(s)"
            )

        for error in errors:
            self.stdout.write(self.style.ERROR(error))

    def sync_professors_with_latest_list(self, official_encadrant_names):
        official_names = {
            normalize_person_name(name)
            for name in official_encadrant_names
            if name and name.strip()
        }
        removed = []
        blocked = []
        removed_users = 0

        for professor in ProfessorProfile.objects.select_related("user").all():
            if normalize_person_name(professor.full_name) in official_names:
                continue

            links = {
                "jurys": JuryMember.objects.filter(professor=professor).count(),
                "presidences": JuryStudent.objects.filter(president=professor).count(),
                "evaluations": Evaluation.objects.filter(professor=professor).count(),
                "notes": Note.objects.filter(professor=professor).count(),
                "students": StudentProfile.objects.filter(encadrant=professor).count(),
            }

            if any(links.values()):
                blocked.append({"name": professor.full_name, **links})
                continue

            user_id = professor.user_id
            ProfessorAvailability.objects.filter(professor=professor).delete()
            removed.append(professor.full_name)
            professor.delete()

            if user_id:
                user = CustomUser.objects.filter(pk=user_id)
                if user.exists():
                    user.delete()
                    removed_users += 1

        return {
            "removed": removed,
            "blocked": blocked,
            "removed_users": removed_users,
        }

    def normalize_header(self, value):
        return normalize_text(value)

    def read_csv_rows(self, csv_path):
        try:
            file = open(csv_path, encoding="utf-8-sig")
        except OSError as exc:
            raise CommandError(f"Impossible d'ouvrir le fichier : {exc}")

        with file:
            reader = csv.DictReader(file)

            if not reader.fieldnames:
                raise CommandError("Le fichier CSV est vide ou sans en-tete.")

            normalized_fieldnames = {
                name: COLUMN_ALIASES.get(self.normalize_header(name))
                for name in reader.fieldnames
            }

            rows = []
            for line_number, raw_row in enumerate(reader, start=2):
                row = {}
                for original_name, value in raw_row.items():
                    target_field = normalized_fieldnames.get(original_name)
                    if target_field:
                        row[target_field] = (value or "").strip()
                rows.append((line_number, row))

            return rows

    def read_xlsx_rows(self, xlsx_path):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError(
                "La bibliotheque 'openpyxl' est requise pour lire les fichiers .xlsx. "
                f"Erreur : {exc}"
            )

        try:
            workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
        except Exception as exc:
            raise CommandError(f"Impossible d'ouvrir le fichier Excel : {exc}")

        worksheet = workbook.active

        header_row_index = None
        header_map = {}

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1
        ):
            normalized_cells = [self.normalize_header(cell) for cell in row]
            matches = sum(1 for cell in normalized_cells if cell in HEADER_KEYWORDS)

            if matches >= 3:
                header_row_index = row_index
                for column_index, cell in enumerate(row):
                    target_field = COLUMN_ALIASES.get(self.normalize_header(cell))
                    if target_field:
                        header_map[column_index] = target_field
                break

        if header_row_index is None:
            raise CommandError(
                "Impossible de detecter la ligne d'en-tete (Matricule, Nom complet, "
                "Filiere, Encadrant) dans les 20 premieres lignes du fichier."
            )

        rows = []

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=header_row_index + 1, values_only=True),
            start=header_row_index + 1,
        ):
            parsed = {}
            for column_index, value in enumerate(row):
                target_field = header_map.get(column_index)
                if target_field:
                    parsed[target_field] = str(value).strip() if value is not None else ""
            rows.append((row_index, parsed))

        return rows
